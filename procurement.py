from datetime import date, datetime
import io

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from audit import record_event
from ui import page_header
from approval_workflow import needs_approval, submit_approval_request


def ensure_procurement_schema(conn):
    cursor = conn.cursor()
    try:
        cursor.execute("""CREATE TABLE IF NOT EXISTS procurement_bookings (
            id BIGSERIAL PRIMARY KEY, booking_number TEXT UNIQUE NOT NULL,
            supplier_id INTEGER NOT NULL REFERENCES suppliers(id), product_id INTEGER NOT NULL REFERENCES products(id),
            booking_date DATE NOT NULL, valid_from DATE, valid_to DATE, booked_liters REAL NOT NULL CHECK(booked_liters>0),
            unit_price REAL NOT NULL CHECK(unit_price>=0), payment_terms TEXT NOT NULL,
            transport_responsibility TEXT NOT NULL, status TEXT NOT NULL DEFAULT 'OPEN', notes TEXT,
            created_by TEXT NOT NULL, created_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            closed_by TEXT, closed_at TIMESTAMPTZ,
            CHECK(status IN ('DRAFT','OPEN','PARTIALLY_USED','COMPLETED','CANCELLED','EXPIRED'))
        )""")
        cursor.execute("""CREATE TABLE IF NOT EXISTS procurement_releases (
            id BIGSERIAL PRIMARY KEY, booking_id BIGINT NOT NULL REFERENCES procurement_bookings(id),
            release_number TEXT UNIQUE NOT NULL, release_date DATE NOT NULL, planned_delivery_date DATE,
            released_liters REAL NOT NULL CHECK(released_liters>0), status TEXT NOT NULL DEFAULT 'OPEN',
            notes TEXT, created_by TEXT NOT NULL, created_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            CHECK(status IN ('OPEN','PARTIALLY_RECEIVED','RECEIVED','CANCELLED'))
        )""")
        cursor.execute("""CREATE TABLE IF NOT EXISTS supplier_claims (
            id BIGSERIAL PRIMARY KEY, tank_transaction_id BIGINT UNIQUE REFERENCES tank_transactions(id),
            booking_id BIGINT REFERENCES procurement_bookings(id), supplier_id INTEGER NOT NULL REFERENCES suppliers(id),
            claim_type TEXT NOT NULL, claim_liters REAL NOT NULL, unit_price REAL NOT NULL DEFAULT 0,
            claim_amount REAL NOT NULL DEFAULT 0, status TEXT NOT NULL DEFAULT 'OPEN', credit_note_number TEXT,
            credit_note_date DATE, notes TEXT, created_by TEXT NOT NULL, created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
            resolved_by TEXT, resolved_at TIMESTAMPTZ,
            CHECK(status IN ('OPEN','SUBMITTED','ACKNOWLEDGED','CREDIT_NOTE_RECEIVED','CLOSED','REJECTED'))
        )""")
        for statement in (
            "ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS booking_id BIGINT REFERENCES procurement_bookings(id)",
            "ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS booking_release_id BIGINT REFERENCES procurement_releases(id)",
            "ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS purchase_type TEXT",
            "ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS unit_price REAL",
            "ALTER TABLE procurement_releases ADD COLUMN IF NOT EXISTS destination_depot_id INTEGER REFERENCES depots(id)",
            "ALTER TABLE procurement_releases ADD COLUMN IF NOT EXISTS destination_tank_id INTEGER REFERENCES storage_tanks(id)",
        ):
            cursor.execute(statement)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_booking_supplier_status ON procurement_bookings(supplier_id,status)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_release_booking_status ON procurement_releases(booking_id,status)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_claim_status ON supplier_claims(status,created_at DESC)")
        conn.commit()
    except Exception:
        conn.rollback()
        raise


def booking_options(conn, supplier_id=None):
    sql = """SELECT b.id,b.booking_number,s.name AS supplier,p.name AS product,b.product_id,b.booked_liters,b.unit_price,
        b.payment_terms,b.transport_responsibility,b.status,
        COALESCE(SUM(CASE WHEN tx.type='IN' THEN tx.accepted_liters ELSE 0 END),0) AS received_liters
        FROM procurement_bookings b JOIN suppliers s ON s.id=b.supplier_id JOIN products p ON p.id=b.product_id
        LEFT JOIN tank_transactions tx ON tx.booking_id=b.id
        WHERE b.status IN ('OPEN','PARTIALLY_USED')"""
    params = []
    if supplier_id:
        sql += " AND b.supplier_id=%s"
        params.append(supplier_id)
    sql += " GROUP BY b.id,s.name,p.name ORDER BY b.booking_date DESC,b.id DESC"
    data = pd.read_sql_query(sql, conn, params=params)
    if not data.empty:
        data["remaining_liters"] = data["booked_liters"] - data["received_liters"]
        data = data[data["remaining_liters"] > 0.005]
        data["label"] = data.apply(lambda r: f"{r['booking_number']} · {r['supplier']} · {r['remaining_liters']:,.2f} L remaining", axis=1)
    return data


def release_options(conn, booking_id):
    return pd.read_sql_query("""SELECT r.id,r.release_number,r.released_liters,r.status,
        COALESCE(SUM(tx.accepted_liters),0) AS received_liters
        FROM procurement_releases r LEFT JOIN tank_transactions tx ON tx.booking_release_id=r.id
        WHERE r.booking_id=%s AND r.status IN ('OPEN','PARTIALLY_RECEIVED')
        GROUP BY r.id ORDER BY r.release_date,r.id""", conn, params=[booking_id])


def refresh_booking_status(conn, booking_id, release_id=None):
    cursor = conn.cursor()
    cursor.execute("""SELECT b.booked_liters,COALESCE(SUM(tx.accepted_liters),0)
        FROM procurement_bookings b LEFT JOIN tank_transactions tx ON tx.booking_id=b.id
        WHERE b.id=%s GROUP BY b.id""", (booking_id,))
    row = cursor.fetchone()
    if row:
        received = float(row[1] or 0)
        status = "COMPLETED" if received >= float(row[0]) - 0.005 else ("PARTIALLY_USED" if received > 0 else "OPEN")
        cursor.execute("UPDATE procurement_bookings SET status=%s WHERE id=%s AND status<>'CANCELLED'", (status, booking_id))
    if release_id:
        cursor.execute("""SELECT r.released_liters,COALESCE(SUM(tx.accepted_liters),0)
            FROM procurement_releases r LEFT JOIN tank_transactions tx ON tx.booking_release_id=r.id
            WHERE r.id=%s GROUP BY r.id""", (release_id,))
        row = cursor.fetchone()
        if row:
            received = float(row[1] or 0)
            status = "RECEIVED" if received >= float(row[0]) - 0.005 else ("PARTIALLY_RECEIVED" if received > 0 else "OPEN")
            cursor.execute("UPDATE procurement_releases SET status=%s WHERE id=%s AND status<>'CANCELLED'", (status, release_id))


def create_variance_claim(conn, tank_transaction_id, booking_id, supplier_id, variance_liters, unit_price, user):
    if variance_liters >= -0.005:
        return None
    cursor = conn.cursor()
    shortage = abs(float(variance_liters))
    cursor.execute("""INSERT INTO supplier_claims(tank_transaction_id,booking_id,supplier_id,claim_type,claim_liters,
        unit_price,claim_amount,notes,created_by) VALUES (%s,%s,%s,'SHORT_RECEIPT',%s,%s,%s,%s,%s)
        ON CONFLICT(tank_transaction_id) DO NOTHING RETURNING id""",
        (tank_transaction_id, booking_id, supplier_id, shortage, unit_price or 0, shortage * float(unit_price or 0),
         "Automatically created from supplier receipt variance.", user))
    row = cursor.fetchone()
    return row[0] if row else None


def _report(bookings, releases, claims, profile):
    wb = Workbook(); ws = wb.active; ws.title = "Booking Summary"
    sheets = [(ws, bookings, "Supplier Bookings"), (wb.create_sheet("Releases"), releases, "Booking Releases"),
              (wb.create_sheet("Claims"), claims, "Supplier Claims")]
    primary = str(profile.get("primary_color", "#8C1C1C")).replace("#", "")
    secondary = str(profile.get("secondary_color", "#172033")).replace("#", "")
    company = profile.get("company_name", "FILLIT")
    for index, (sheet, frame, title) in enumerate(sheets):
        sheet.sheet_view.showGridLines = False
        end = max(len(frame.columns), 8); end_letter = get_column_letter(end)
        sheet.merge_cells(f"A1:{end_letter}2"); sheet["A1"] = f"{company} | {title}"
        sheet["A1"].fill = PatternFill("solid", fgColor=secondary); sheet["A1"].font = Font(size=20,bold=True,color="FFFFFF")
        sheet["A1"].alignment = Alignment(vertical="center")
        sheet.merge_cells(f"A3:{end_letter}3"); sheet["A3"] = f"Generated {datetime.now():%d %b %Y, %I:%M %p} (Dubai)"
        sheet["A3"].fill = PatternFill("solid", fgColor=primary); sheet["A3"].font = Font(color="FFFFFF")
        if frame.empty:
            sheet["A5"] = "No records available"
            continue
        headers = [str(c).replace("_", " ").title() for c in frame.columns]
        for col, value in enumerate(headers, 1):
            cell=sheet.cell(5,col,value); cell.fill=PatternFill("solid",fgColor=primary); cell.font=Font(bold=True,color="FFFFFF")
        for row_idx, values in enumerate(frame.itertuples(index=False,name=None), 6):
            for col_idx, value in enumerate(values, 1):
                if isinstance(value, pd.Timestamp): value=value.to_pydatetime().replace(tzinfo=None)
                sheet.cell(row_idx,col_idx,value)
        last=5+len(frame); table=Table(displayName=f"ProcurementTable{index+1}",ref=f"A5:{get_column_letter(len(headers))}{last}")
        table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True,showColumnStripes=False)
        sheet.add_table(table); sheet.freeze_panes="A6"; sheet.auto_filter.ref=table.ref
        for col in range(1,len(headers)+1): sheet.column_dimensions[get_column_letter(col)].width=min(max(len(headers[col-1])+3,14),28)
        sheet.conditional_formatting.add(f"A6:{get_column_letter(len(headers))}{last}",FormulaRule(formula=[f'ISNUMBER(SEARCH("OPEN",${get_column_letter(len(headers))}6))'],fill=PatternFill("solid",fgColor="FEF0C7")))
    output=io.BytesIO(); wb.save(output); return output.getvalue()


def render_procurement(conn):
    ensure_procurement_schema(conn)
    page_header("Supplier Bookings & Procurement", "Control advance bookings, releases, receipt consumption and supplier claims.")
    user=st.session_state.get("user","System")
    suppliers=pd.read_sql_query("SELECT id,name FROM suppliers WHERE COALESCE(status,'ACTIVE')='ACTIVE' ORDER BY name",conn)
    products=pd.read_sql_query("SELECT id,name FROM products WHERE active=TRUE ORDER BY name",conn)
    supplier_map=dict(zip(suppliers["name"],suppliers["id"])); product_map=dict(zip(products["name"],products["id"]))
    tab_overview,tab_booking,tab_release,tab_claims,tab_control,tab_report=st.tabs(["Overview","New booking","Create release","Supplier claims","Cancellation control","Reports"])
    with tab_booking:
        with st.form("new_procurement_booking"):
            a,b=st.columns(2); supplier=a.selectbox("Supplier",list(supplier_map)); product=b.selectbox("Product",list(product_map))
            c,d,e=st.columns(3); number=c.text_input("Booking / contract number"); booking_date=d.date_input("Booking date",date.today()); valid_to=e.date_input("Valid until",date.today())
            f,g=st.columns(2); liters=f.number_input("Booked quantity",min_value=0.0); price=g.number_input("Unit price",min_value=0.0,format="%.4f")
            h,i=st.columns(2); terms=h.selectbox("Payment terms",["Advance paid","Credit 7 days","Credit 15 days","Credit 30 days","Cash on delivery","Other"]); transport=i.selectbox("Transport responsibility",["Supplier delivery","Company collection","Third-party transporter"])
            notes=st.text_area("Booking notes"); submit=st.form_submit_button("Create supplier booking",type="primary")
        if submit:
            if not number.strip() or liters<=0: st.error("Booking number and a quantity greater than zero are required.")
            else:
                try:
                    booking_value=float(liters)*float(price)
                    if needs_approval(conn,"BOOKING_VALUE",booking_value):
                        check=conn.cursor(); check.execute("""SELECT id FROM approval_requests
                            WHERE request_kind='SUPPLIER_BOOKING' AND status='PENDING'
                            AND payload->>'booking_number'=%s LIMIT 1""",(number.strip(),))
                        duplicate=check.fetchone()
                        if duplicate: raise ValueError(f"Booking {number.strip()} is already waiting as AP-{duplicate[0]}.")
                        payload={"booking_number":number.strip(),"supplier_id":int(supplier_map[supplier]),"product_id":int(product_map[product]),
                                 "booking_date":str(booking_date),"valid_to":str(valid_to),"liters":liters,"unit_price":price,
                                 "payment_terms":terms,"transport":transport,"notes":notes or None}
                        request_id=submit_approval_request(conn,"SUPPLIER_BOOKING",f"Supplier booking · {number.strip()} · {liters:,.2f} L",payload,user,liters,booking_value)
                        st.success(f"AP-{request_id} submitted for approval. The booking is not active yet."); st.rerun()
                    else:
                        cur=conn.cursor(); cur.execute("""INSERT INTO procurement_bookings(booking_number,supplier_id,product_id,booking_date,valid_from,valid_to,booked_liters,unit_price,payment_terms,transport_responsibility,status,notes,created_by)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'OPEN',%s,%s) RETURNING id""",(number.strip(),supplier_map[supplier],product_map[product],booking_date,booking_date,valid_to,liters,price,terms,transport,notes or None,user)); booking_id=cur.fetchone()[0]; conn.commit()
                        record_event(conn,"CREATE_BOOKING","Procurement","Supplier Booking",booking_id,f"Created booking {number.strip()} for {liters:,.2f} L"); st.success(f"Booking BK-{booking_id} created."); st.rerun()
                except Exception as error: conn.rollback(); st.error(str(error))
    bookings=booking_options(conn)
    with tab_release:
        if bookings.empty: st.info("Create an open supplier booking first.")
        else:
            booking_map=dict(zip(bookings["label"],bookings["id"])); selected=st.selectbox("Booking",list(booking_map)); row=bookings[bookings["id"]==booking_map[selected]].iloc[0]
            st.info(f"Remaining booking balance: {row['remaining_liters']:,.2f} L")
            destinations=pd.read_sql_query("""SELECT t.id,t.depot_id,CONCAT(d.code,' · ',t.code,' · ',t.name) AS label FROM storage_tanks t JOIN depots d ON d.id=t.depot_id WHERE t.product_id=%s AND t.status<>'OUT_OF_SERVICE' ORDER BY d.code,t.code""",conn,params=[int(row["product_id"])]); destination_map=dict(zip(destinations["label"],destinations["id"]))
            with st.form("new_release"):
                a,b=st.columns(2); number=a.text_input("Release number"); planned=b.date_input("Planned delivery / collection date",date.today()); destination=st.selectbox("Planned destination tank",["Not allocated yet"]+list(destination_map)); liters=st.number_input("Release quantity",min_value=0.0,max_value=float(row["remaining_liters"])); notes=st.text_area("Release notes"); submit=st.form_submit_button("Create release",type="primary")
            if submit:
                if not number.strip() or liters<=0: st.error("Release number and quantity are required.")
                else:
                    try:
                        tank_id=None if destination=="Not allocated yet" else int(destination_map[destination]); depot_id=None if tank_id is None else int(destinations[destinations["id"]==tank_id].iloc[0]["depot_id"])
                        cur=conn.cursor(); cur.execute("""INSERT INTO procurement_releases(booking_id,release_number,release_date,planned_delivery_date,released_liters,destination_depot_id,destination_tank_id,notes,created_by) VALUES (%s,%s,CURRENT_DATE,%s,%s,%s,%s,%s,%s) RETURNING id""",(int(row["id"]),number.strip(),planned,liters,depot_id,tank_id,notes or None,user)); release_id=cur.fetchone()[0]; conn.commit(); record_event(conn,"CREATE_RELEASE","Procurement","Booking Release",release_id,f"Released {liters:,.2f} L against {row['booking_number']}"); st.success(f"Release RL-{release_id} created."); st.rerun()
                    except Exception as error: conn.rollback(); st.error(str(error))
    with tab_overview:
        all_bookings=pd.read_sql_query("""SELECT b.id,b.booking_number,s.name AS supplier,p.name AS product,b.booking_date,b.valid_to,b.booked_liters,b.unit_price,b.payment_terms,b.transport_responsibility,b.status,COALESCE(SUM(tx.accepted_liters),0) AS received_liters,b.booked_liters-COALESCE(SUM(tx.accepted_liters),0) AS remaining_liters,b.created_by FROM procurement_bookings b JOIN suppliers s ON s.id=b.supplier_id JOIN products p ON p.id=b.product_id LEFT JOIN tank_transactions tx ON tx.booking_id=b.id GROUP BY b.id,s.name,p.name ORDER BY b.id DESC""",conn)
        c1,c2,c3,c4=st.columns(4); c1.metric("Bookings",len(all_bookings)); c2.metric("Open quantity",f"{all_bookings['remaining_liters'].clip(lower=0).sum():,.0f} L" if not all_bookings.empty else "0 L"); c3.metric("Booked value",f"{(all_bookings['booked_liters']*all_bookings['unit_price']).sum():,.2f}" if not all_bookings.empty else "0"); c4.metric("Open claims",pd.read_sql_query("SELECT COUNT(*) AS n FROM supplier_claims WHERE status NOT IN ('CLOSED','REJECTED')",conn).iloc[0,0])
        st.dataframe(all_bookings,use_container_width=True,hide_index=True,height=430)
    with tab_claims:
        claims=pd.read_sql_query("""SELECT c.id,c.created_at,s.name AS supplier,b.booking_number,c.claim_type,c.claim_liters,c.unit_price,c.claim_amount,c.status,c.credit_note_number,c.credit_note_date,c.notes,c.created_by FROM supplier_claims c JOIN suppliers s ON s.id=c.supplier_id LEFT JOIN procurement_bookings b ON b.id=c.booking_id ORDER BY c.id DESC""",conn)
        st.dataframe(claims,use_container_width=True,hide_index=True,height=360)
        open_claims=claims[~claims["status"].isin(["CLOSED","REJECTED"])] if not claims.empty else claims
        if not open_claims.empty:
            claim_map={f"CL-{int(r.id)} · {r.supplier} · {r.claim_liters:,.2f} L":int(r.id) for r in open_claims.itertuples()}
            with st.form("claim_update"):
                selected=st.selectbox("Claim",list(claim_map)); status=st.selectbox("New status",["SUBMITTED","ACKNOWLEDGED","CREDIT_NOTE_RECEIVED","CLOSED","REJECTED"]); credit=st.text_input("Credit note number"); credit_date=st.date_input("Credit note date",date.today()); notes=st.text_area("Resolution notes"); submit=st.form_submit_button("Update claim",type="primary")
            if submit:
                claim_id=claim_map[selected]
                if status in ("CLOSED","REJECTED"):
                    if not notes.strip(): st.error("Resolution notes are required before closing or rejecting a claim.")
                    else:
                        check=conn.cursor(); check.execute("""SELECT id FROM approval_requests WHERE request_kind='CLAIM_RESOLUTION'
                            AND status='PENDING' AND payload->>'claim_id'=%s LIMIT 1""",(str(claim_id),)); duplicate=check.fetchone()
                        if duplicate: st.error(f"This claim is already waiting for approval as AP-{duplicate[0]}.")
                        else:
                            payload={"claim_id":claim_id,"target_status":status,"credit_note_number":credit or None,
                                     "credit_note_date":str(credit_date) if credit else None,"notes":notes.strip()}
                            request_id=submit_approval_request(conn,"CLAIM_RESOLUTION",f"Supplier claim {status.lower()} · CL-{claim_id}",payload,user)
                            st.success(f"AP-{request_id} submitted. The claim remains unchanged until approval."); st.rerun()
                else:
                    cur=conn.cursor(); cur.execute("""UPDATE supplier_claims SET status=%s,credit_note_number=%s,
                        credit_note_date=%s,notes=COALESCE(%s,notes) WHERE id=%s""",
                        (status,credit or None,credit_date if credit else None,notes or None,claim_id)); conn.commit()
                    record_event(conn,"UPDATE_CLAIM","Procurement","Supplier Claim",claim_id,f"Claim status changed to {status}"); st.success("Claim updated."); st.rerun()
    with tab_control:
        st.subheader("Controlled cancellations")
        st.caption("Cancellation requests do not change booking or release balances until a different authorized user approves them.")
        cancel_bookings=pd.read_sql_query("""SELECT b.id,b.booking_number,s.name AS supplier,p.name AS product,b.status,
            b.booked_liters,COALESCE(SUM(tx.accepted_liters),0) AS received_liters
            FROM procurement_bookings b JOIN suppliers s ON s.id=b.supplier_id JOIN products p ON p.id=b.product_id
            LEFT JOIN tank_transactions tx ON tx.booking_id=b.id WHERE b.status IN ('OPEN','PARTIALLY_USED')
            GROUP BY b.id,s.name,p.name ORDER BY b.id DESC""",conn)
        cancel_releases=pd.read_sql_query("""SELECT r.id,r.release_number,b.booking_number,r.released_liters,r.status,
            COALESCE(SUM(tx.accepted_liters),0) AS received_liters
            FROM procurement_releases r JOIN procurement_bookings b ON b.id=r.booking_id
            LEFT JOIN tank_transactions tx ON tx.booking_release_id=r.id WHERE r.status='OPEN'
            GROUP BY r.id,b.booking_number ORDER BY r.id DESC""",conn)
        booking_cancel,release_cancel=st.tabs(["Cancel booking","Cancel release"])
        with booking_cancel:
            if cancel_bookings.empty: st.info("No open bookings are available for cancellation.")
            else:
                booking_choices={f"BK-{int(r.id)} · {r.booking_number} · {r.supplier} · {r.booked_liters-r.received_liters:,.2f} L remaining":int(r.id) for r in cancel_bookings.itertuples()}
                with st.form("booking_cancellation_request"):
                    selected_booking=st.selectbox("Booking",list(booking_choices)); reason=st.text_area("Cancellation reason"); reference=st.text_input("Authorization / supporting reference"); submit_cancel=st.form_submit_button("Submit booking cancellation",type="primary")
                if submit_cancel:
                    booking_id=booking_choices[selected_booking]
                    if len(reason.strip())<5 or not reference.strip(): st.error("Enter a clear reason and supporting reference.")
                    else:
                        check=conn.cursor(); check.execute("""SELECT id FROM approval_requests WHERE request_kind='BOOKING_CANCELLATION'
                            AND status='PENDING' AND payload->>'booking_id'=%s LIMIT 1""",(str(booking_id),)); duplicate=check.fetchone()
                        if duplicate: st.error(f"This booking is already waiting as AP-{duplicate[0]}.")
                        else:
                            request_id=submit_approval_request(conn,"BOOKING_CANCELLATION",f"Cancel supplier booking · BK-{booking_id}",
                                {"booking_id":booking_id,"reason":reason.strip(),"reference":reference.strip()},user)
                            st.success(f"AP-{request_id} submitted. The booking remains active until approval."); st.rerun()
        with release_cancel:
            available_releases=cancel_releases[cancel_releases["received_liters"]<=0.005] if not cancel_releases.empty else cancel_releases
            if available_releases.empty: st.info("No unused open releases are available for cancellation.")
            else:
                release_choices={f"RL-{int(r.id)} · {r.release_number} · {r.booking_number} · {r.released_liters:,.2f} L":int(r.id) for r in available_releases.itertuples()}
                with st.form("release_cancellation_request"):
                    selected_release=st.selectbox("Release",list(release_choices)); reason=st.text_area("Cancellation reason",key="release_cancel_reason"); reference=st.text_input("Authorization / supporting reference",key="release_cancel_reference"); submit_cancel=st.form_submit_button("Submit release cancellation",type="primary")
                if submit_cancel:
                    release_id=release_choices[selected_release]
                    if len(reason.strip())<5 or not reference.strip(): st.error("Enter a clear reason and supporting reference.")
                    else:
                        check=conn.cursor(); check.execute("""SELECT id FROM approval_requests WHERE request_kind='RELEASE_CANCELLATION'
                            AND status='PENDING' AND payload->>'release_id'=%s LIMIT 1""",(str(release_id),)); duplicate=check.fetchone()
                        if duplicate: st.error(f"This release is already waiting as AP-{duplicate[0]}.")
                        else:
                            request_id=submit_approval_request(conn,"RELEASE_CANCELLATION",f"Cancel supplier release · RL-{release_id}",
                                {"release_id":release_id,"reason":reason.strip(),"reference":reference.strip()},user)
                            st.success(f"AP-{request_id} submitted. The release remains open until approval."); st.rerun()
    with tab_report:
        all_bookings=pd.read_sql_query("SELECT * FROM procurement_bookings ORDER BY id DESC",conn); releases=pd.read_sql_query("SELECT * FROM procurement_releases ORDER BY id DESC",conn); claims=pd.read_sql_query("SELECT * FROM supplier_claims ORDER BY id DESC",conn)
        report=_report(all_bookings,releases,claims,st.session_state.get("company_profile",{})); st.download_button("Download procurement report",report,f"procurement_{datetime.now():%Y%m%d_%H%M%S}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
