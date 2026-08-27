from datetime import date
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from audit import record_event
from ui import page_header


DECISION_ROLES = {"ADMIN", "INVENTORY_MANAGER", "APPROVER"}
OPERATING_ROLES = DECISION_ROLES | {"STOREKEEPER", "PROCUREMENT_USER", "OPERATOR"}


def ensure_quality_schema(conn):
    cursor = conn.cursor()
    try:
        cursor.execute("""CREATE TABLE IF NOT EXISTS product_quality_specs(
            id BIGSERIAL PRIMARY KEY,product_id INTEGER NOT NULL REFERENCES products(id),effective_from DATE NOT NULL,
            density_min REAL,density_max REAL,temperature_max REAL,water_max_ppm REAL,sulfur_max_ppm REAL,
            flash_point_min REAL,notes TEXT,active BOOLEAN NOT NULL DEFAULT TRUE,created_by TEXT,
            created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP)""")
        cursor.execute("""CREATE TABLE IF NOT EXISTS fuel_batches(
            id BIGSERIAL PRIMARY KEY,batch_number TEXT NOT NULL,supplier_id INTEGER NOT NULL REFERENCES suppliers(id),
            product_id INTEGER NOT NULL REFERENCES products(id),manufacture_date DATE,expiry_date DATE,coa_reference TEXT,
            status TEXT NOT NULL DEFAULT 'QUARANTINE' CHECK(status IN ('DRAFT','QUARANTINE','RELEASED','REJECTED','CLOSED')),
            quarantine_reason TEXT,decision_reason TEXT,decision_reference TEXT,created_by TEXT,
            created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,decision_by TEXT,decision_at TIMESTAMPTZ,
            UNIQUE(supplier_id,batch_number))""")
        cursor.execute("""CREATE TABLE IF NOT EXISTS quality_inspections(
            id BIGSERIAL PRIMARY KEY,batch_id BIGINT NOT NULL REFERENCES fuel_batches(id) ON DELETE CASCADE,
            inspected_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,sample_reference TEXT NOT NULL,
            density REAL,temperature_c REAL,water_ppm REAL,sulfur_ppm REAL,flash_point_c REAL,appearance TEXT,
            result TEXT NOT NULL CHECK(result IN ('PENDING','PASS','FAIL')),exceptions TEXT,notes TEXT,
            inspected_by TEXT,created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP)""")
        cursor.execute("ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS batch_id BIGINT REFERENCES fuel_batches(id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_batches_status ON fuel_batches(status,created_at DESC)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_inspections_batch ON quality_inspections(batch_id,inspected_at DESC)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_tank_transactions_batch ON tank_transactions(batch_id)")
        conn.commit()
    except Exception:
        conn.rollback(); raise


def _maps(conn):
    suppliers = pd.read_sql_query("SELECT id,name FROM suppliers WHERE COALESCE(status,'ACTIVE')='ACTIVE' ORDER BY name", conn)
    products = pd.read_sql_query("SELECT id,name FROM products WHERE active=TRUE ORDER BY name", conn)
    return dict(zip(suppliers.name, suppliers.id)), dict(zip(products.name, products.id))


def _batches(conn):
    return pd.read_sql_query("""SELECT b.id,b.batch_number,s.name supplier,p.name product,b.manufacture_date,b.expiry_date,
        b.coa_reference,b.status,b.quarantine_reason,b.decision_reason,b.decision_reference,b.created_by,b.created_at,
        b.decision_by,b.decision_at,COALESCE(SUM(x.accepted_liters),SUM(x.liters),0) linked_liters,COUNT(x.id) linked_receipts
        FROM fuel_batches b JOIN suppliers s ON s.id=b.supplier_id JOIN products p ON p.id=b.product_id
        LEFT JOIN tank_transactions x ON x.batch_id=b.id GROUP BY b.id,s.name,p.name ORDER BY b.id DESC""", conn)


def _latest_inspection(conn, batch_id):
    return pd.read_sql_query("SELECT * FROM quality_inspections WHERE batch_id=%s ORDER BY inspected_at DESC,id DESC LIMIT 1", conn, params=[batch_id])


def _evaluate(spec, values):
    checks=[]
    rules=[("Density minimum","density","density_min",">="),("Density maximum","density","density_max","<="),
           ("Temperature maximum","temperature_c","temperature_max","<="),("Water maximum","water_ppm","water_max_ppm","<="),
           ("Sulfur maximum","sulfur_ppm","sulfur_max_ppm","<="),("Flash point minimum","flash_point_c","flash_point_min",">=")]
    for label,value_key,limit_key,op in rules:
        limit=spec.get(limit_key); value=values.get(value_key)
        if limit is None or pd.isna(limit) or value is None: continue
        ok=value>=float(limit) if op==">=" else value<=float(limit)
        if not ok: checks.append(f"{label}: {value:g} must be {op} {float(limit):g}")
    return checks


def _quality_report(batches, receipts, inspections, specs, company):
    wb=Workbook(); wb.remove(wb.active)
    red="9E1B1B"; navy="111827"; pale="F3F6FA"; green="DDF3E4"; thin=Side(style="thin",color="D8DEE8")
    summary=wb.create_sheet("Executive Summary")
    summary.append([company,"Product, Batch & Fuel Quality Control"]); summary.merge_cells("A1:F1")
    summary.append(["Generated",pd.Timestamp.now().strftime("%d %b %Y %H:%M"),"Registered batches",len(batches),"Quarantined",int((batches.status=="QUARANTINE").sum()) if not batches.empty else 0])
    summary.append(["Released",int((batches.status=="RELEASED").sum()) if not batches.empty else 0,"Rejected",int((batches.status=="REJECTED").sum()) if not batches.empty else 0,"Linked litres",float(batches.linked_liters.sum()) if not batches.empty else 0])
    datasets={"Batch Register":batches,"Receipt Traceability":receipts,"Inspection Results":inspections,"Specifications":specs}
    for name,data in datasets.items():
        ws=wb.create_sheet(name); clean=data.copy()
        for col in clean.columns:
            if pd.api.types.is_datetime64_any_dtype(clean[col]): clean[col]=clean[col].astype(str)
        ws.append(list(clean.columns))
        for row in clean.where(pd.notna(clean),None).itertuples(index=False,name=None): ws.append(list(row))
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        for cell in ws[1]: cell.fill=PatternFill("solid",fgColor=navy); cell.font=Font(color="FFFFFF",bold=True); cell.alignment=Alignment(wrap_text=True)
        for column in ws.columns:
            letter=column[0].column_letter; ws.column_dimensions[letter].width=min(max(12,max(len(str(c.value or "")) for c in column)+2),34)
        for row in ws.iter_rows():
            for cell in row: cell.border=Border(bottom=thin)
    summary["A1"].fill=PatternFill("solid",fgColor=red); summary["A1"].font=Font(color="FFFFFF",bold=True,size=18)
    summary["A3"].fill=PatternFill("solid",fgColor=green); summary.column_dimensions["A"].width=22
    for col in "BCDEF": summary.column_dimensions[col].width=20
    out=BytesIO(); wb.save(out); return out.getvalue()


def render_product_quality(conn):
    user=st.session_state.get("user","System"); role=st.session_state.get("role","VIEWER")
    company=st.session_state.get("company_profile",{}).get("company_name","Company")
    page_header("Product & Quality Control","Trace every supplier receipt to a controlled fuel batch, inspection and release decision.")
    batches=_batches(conn)
    c1,c2,c3,c4=st.columns(4)
    c1.metric("Registered batches",len(batches)); c2.metric("In quarantine",int((batches.status=="QUARANTINE").sum()) if not batches.empty else 0)
    c3.metric("Released",int((batches.status=="RELEASED").sum()) if not batches.empty else 0); c4.metric("Rejected",int((batches.status=="REJECTED").sum()) if not batches.empty else 0)
    register,create,link,inspect,decision,specs_tab,trace=st.tabs(["Batch register","Create batch","Link receipts","Quality inspection","Release decision","Specifications","Traceability report"])
    supplier_map,product_map=_maps(conn)
    with register:
        if batches.empty: st.info("No fuel batches have been registered yet.")
        else:
            status=st.multiselect("Status",sorted(batches.status.unique()),default=sorted(batches.status.unique())); search=st.text_input("Search batch, supplier, product or COA")
            view=batches[batches.status.isin(status)]
            if search: view=view[view.astype(str).agg(" ".join,axis=1).str.contains(search,case=False,na=False)]
            st.dataframe(view,use_container_width=True,hide_index=True,height=430)
    with create:
        if role not in OPERATING_ROLES: st.info("Your role has view-only access.")
        elif not supplier_map or not product_map: st.warning("Create an active supplier and product first.")
        else:
            with st.form("create_quality_batch"):
                a,b=st.columns(2); number=a.text_input("Supplier batch / lot number"); supplier=b.selectbox("Supplier",list(supplier_map))
                a,b=st.columns(2); product=a.selectbox("Product",list(product_map)); coa=b.text_input("Certificate of analysis reference")
                a,b=st.columns(2); manufactured=a.date_input("Manufacture date",value=None); expiry=b.date_input("Expiry date",value=None)
                reason=st.text_area("Quarantine note",value="Awaiting quality inspection and authorized release.")
                submitted=st.form_submit_button("Create batch in quarantine",type="primary")
            if submitted:
                if len(number.strip())<2 or len(coa.strip())<2: st.error("Enter the supplier batch number and COA reference.")
                elif manufactured and expiry and expiry<manufactured: st.error("Expiry date cannot be earlier than manufacture date.")
                else:
                    try:
                        cur=conn.cursor(); cur.execute("""INSERT INTO fuel_batches(batch_number,supplier_id,product_id,manufacture_date,expiry_date,coa_reference,quarantine_reason,created_by)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",(number.strip(),supplier_map[supplier],product_map[product],manufactured,expiry,coa.strip(),reason.strip() or None,user)); batch_id=cur.fetchone()[0]; conn.commit()
                        record_event(conn,"CREATE_BATCH","Product Quality","Fuel Batch",batch_id,f"Created batch {number.strip()} in quarantine"); st.success(f"Batch FB-{batch_id} created in quarantine."); st.rerun()
                    except Exception as error: conn.rollback(); st.error(str(error))
    with link:
        eligible=pd.read_sql_query("""SELECT x.id,x.movement_at,d.code depot,t.code tank,s.name supplier,p.name product,
            COALESCE(x.accepted_liters,x.liters) accepted_liters,x.reference,x.batch_id,x.supplier_id,x.product_id
            FROM tank_transactions x JOIN storage_tanks t ON t.id=x.tank_id JOIN depots d ON d.id=t.depot_id
            JOIN suppliers s ON s.id=x.supplier_id JOIN products p ON p.id=x.product_id
            WHERE x.movement_category='SUPPLIER_RECEIPT' ORDER BY x.id DESC""",conn)
        if eligible.empty: st.info("No supplier receipts are available for batch linking.")
        else:
            unlinked=eligible[eligible.batch_id.isna()]
            st.caption("This action adds traceability only. It never changes the posted receipt quantity or tank balance.")
            if unlinked.empty: st.success("All supplier receipts are linked to batches.")
            elif role not in OPERATING_ROLES: st.dataframe(unlinked,use_container_width=True,hide_index=True)
            else:
                receipt_labels={f"STX-{int(r.id)} · {r.supplier} · {r.product} · {float(r.accepted_liters):,.2f} L":r for r in unlinked.itertuples()}
                selected=st.selectbox("Unlinked supplier receipt",list(receipt_labels)); row=receipt_labels[selected]
                matching=batches[(batches.supplier==row.supplier)&(batches.product==row.product)&batches.status.isin(["DRAFT","QUARANTINE","RELEASED"])]
                if matching.empty: st.warning("Create a batch for the same supplier and product before linking this receipt.")
                else:
                    choices={f"FB-{int(r.id)} · {r.batch_number} · {r.status}":int(r.id) for r in matching.itertuples()}; target=st.selectbox("Matching batch",list(choices))
                    if st.button("Link receipt to batch",type="primary"):
                        cur=conn.cursor(); cur.execute("UPDATE tank_transactions SET batch_id=%s WHERE id=%s AND batch_id IS NULL",(choices[target],int(row.id)))
                        if cur.rowcount!=1: conn.rollback(); st.error("The receipt was already linked. Refresh and check again.")
                        else:
                            conn.commit(); record_event(conn,"LINK_RECEIPT_BATCH","Product Quality","Tank Transaction",int(row.id),f"Linked STX-{int(row.id)} to FB-{choices[target]} without changing quantity"); st.success("Receipt linked. Posted litres and tank balance were not changed."); st.rerun()
    with inspect:
        available=batches[batches.status.isin(["DRAFT","QUARANTINE"])]
        if available.empty: st.info("No draft or quarantined batch is waiting for inspection.")
        elif role not in OPERATING_ROLES: st.info("Your role has view-only access.")
        else:
            choices={f"FB-{int(r.id)} · {r.batch_number} · {r.product}":int(r.id) for r in available.itertuples()}; label=st.selectbox("Batch to inspect",list(choices)); batch_id=choices[label]
            batch_row=available[available.id==batch_id].iloc[0]
            spec_df=pd.read_sql_query("SELECT * FROM product_quality_specs WHERE product_id=(SELECT product_id FROM fuel_batches WHERE id=%s) AND active=TRUE ORDER BY effective_from DESC,id DESC LIMIT 1",conn,params=[batch_id])
            with st.form("quality_inspection"):
                sample=st.text_input("Sample / laboratory reference"); a,b,c=st.columns(3)
                density=a.number_input("Density",min_value=0.0,value=None); temperature=b.number_input("Temperature °C",value=None); water=c.number_input("Water ppm",min_value=0.0,value=None)
                a,b,c=st.columns(3); sulfur=a.number_input("Sulfur ppm",min_value=0.0,value=None); flash=b.number_input("Flash point °C",value=None); appearance=c.selectbox("Appearance",["Clear and bright","Acceptable","Cloudy","Sediment present","Other"])
                notes=st.text_area("Inspection notes"); submit=st.form_submit_button("Record inspection",type="primary")
            if submit:
                if len(sample.strip())<2: st.error("Enter the sample or laboratory reference.")
                else:
                    values={"density":density,"temperature_c":temperature,"water_ppm":water,"sulfur_ppm":sulfur,"flash_point_c":flash}
                    if spec_df.empty: exceptions=["No active product specification is configured"]; result="PENDING"
                    else: exceptions=_evaluate(spec_df.iloc[0].to_dict(),values); result="FAIL" if exceptions else "PASS"
                    cur=conn.cursor(); cur.execute("""INSERT INTO quality_inspections(batch_id,sample_reference,density,temperature_c,water_ppm,sulfur_ppm,flash_point_c,appearance,result,exceptions,notes,inspected_by)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",(batch_id,sample.strip(),density,temperature,water,sulfur,flash,appearance,result,"; ".join(exceptions) or None,notes.strip() or None,user)); inspection_id=cur.fetchone()[0]; conn.commit()
                    record_event(conn,"QUALITY_INSPECTION","Product Quality","Quality Inspection",inspection_id,f"Inspection result {result} for FB-{batch_id}",severity="WARNING" if result!="PASS" else "INFO"); st.success(f"QI-{inspection_id} recorded with result {result}."); st.rerun()
    with decision:
        pending=batches[batches.status.isin(["DRAFT","QUARANTINE","RELEASED"])]
        if pending.empty: st.info("No batch is waiting for a decision.")
        elif role not in DECISION_ROLES: st.info("Only an authorized approver can release, quarantine or reject a batch.")
        else:
            choices={f"FB-{int(r.id)} · {r.batch_number} · {r.status}":int(r.id) for r in pending.itertuples()}; label=st.selectbox("Batch",list(choices),key="decision_batch"); batch_id=choices[label]
            latest=_latest_inspection(conn,batch_id)
            if latest.empty: st.warning("No inspection is recorded. This batch cannot be released.")
            else: st.dataframe(latest[["id","inspected_at","sample_reference","result","exceptions","inspected_by"]],use_container_width=True,hide_index=True)
            with st.form("batch_decision"):
                action=st.selectbox("Decision",["RELEASED","QUARANTINE","REJECTED"]); reason=st.text_area("Decision reason"); reference=st.text_input("Authorization / supporting reference"); submit=st.form_submit_button("Confirm batch decision",type="primary")
            if submit:
                latest_result=None if latest.empty else latest.iloc[0].result
                if action=="RELEASED" and latest_result!="PASS": st.error("Release is blocked until the latest quality inspection has passed.")
                elif len(reason.strip())<5 or len(reference.strip())<2: st.error("Enter a clear reason and authorization reference.")
                else:
                    cur=conn.cursor(); cur.execute("UPDATE fuel_batches SET status=%s,decision_reason=%s,decision_reference=%s,decision_by=%s,decision_at=CURRENT_TIMESTAMP WHERE id=%s",(action,reason.strip(),reference.strip(),user,batch_id)); conn.commit()
                    record_event(conn,"BATCH_DECISION","Product Quality","Fuel Batch",batch_id,f"Batch changed to {action}; authorization {reference.strip()}",severity="WARNING" if action!="RELEASED" else "INFO"); st.success(f"FB-{batch_id} is now {action}."); st.rerun()
    with specs_tab:
        specs=pd.read_sql_query("""SELECT q.id,p.name product,q.effective_from,q.density_min,q.density_max,q.temperature_max,q.water_max_ppm,q.sulfur_max_ppm,q.flash_point_min,q.active,q.created_by,q.created_at FROM product_quality_specs q JOIN products p ON p.id=q.product_id ORDER BY q.id DESC""",conn)
        st.dataframe(specs,use_container_width=True,hide_index=True)
        if role in {"ADMIN","INVENTORY_MANAGER"} and product_map:
            with st.expander("Create a new active specification"):
                with st.form("quality_spec"):
                    product=st.selectbox("Product",list(product_map),key="spec_product"); effective=st.date_input("Effective from",date.today())
                    a,b,c=st.columns(3); dmin=a.number_input("Density minimum",min_value=0.0,value=None); dmax=b.number_input("Density maximum",min_value=0.0,value=None); tmax=c.number_input("Temperature maximum °C",value=None)
                    a,b,c=st.columns(3); wmax=a.number_input("Water maximum ppm",min_value=0.0,value=None); smax=b.number_input("Sulfur maximum ppm",min_value=0.0,value=None); fmin=c.number_input("Flash point minimum °C",value=None)
                    notes=st.text_area("Specification notes"); submit=st.form_submit_button("Activate specification",type="primary")
                if submit:
                    if dmin is not None and dmax is not None and dmin>dmax: st.error("Density minimum cannot exceed density maximum.")
                    elif all(v is None for v in [dmin,dmax,tmax,wmax,smax,fmin]): st.error("Enter at least one quality limit.")
                    else:
                        cur=conn.cursor(); cur.execute("UPDATE product_quality_specs SET active=FALSE WHERE product_id=%s",(product_map[product],)); cur.execute("""INSERT INTO product_quality_specs(product_id,effective_from,density_min,density_max,temperature_max,water_max_ppm,sulfur_max_ppm,flash_point_min,notes,created_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",(product_map[product],effective,dmin,dmax,tmax,wmax,smax,fmin,notes.strip() or None,user)); spec_id=cur.fetchone()[0]; conn.commit(); record_event(conn,"ACTIVATE_QUALITY_SPEC","Product Quality","Product Quality Specification",spec_id,f"Activated specification for {product}"); st.success(f"Specification QS-{spec_id} activated."); st.rerun()
    with trace:
        receipts=pd.read_sql_query("""SELECT x.id transaction_id,x.movement_at,d.code depot,t.code tank,s.name supplier,p.name product,COALESCE(x.accepted_liters,x.liters) accepted_liters,x.reference,b.id batch_id,b.batch_number,b.status batch_status FROM tank_transactions x JOIN storage_tanks t ON t.id=x.tank_id JOIN depots d ON d.id=t.depot_id LEFT JOIN suppliers s ON s.id=x.supplier_id LEFT JOIN products p ON p.id=x.product_id LEFT JOIN fuel_batches b ON b.id=x.batch_id WHERE x.movement_category='SUPPLIER_RECEIPT' ORDER BY x.id DESC""",conn)
        inspections=pd.read_sql_query("""SELECT q.id,b.batch_number,s.name supplier,p.name product,q.inspected_at,q.sample_reference,q.density,q.temperature_c,q.water_ppm,q.sulfur_ppm,q.flash_point_c,q.appearance,q.result,q.exceptions,q.inspected_by FROM quality_inspections q JOIN fuel_batches b ON b.id=q.batch_id JOIN suppliers s ON s.id=b.supplier_id JOIN products p ON p.id=b.product_id ORDER BY q.id DESC""",conn)
        specs=pd.read_sql_query("""SELECT q.id,p.name product,q.effective_from,q.density_min,q.density_max,q.temperature_max,q.water_max_ppm,q.sulfur_max_ppm,q.flash_point_min,q.active,q.created_by FROM product_quality_specs q JOIN products p ON p.id=q.product_id ORDER BY q.id DESC""",conn)
        st.dataframe(receipts,use_container_width=True,hide_index=True,height=300)
        report=_quality_report(batches,receipts,inspections,specs,company)
        st.download_button("Download batch traceability & quality report",report,"product_batch_quality_report.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
