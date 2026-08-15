from datetime import date, datetime, timedelta
import io

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ui import page_header


def ensure_forecasting_schema(conn):
    cursor=conn.cursor()
    try:
        cursor.execute("ALTER TABLE procurement_releases ADD COLUMN IF NOT EXISTS destination_depot_id INTEGER REFERENCES depots(id)")
        cursor.execute("ALTER TABLE procurement_releases ADD COLUMN IF NOT EXISTS destination_tank_id INTEGER REFERENCES storage_tanks(id)")
        conn.commit()
    except Exception:
        conn.rollback(); raise


def calculate_forecast(conn,lookback_days,horizon_days,safety_days,growth_percent):
    ensure_forecasting_schema(conn)
    tanks=pd.read_sql_query("""SELECT t.id,CONCAT(d.code,' · ',t.code) AS tank,d.code AS depot,t.code,p.name AS product,
        t.safe_capacity_liters,t.minimum_stock_liters,t.reorder_level_liters,
        COALESCE(SUM(CASE WHEN tx.type='IN' THEN tx.liters ELSE -tx.liters END),0) AS current_stock
        FROM storage_tanks t JOIN depots d ON d.id=t.depot_id JOIN products p ON p.id=t.product_id
        LEFT JOIN tank_transactions tx ON tx.tank_id=t.id WHERE t.status<>'OUT_OF_SERVICE'
        GROUP BY t.id,d.code,p.name ORDER BY d.code,t.code""",conn)
    if tanks.empty: return tanks,pd.DataFrame(),pd.DataFrame()
    start=date.today()-timedelta(days=lookback_days-1)
    usage=pd.read_sql_query("""SELECT tank_id,DATE(movement_at AT TIME ZONE 'Asia/Dubai') AS day,SUM(liters) AS issued_liters
        FROM tank_transactions WHERE type='OUT' AND movement_category IN ('TANK_TO_TRUCK','STANDARD')
        AND movement_at >= %s GROUP BY tank_id,DATE(movement_at AT TIME ZONE 'Asia/Dubai') ORDER BY day""",conn,params=[start])
    incoming=pd.read_sql_query("""SELECT r.destination_tank_id,SUM(r.released_liters-COALESCE(received.received,0)) AS incoming_liters
        FROM procurement_releases r LEFT JOIN (SELECT booking_release_id,SUM(accepted_liters) AS received FROM tank_transactions WHERE booking_release_id IS NOT NULL GROUP BY booking_release_id) received ON received.booking_release_id=r.id
        WHERE r.status IN ('OPEN','PARTIALLY_RECEIVED') AND r.destination_tank_id IS NOT NULL
        AND COALESCE(r.planned_delivery_date,CURRENT_DATE)<=%s GROUP BY r.destination_tank_id""",conn,params=[date.today()+timedelta(days=horizon_days)])
    unallocated=pd.read_sql_query("""SELECT r.release_number,s.name AS supplier,p.name AS product,r.planned_delivery_date,r.released_liters,r.status
        FROM procurement_releases r JOIN procurement_bookings b ON b.id=r.booking_id JOIN suppliers s ON s.id=b.supplier_id JOIN products p ON p.id=b.product_id
        WHERE r.status IN ('OPEN','PARTIALLY_RECEIVED') AND r.destination_tank_id IS NULL ORDER BY r.planned_delivery_date""",conn)
    totals=usage.groupby("tank_id")["issued_liters"].sum() if not usage.empty else pd.Series(dtype=float)
    incoming_map=dict(zip(incoming["destination_tank_id"],incoming["incoming_liters"])) if not incoming.empty else {}
    rows=[]
    growth=1+growth_percent/100
    for tank in tanks.itertuples():
        historical=float(totals.get(tank.id,0)); daily=historical/lookback_days; forecast_daily=daily*growth
        arriving=float(incoming_map.get(tank.id,0) or 0); available=float(tank.current_stock)+arriving
        demand=forecast_daily*horizon_days; safety=forecast_daily*safety_days
        projected=available-demand; days_cover=available/forecast_daily if forecast_daily>0 else None
        stockout=date.today()+timedelta(days=max(int(available/forecast_daily),0)) if forecast_daily>0 else None
        target=max(float(tank.reorder_level_liters or 0),demand+safety)
        order=max(target-available,0); minimum=float(tank.minimum_stock_liters or 0)
        if float(tank.current_stock)<=minimum: risk="Critical"
        elif order>0 and (days_cover is None or days_cover<=safety_days): risk="Reorder"
        elif order>0: risk="Monitor"
        else: risk="Healthy"
        rows.append({"Tank":tank.tank,"Depot":tank.depot,"Product":tank.product,"Current Stock (L)":float(tank.current_stock),"Incoming (L)":arriving,
            "Historical Usage (L)":historical,"Average Daily Use (L)":daily,"Forecast Daily Use (L)":forecast_daily,"Forecast Demand (L)":demand,
            "Safety Stock (L)":safety,"Projected Balance (L)":projected,"Days Cover":days_cover,"Estimated Stock-out":stockout,
            "Suggested Order (L)":order,"Safe Capacity (L)":float(tank.safe_capacity_liters),"Risk":risk})
    return pd.DataFrame(rows),usage,unallocated


def _excel_report(forecast,usage,unallocated,profile,assumptions):
    wb=Workbook(); sheets=[(wb.active,"Forecast",forecast),(wb.create_sheet(),"Daily Demand",usage),(wb.create_sheet(),"Unallocated Releases",unallocated)]
    primary=str(profile.get("primary_color","#8C1C1C")).replace("#",""); secondary=str(profile.get("secondary_color","#172033")).replace("#",""); company=profile.get("company_name","FILLIT")
    for index,(ws,name,frame) in enumerate(sheets):
        ws.title=name; ws.sheet_view.showGridLines=False; end=max(len(frame.columns),8); letter=get_column_letter(end)
        ws.merge_cells(f"A1:{letter}2"); ws["A1"]=f"{company} | Inventory {name}"; ws["A1"].fill=PatternFill("solid",fgColor=secondary); ws["A1"].font=Font(size=20,bold=True,color="FFFFFF"); ws["A1"].alignment=Alignment(vertical="center")
        ws.merge_cells(f"A3:{letter}3"); ws["A3"]=f"Generated {datetime.now():%d %b %Y, %I:%M %p} (Dubai) · {assumptions}"; ws["A3"].fill=PatternFill("solid",fgColor=primary); ws["A3"].font=Font(color="FFFFFF")
        if frame.empty: ws["A5"]="No records available"; continue
        headers=list(frame.columns)
        for col,h in enumerate(headers,1): c=ws.cell(5,col,h); c.fill=PatternFill("solid",fgColor=primary); c.font=Font(bold=True,color="FFFFFF"); c.alignment=Alignment(wrap_text=True)
        for r,values in enumerate(frame.itertuples(index=False,name=None),6):
            for c,value in enumerate(values,1):
                if isinstance(value,pd.Timestamp): value=value.to_pydatetime().replace(tzinfo=None)
                ws.cell(r,c,value)
        last=5+len(frame); table=Table(displayName=f"ForecastTable{index+1}",ref=f"A5:{get_column_letter(len(headers))}{last}"); table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True); ws.add_table(table); ws.freeze_panes="A6"
        for c,h in enumerate(headers,1):
            ws.column_dimensions[get_column_letter(c)].width=18 if "(" in str(h) or "Date" in str(h) or "Stock" in str(h) else min(max(len(str(h))+3,13),24)
            if "(L)" in str(h):
                for r in range(6,last+1): ws.cell(r,c).number_format='#,##0.00 "L";[Red]-#,##0.00 "L"'
            if "Date" in str(h) or "Stock-out" in str(h):
                for r in range(6,last+1): ws.cell(r,c).number_format="dd mmm yyyy"
        if "Risk" in headers:
            risk_col=get_column_letter(headers.index("Risk")+1)
            for status,colour in [("Critical","FEE4E2"),("Reorder","FEF0C7"),("Monitor","FFF7CC"),("Healthy","DCFCE7")]: ws.conditional_formatting.add(f"A6:{get_column_letter(len(headers))}{last}",FormulaRule(formula=[f'${risk_col}6="{status}"'],fill=PatternFill("solid",fgColor=colour)))
    out=io.BytesIO(); wb.save(out); return out.getvalue()


def render_forecasting(conn):
    page_header("Inventory Forecasting", "Predict demand, stock-out risk and purchasing requirements from actual storage movements.")
    a,b,c,d=st.columns(4); lookback=a.selectbox("History period",[7,14,30,60,90],index=2,format_func=lambda x:f"Last {x} days"); horizon=b.selectbox("Forecast period",[7,14,30,60,90],index=2,format_func=lambda x:f"Next {x} days"); safety=c.number_input("Safety-stock days",min_value=0,max_value=90,value=7); growth=d.number_input("Demand adjustment %",min_value=-50.0,max_value=200.0,value=0.0)
    forecast,usage,unallocated=calculate_forecast(conn,lookback,horizon,safety,growth)
    if forecast.empty: st.warning("Create storage tanks before running a forecast."); return
    c1,c2,c3,c4=st.columns(4); c1.metric("Current storage",f"{forecast['Current Stock (L)'].sum():,.0f} L"); c2.metric("Forecast demand",f"{forecast['Forecast Demand (L)'].sum():,.0f} L"); c3.metric("Confirmed incoming",f"{forecast['Incoming (L)'].sum():,.0f} L"); c4.metric("Suggested purchasing",f"{forecast['Suggested Order (L)'].sum():,.0f} L")
    st.plotly_chart(px.bar(forecast,x="Tank",y=["Current Stock (L)","Forecast Demand (L)","Incoming (L)"],barmode="group",title="Stock, demand and confirmed incoming"),use_container_width=True)
    st.dataframe(forecast,use_container_width=True,hide_index=True,height=420,column_config={c:st.column_config.NumberColumn(format="%.2f L") for c in forecast.columns if "(L)" in c})
    if not unallocated.empty: st.warning(f"{len(unallocated)} open supplier release(s) are not included because no destination tank is allocated."); st.dataframe(unallocated,use_container_width=True,hide_index=True)
    assumptions=f"Lookback {lookback} days · Horizon {horizon} days · Safety {safety} days · Demand adjustment {growth:+.1f}%"
    report=_excel_report(forecast,usage,unallocated,st.session_state.get("company_profile",{}),assumptions)
    st.download_button("Download forecasting report",report,f"inventory_forecast_{datetime.now():%Y%m%d_%H%M%S}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
