from datetime import date, datetime, timedelta
import io
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from ui import page_header

REPORTS={
"Tank Inventory":("Inventory","""SELECT d.code AS depot,t.code AS tank,t.name,p.name AS product,t.safe_capacity_liters,t.minimum_stock_liters,t.reorder_level_liters,COALESCE(SUM(CASE WHEN x.type='IN' THEN x.liters ELSE -x.liters END),0) AS balance_liters,t.status FROM storage_tanks t JOIN depots d ON d.id=t.depot_id JOIN products p ON p.id=t.product_id LEFT JOIN tank_transactions x ON x.tank_id=t.id GROUP BY t.id,d.code,p.name ORDER BY d.code,t.code""",None),
"Truck Inventory":("Inventory","""SELECT CONCAT(t.emirate,' ',t.plate_code,' ',t.plate_number) AS truck,p.name AS product,t.capacity_liters,t.minimum_stock_liters,t.reorder_level_liters,COALESCE(SUM(CASE WHEN x.type='IN' THEN x.liters ELSE -x.liters END),0) AS balance_liters,t.operational_status FROM trucks t LEFT JOIN products p ON p.id=t.product_id LEFT JOIN transactions x ON x.truck_id=t.id GROUP BY t.id,p.name ORDER BY truck""",None),
"Storage Movements":("Operations","""SELECT x.id,x.movement_at,d.code AS depot,t.code AS tank,p.name AS product,x.type,x.liters,x.movement_category,s.name AS supplier,x.ordered_liters,x.dispatched_liters,x.accepted_liters,x.variance_liters,x.reference,x.purchase_type,x.created_by FROM tank_transactions x JOIN storage_tanks t ON t.id=x.tank_id JOIN depots d ON d.id=t.depot_id LEFT JOIN products p ON p.id=x.product_id LEFT JOIN suppliers s ON s.id=x.supplier_id WHERE DATE(x.movement_at AT TIME ZONE 'Asia/Dubai') BETWEEN %s AND %s ORDER BY x.id DESC""","date"),
"Truck Transactions":("Operations","""SELECT x.id,x.date,CONCAT(t.emirate,' ',t.plate_code,' ',t.plate_number) AS truck,p.name AS product,x.type,x.liters,x.movement_category,x.ticket_number,x.created_by,x.record_status FROM transactions x JOIN trucks t ON t.id=x.truck_id LEFT JOIN products p ON p.id=x.product_id WHERE x.date::date BETWEEN %s AND %s ORDER BY x.id DESC""","date"),
"Supplier Bookings":("Procurement","""SELECT b.id,b.booking_number,s.name AS supplier,p.name AS product,b.booking_date,b.valid_to,b.booked_liters,b.unit_price,b.payment_terms,b.transport_responsibility,b.status,b.created_by FROM procurement_bookings b JOIN suppliers s ON s.id=b.supplier_id JOIN products p ON p.id=b.product_id WHERE b.booking_date BETWEEN %s AND %s ORDER BY b.id DESC""","date"),
"Booking Releases":("Procurement","""SELECT r.id,r.release_number,b.booking_number,s.name AS supplier,p.name AS product,r.release_date,r.planned_delivery_date,r.released_liters,d.code AS depot,t.code AS tank,r.status,r.created_by FROM procurement_releases r JOIN procurement_bookings b ON b.id=r.booking_id JOIN suppliers s ON s.id=b.supplier_id JOIN products p ON p.id=b.product_id LEFT JOIN depots d ON d.id=r.destination_depot_id LEFT JOIN storage_tanks t ON t.id=r.destination_tank_id WHERE r.release_date BETWEEN %s AND %s ORDER BY r.id DESC""","date"),
"Supplier Claims":("Procurement","""SELECT c.id,c.created_at,s.name AS supplier,b.booking_number,c.claim_type,c.claim_liters,c.unit_price,c.claim_amount,c.status,c.credit_note_number,c.credit_note_date,c.created_by FROM supplier_claims c JOIN suppliers s ON s.id=c.supplier_id LEFT JOIN procurement_bookings b ON b.id=c.booking_id WHERE DATE(c.created_at AT TIME ZONE 'Asia/Dubai') BETWEEN %s AND %s ORDER BY c.id DESC""","date"),
"Stock in Transit":("Inventory","""SELECT m.id,m.transfer_number,sd.code source_depot,st.code source_tank,dd.code destination_depot,dt.code destination_tank,p.name product,m.planned_liters,m.dispatched_liters,m.received_liters,m.variance_liters,m.status,m.vehicle_number,m.driver_name,m.seal_number,m.dispatched_at,m.received_at FROM inventory_transfers m JOIN storage_tanks st ON st.id=m.source_tank_id JOIN depots sd ON sd.id=st.depot_id JOIN storage_tanks dt ON dt.id=m.destination_tank_id JOIN depots dd ON dd.id=dt.depot_id JOIN products p ON p.id=m.product_id WHERE DATE(m.created_at AT TIME ZONE 'Asia/Dubai') BETWEEN %s AND %s ORDER BY m.id DESC""","date"),
"Tank Cycle Counts":("Control","""SELECT c.id,c.counted_at,d.code depot,t.code tank,c.system_liters,c.physical_liters,c.variance_liters,c.variance_percent,c.measurement_method,c.reference,c.reason,c.status,c.counted_by,c.reviewed_by FROM tank_cycle_counts c JOIN storage_tanks t ON t.id=c.tank_id JOIN depots d ON d.id=t.depot_id WHERE DATE(c.counted_at AT TIME ZONE 'Asia/Dubai') BETWEEN %s AND %s ORDER BY c.id DESC""","date"),
"Inventory Incidents":("Control","""SELECT i.id,i.incident_number,i.occurred_at,i.asset_reference,i.incident_type,p.name product,i.estimated_liters,i.description,i.immediate_action,i.status,i.root_cause,i.corrective_action,i.owner,i.created_by FROM inventory_incidents i LEFT JOIN products p ON p.id=i.product_id WHERE DATE(i.occurred_at AT TIME ZONE 'Asia/Dubai') BETWEEN %s AND %s ORDER BY i.id DESC""","date"),
"Batch Register":("Quality","""SELECT b.id,b.batch_number,s.name supplier,p.name product,b.manufacture_date,b.expiry_date,b.coa_reference,b.status,b.decision_reason,b.decision_reference,b.created_by,b.decision_by FROM fuel_batches b JOIN suppliers s ON s.id=b.supplier_id JOIN products p ON p.id=b.product_id WHERE DATE(b.created_at AT TIME ZONE 'Asia/Dubai') BETWEEN %s AND %s ORDER BY b.id DESC""","date"),
"Quality Inspections":("Quality","""SELECT q.id,q.inspected_at,b.batch_number,p.name product,q.sample_reference,q.density,q.temperature_c,q.water_ppm,q.sulfur_ppm,q.flash_point_c,q.appearance,q.result,q.exceptions,q.inspected_by FROM quality_inspections q JOIN fuel_batches b ON b.id=q.batch_id JOIN products p ON p.id=b.product_id WHERE DATE(q.inspected_at AT TIME ZONE 'Asia/Dubai') BETWEEN %s AND %s ORDER BY q.id DESC""","date"),
"Receipt Invoice Match":("Finance","""SELECT i.id,i.invoice_date,i.supplier_invoice_number,s.name supplier,p.name product,x.id receipt_id,i.invoiced_liters,i.invoiced_unit_price,i.quantity_variance,i.price_variance,i.invoice_total,i.landed_cost_total,i.landed_unit_cost,i.status,i.created_by,i.approved_by FROM supplier_receipt_invoices i JOIN tank_transactions x ON x.id=i.tank_transaction_id JOIN suppliers s ON s.id=x.supplier_id JOIN products p ON p.id=x.product_id WHERE i.invoice_date BETWEEN %s AND %s ORDER BY i.id DESC""","date"),
"Reconciliations":("Control","""SELECT r.id,r.reading_at,CONCAT(t.emirate,' ',t.plate_code,' ',t.plate_number) AS truck,r.system_quantity,r.physical_quantity,r.variance_quantity,r.variance_percent,r.reason,r.status,r.recorded_by,r.reviewed_by FROM stock_reconciliations r JOIN trucks t ON t.id=r.truck_id WHERE DATE(r.reading_at AT TIME ZONE 'Asia/Dubai') BETWEEN %s AND %s ORDER BY r.id DESC""","date"),
"Audit Activity":("Governance","""SELECT id,occurred_at,username,user_role,action,module,entity_type,entity_id,description,status,severity FROM audit_events WHERE DATE(occurred_at AT TIME ZONE 'Asia/Dubai') BETWEEN %s AND %s ORDER BY id DESC""","date")}

def _excel(name,data,profile,period):
 wb=Workbook(); ws=wb.active; ws.title="Report"; primary=str(profile.get("primary_color","#8C1C1C")).replace("#",""); secondary=str(profile.get("secondary_color","#172033")).replace("#",""); company=profile.get("company_name","FILLIT"); end=max(len(data.columns),8); letter=get_column_letter(end)
 ws.merge_cells(f"A1:{letter}2"); ws["A1"]=f"{company} | {name}"; ws["A1"].fill=PatternFill("solid",fgColor=secondary); ws["A1"].font=Font(size=20,bold=True,color="FFFFFF"); ws["A1"].alignment=Alignment(vertical="center")
 ws.merge_cells(f"A3:{letter}3"); ws["A3"]=f"Period {period} · Generated {datetime.now():%d %b %Y, %I:%M %p} (Dubai) · {len(data):,} records"; ws["A3"].fill=PatternFill("solid",fgColor=primary); ws["A3"].font=Font(color="FFFFFF")
 if data.empty: ws["A5"]="No matching records"
 else:
  for c,h in enumerate(data.columns,1): cell=ws.cell(5,c,str(h).replace("_"," ").title()); cell.fill=PatternFill("solid",fgColor=primary); cell.font=Font(bold=True,color="FFFFFF"); cell.alignment=Alignment(wrap_text=True)
  for r,values in enumerate(data.itertuples(index=False,name=None),6):
   for c,v in enumerate(values,1):
    if isinstance(v,pd.Timestamp): v=v.to_pydatetime().replace(tzinfo=None)
    ws.cell(r,c,v)
  last=5+len(data); table=Table(displayName="MasterReportTable",ref=f"A5:{get_column_letter(len(data.columns))}{last}"); table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True); ws.add_table(table); ws.freeze_panes="A6"
  for c,h in enumerate(data.columns,1):
   ws.column_dimensions[get_column_letter(c)].width=min(max(len(str(h))+4,14),28)
   if "liter" in str(h).lower() or "quantity" in str(h).lower() or "balance" in str(h).lower():
    for r in range(6,last+1): ws.cell(r,c).number_format='#,##0.00 "L";[Red]-#,##0.00 "L"'
 out=io.BytesIO(); wb.save(out); return out.getvalue()

def render_master_reports(conn):
 page_header("Master Report Centre","Generate management, operational, procurement, control and audit reports from one place.")
 c1,c2,c3=st.columns([2,1,1]); name=c1.selectbox("Report",list(REPORTS),format_func=lambda x:f"{REPORTS[x][0]} · {x}"); start=c2.date_input("From",date.today()-timedelta(days=30)); end=c3.date_input("To",date.today())
 if start>end: st.error("From date cannot be after To date."); return
 category,sql,dated=REPORTS[name]
 try: data=pd.read_sql_query(sql,conn,params=[start,end] if dated else None)
 except Exception as error: st.error(f"This report is not available yet: {error}"); return
 search=st.text_input("Search within this report",placeholder="Supplier, truck, tank, reference, user, status...").strip().lower()
 if search and not data.empty: data=data[data.astype(str).apply(lambda row:row.str.lower().str.contains(search,na=False).any(),axis=1)]
 a,b,c=st.columns(3); a.metric("Report category",category); b.metric("Matching records",f"{len(data):,}"); c.metric("Report period",f"{(end-start).days+1} days")
 st.dataframe(data,use_container_width=True,hide_index=True,height=500)
 report=_excel(name,data,st.session_state.get("company_profile",{}),f"{start:%d %b %Y} to {end:%d %b %Y}")
 st.download_button("Download professional Excel report",report,f"{name.lower().replace(' ','_')}_{datetime.now():%Y%m%d_%H%M%S}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
