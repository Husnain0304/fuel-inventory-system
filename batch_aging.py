from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from audit import record_event
from ui import page_header


EDIT_ROLES={"ADMIN","INVENTORY_MANAGER","STOREKEEPER","OPERATOR"}


def ensure_batch_aging_schema(conn):
    c=conn.cursor()
    try:
        c.execute("""CREATE TABLE IF NOT EXISTS batch_issue_allocations(
            id BIGSERIAL PRIMARY KEY,batch_id BIGINT NOT NULL REFERENCES fuel_batches(id),
            tank_transaction_id BIGINT UNIQUE REFERENCES tank_transactions(id),allocated_liters REAL NOT NULL CHECK(allocated_liters>0),
            allocation_method TEXT NOT NULL DEFAULT 'FEFO',exception_reason TEXT,created_by TEXT,
            created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP)""")
        c.execute("""CREATE TABLE IF NOT EXISTS batch_aging_settings(
            id INTEGER PRIMARY KEY DEFAULT 1,expiry_warning_days INTEGER NOT NULL DEFAULT 60,
            critical_warning_days INTEGER NOT NULL DEFAULT 30,updated_by TEXT,updated_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP)""")
        c.execute("INSERT INTO batch_aging_settings(id) VALUES (1) ON CONFLICT(id) DO NOTHING")
        c.execute("CREATE INDEX IF NOT EXISTS idx_batch_alloc_batch ON batch_issue_allocations(batch_id)")
        conn.commit()
    except Exception: conn.rollback(); raise


def _settings(conn):
    return pd.read_sql_query("SELECT expiry_warning_days,critical_warning_days FROM batch_aging_settings WHERE id=1",conn).iloc[0]


def _position(conn):
    return pd.read_sql_query("""SELECT b.id,b.batch_number,s.name supplier,p.name product,b.expiry_date,b.status,
        COALESCE(r.received_liters,0) received_liters,COALESCE(a.allocated_liters,0) allocated_liters,
        GREATEST(COALESCE(r.received_liters,0)-COALESCE(a.allocated_liters,0),0) available_liters,
        CASE WHEN b.expiry_date IS NULL THEN NULL ELSE b.expiry_date-CURRENT_DATE END days_to_expiry,
        CASE WHEN b.status!='RELEASED' THEN 'BLOCKED'
             WHEN b.expiry_date<CURRENT_DATE THEN 'EXPIRED'
             WHEN b.expiry_date<=CURRENT_DATE+(SELECT critical_warning_days FROM batch_aging_settings WHERE id=1) THEN 'CRITICAL'
             WHEN b.expiry_date<=CURRENT_DATE+(SELECT expiry_warning_days FROM batch_aging_settings WHERE id=1) THEN 'WARNING'
             ELSE 'NORMAL' END aging_status
        FROM fuel_batches b JOIN suppliers s ON s.id=b.supplier_id JOIN products p ON p.id=b.product_id
        LEFT JOIN (SELECT batch_id,SUM(COALESCE(accepted_liters,liters)) received_liters FROM tank_transactions WHERE batch_id IS NOT NULL AND type='IN' GROUP BY batch_id) r ON r.batch_id=b.id
        LEFT JOIN (SELECT batch_id,SUM(allocated_liters) allocated_liters FROM batch_issue_allocations GROUP BY batch_id) a ON a.batch_id=b.id
        ORDER BY CASE WHEN b.status='RELEASED' THEN 0 ELSE 1 END,b.expiry_date NULLS LAST,b.id""",conn)


def _report(position,allocations,settings,company):
    wb=Workbook(); wb.remove(wb.active)
    for name,data in {"Executive Summary":pd.DataFrame({"Control":["Company","Warning days","Critical days","Available released litres","Expired/critical batches"],"Value":[company,int(settings.expiry_warning_days),int(settings.critical_warning_days),float(position.loc[position.status.eq('RELEASED'),'available_liters'].sum()),int(position.aging_status.isin(['EXPIRED','CRITICAL']).sum())]}),"Batch Aging":position,"Issue Allocations":allocations}.items():
        ws=wb.create_sheet(name); ws.append(list(data.columns))
        for row in data.where(pd.notna(data),None).itertuples(index=False,name=None): ws.append([str(v) if isinstance(v,pd.Timestamp) else v for v in row])
        for cell in ws[1]: cell.fill=PatternFill("solid",fgColor="111827"); cell.font=Font(color="FFFFFF",bold=True)
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(max(14,max(len(str(x.value or "")) for x in col)+2),32)
    wb["Executive Summary"]["A1"].fill=PatternFill("solid",fgColor="9E1B1B")
    out=BytesIO(); wb.save(out); return out.getvalue()


def render_batch_aging(conn):
    role=st.session_state.get("role","VIEWER"); user=st.session_state.get("user","System"); settings=_settings(conn); position=_position(conn)
    page_header("Batch Aging & FEFO Control","Use released fuel in earliest-expiry order and prevent expired or quarantined stock from being selected.")
    released=position[position.status.eq("RELEASED")]
    a,b,c,d=st.columns(4); a.metric("Released available",f"{released.available_liters.sum():,.0f} L"); b.metric("Expiry warnings",int(position.aging_status.eq("WARNING").sum())); c.metric("Critical / expired",int(position.aging_status.isin(["CRITICAL","EXPIRED"]).sum())); d.metric("Blocked batches",int(position.aging_status.eq("BLOCKED").sum()))
    overview,allocate,history,configuration=st.tabs(["FEFO stock position","Allocate issued fuel","Allocation history","Settings & report"])
    with overview:
        st.caption("FEFO means First Expiry, First Out. The top released batch with available stock is the recommended batch to consume first.")
        st.dataframe(position,use_container_width=True,hide_index=True,height=460)
    with allocate:
        issues=pd.read_sql_query("""SELECT x.id,x.movement_at,d.code depot,t.code tank,p.name product,x.liters,x.reference,x.product_id
            FROM tank_transactions x JOIN storage_tanks t ON t.id=x.tank_id JOIN depots d ON d.id=t.depot_id LEFT JOIN products p ON p.id=x.product_id
            LEFT JOIN batch_issue_allocations a ON a.tank_transaction_id=x.id WHERE x.type='OUT' AND a.id IS NULL ORDER BY x.id DESC""",conn)
        if issues.empty: st.success("Every storage issue is allocated to a batch.")
        elif role not in EDIT_ROLES: st.dataframe(issues,use_container_width=True,hide_index=True)
        else:
            labels={f"STX-{int(r.id)} · {r.depot}/{r.tank} · {r.product} · {float(r.liters):,.2f} L":r for r in issues.itertuples()}; choice=st.selectbox("Unallocated storage issue",list(labels)); issue=labels[choice]
            eligible=position[(position.status.eq("RELEASED"))&(position.product.eq(issue.product))&(position.available_liters>0)&(~position.aging_status.eq("EXPIRED"))]
            if eligible.empty: st.error("No released, unexpired batch with available quantity matches this product.")
            else:
                recommended=eligible.iloc[0]; options={f"FB-{int(r.id)} · {r.batch_number} · expires {r.expiry_date} · {float(r.available_liters):,.2f} L":r for r in eligible.itertuples()}; selected=st.selectbox("Batch",list(options)); batch=options[selected]
                if int(batch.id)==int(recommended.id): st.success("FEFO recommendation followed: this is the earliest-expiring eligible batch.")
                reason=st.text_area("FEFO exception reason",disabled=int(batch.id)==int(recommended.id),help="Required only when a later-expiring batch is selected.")
                if st.button("Allocate issue to batch",type="primary"):
                    if float(issue.liters)>float(batch.available_liters)+0.005: st.error("The selected batch does not have enough available litres.")
                    elif int(batch.id)!=int(recommended.id) and len(reason.strip())<5: st.error("Enter a clear reason for overriding the FEFO recommendation.")
                    else:
                        c=conn.cursor(); c.execute("INSERT INTO batch_issue_allocations(batch_id,tank_transaction_id,allocated_liters,allocation_method,exception_reason,created_by) VALUES (%s,%s,%s,'FEFO',%s,%s) RETURNING id",(int(batch.id),int(issue.id),float(issue.liters),reason.strip() or None,user)); allocation_id=c.fetchone()[0]; conn.commit(); record_event(conn,"ALLOCATE_BATCH_ISSUE","Batch Aging","Batch Allocation",allocation_id,f"Allocated STX-{int(issue.id)} to FB-{int(batch.id)} using FEFO"); st.success(f"Allocation BA-{allocation_id} recorded. Inventory quantity was not changed."); st.rerun()
    allocations=pd.read_sql_query("""SELECT a.id,a.created_at,x.id tank_transaction_id,b.batch_number,p.name product,a.allocated_liters,a.allocation_method,a.exception_reason,a.created_by FROM batch_issue_allocations a JOIN fuel_batches b ON b.id=a.batch_id JOIN products p ON p.id=b.product_id JOIN tank_transactions x ON x.id=a.tank_transaction_id ORDER BY a.id DESC""",conn)
    with history: st.dataframe(allocations,use_container_width=True,hide_index=True,height=450)
    with configuration:
        if role in {"ADMIN","INVENTORY_MANAGER"}:
            with st.form("aging_settings"):
                warning=st.number_input("Expiry warning days",min_value=1,max_value=365,value=int(settings.expiry_warning_days)); critical=st.number_input("Critical warning days",min_value=1,max_value=180,value=int(settings.critical_warning_days)); save=st.form_submit_button("Save aging limits",type="primary")
            if save:
                if critical>=warning: st.error("Critical days must be lower than warning days.")
                else:
                    c=conn.cursor(); c.execute("UPDATE batch_aging_settings SET expiry_warning_days=%s,critical_warning_days=%s,updated_by=%s,updated_at=CURRENT_TIMESTAMP WHERE id=1",(warning,critical,user)); conn.commit(); record_event(conn,"UPDATE_AGING_LIMITS","Batch Aging","Configuration",1,f"Warning {warning} days; critical {critical} days"); st.success("Aging limits saved."); st.rerun()
        company=st.session_state.get("company_profile",{}).get("company_name","Company"); st.download_button("Download batch aging & FEFO report",_report(position,allocations,settings,company),"batch_aging_fefo_report.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
