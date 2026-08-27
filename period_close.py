from datetime import date, datetime
import io
import json

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from approval_workflow import submit_approval_request
from audit import record_event
from ui import page_header
from valuation import calculate_valuation


def ensure_period_close_schema(conn):
    cursor = conn.cursor()
    try:
        cursor.execute("""CREATE TABLE IF NOT EXISTS inventory_periods(
            id BIGSERIAL PRIMARY KEY, period_name TEXT NOT NULL, start_date DATE NOT NULL,
            end_date DATE NOT NULL, status TEXT NOT NULL DEFAULT 'OPEN'
                CHECK(status IN ('OPEN','UNDER_REVIEW','CLOSED','REOPEN_REQUESTED')),
            close_request_id BIGINT, reopen_request_id BIGINT, close_notes TEXT,
            created_by TEXT NOT NULL, created_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            submitted_by TEXT, submitted_at TIMESTAMPTZ, closed_by TEXT, closed_at TIMESTAMPTZ,
            reopened_by TEXT, reopened_at TIMESTAMPTZ,
            closing_quantity REAL, closing_value REAL, currency TEXT,
            UNIQUE(start_date,end_date), CHECK(end_date>=start_date))""")
        cursor.execute("""CREATE TABLE IF NOT EXISTS period_close_checks(
            id BIGSERIAL PRIMARY KEY, period_id BIGINT NOT NULL REFERENCES inventory_periods(id) ON DELETE CASCADE,
            check_key TEXT NOT NULL, check_name TEXT NOT NULL, result_status TEXT NOT NULL,
            actual_value REAL, detail TEXT, checked_by TEXT NOT NULL,
            checked_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP)""")
        cursor.execute("""CREATE TABLE IF NOT EXISTS inventory_valuation_snapshots(
            id BIGSERIAL PRIMARY KEY, period_id BIGINT NOT NULL REFERENCES inventory_periods(id) ON DELETE CASCADE,
            asset_type TEXT NOT NULL, depot TEXT, asset TEXT NOT NULL, product TEXT,
            quantity_liters REAL NOT NULL, unit_cost REAL NOT NULL, inventory_value REAL NOT NULL,
            captured_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(period_id,asset_type,asset))""")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_inventory_periods_status ON inventory_periods(status,end_date)")
        cursor.execute("ALTER TABLE inventory_periods ADD COLUMN IF NOT EXISTS currency TEXT")
        conn.commit()
    except Exception:
        conn.rollback(); raise


def assert_period_open(conn, movement_date):
    day = pd.to_datetime(movement_date).date()
    cursor = conn.cursor()
    cursor.execute("""SELECT id,period_name FROM inventory_periods
        WHERE status='CLOSED' AND %s BETWEEN start_date AND end_date ORDER BY id DESC LIMIT 1""", (day,))
    row = cursor.fetchone()
    if row:
        raise ValueError(f"{row[1]} is closed. Transactions dated {day:%d %b %Y} cannot be added or changed. Submit a period reopening request first.")


def _safe_count(conn, query, params=None):
    cursor = conn.cursor()
    try:
        cursor.execute(query, params or ()); return int(cursor.fetchone()[0] or 0)
    except Exception:
        conn.rollback(); return 0


def run_close_checks(conn, period_id, user, save=True):
    cursor = conn.cursor(); cursor.execute("SELECT start_date,end_date FROM inventory_periods WHERE id=%s",(period_id,)); row=cursor.fetchone()
    if not row: raise ValueError("The selected accounting period no longer exists.")
    start_date,end_date=row
    position,ledger,claims,commitments=calculate_valuation(conn,end_date)
    pending_approvals=_safe_count(conn,"""SELECT COUNT(*) FROM approval_requests WHERE status='PENDING'
        AND request_kind NOT IN ('PERIOD_CLOSE','PERIOD_REOPEN') AND requested_at::date<=%s""",(end_date,))
    pending_counts=_safe_count(conn,"SELECT COUNT(*) FROM stock_reconciliations WHERE status='PENDING' AND reading_at::date<=%s",(end_date,))
    pending_tank_counts=_safe_count(conn,"SELECT COUNT(*) FROM tank_cycle_counts WHERE status='PENDING' AND counted_at::date<=%s",(end_date,))
    pending_changes=_safe_count(conn,"SELECT COUNT(*) FROM transaction_change_requests WHERE status='PENDING' AND requested_at::date<=%s",(end_date,))
    open_transit=_safe_count(conn,"SELECT COUNT(*) FROM inventory_transfers WHERE status='IN_TRANSIT' AND dispatched_at::date<=%s",(end_date,))
    costing_pending=_safe_count(conn,"SELECT COUNT(*) FROM supplier_receipt_invoices WHERE status IN ('MATCHED','EXCEPTION') AND invoice_date<=%s",(end_date,))
    expired_released=_safe_count(conn,"SELECT COUNT(*) FROM fuel_batches WHERE status='RELEASED' AND expiry_date<=%s",(end_date,))
    expired_calibration=_safe_count(conn,"SELECT COUNT(*) FROM (SELECT DISTINCT ON (tank_id) tank_id,next_due_date FROM tank_calibrations WHERE calibration_date<=%s ORDER BY tank_id,calibration_date DESC,id DESC) q WHERE next_due_date<=%s",(end_date,end_date))
    open_incidents=_safe_count(conn,"SELECT COUNT(*) FROM inventory_incidents WHERE status='OPEN' AND occurred_at::date<=%s",(end_date,))
    negative=int((position["Quantity (L)"] < -0.005).sum()) if not position.empty else 0
    fallback=int(ledger["Cost Source"].isin(["System default cost"]).sum()) if not ledger.empty else 0
    open_claims=int((~claims["status"].isin(["CLOSED","REJECTED"])).sum()) if not claims.empty else 0
    rows=[
        ("PENDING_APPROVALS","Pending approvals",pending_approvals,"BLOCKER" if pending_approvals else "PASS","Complete or reject requests dated on or before period end."),
        ("PENDING_COUNTS","Pending physical counts",pending_counts,"BLOCKER" if pending_counts else "PASS","Complete inventory reconciliations before closing."),
        ("PENDING_TANK_COUNTS","Pending tank cycle counts",pending_tank_counts,"BLOCKER" if pending_tank_counts else "PASS","Review every tank count dated on or before period end."),
        ("PENDING_CHANGES","Pending transaction changes",pending_changes,"BLOCKER" if pending_changes else "PASS","Complete correction and reversal requests before closing."),
        ("OPEN_TRANSIT","Stock still in transit",open_transit,"BLOCKER" if open_transit else "PASS","Receive or investigate all transfers dispatched on or before period end."),
        ("PENDING_COSTING","Unapproved receipt costing",costing_pending,"BLOCKER" if costing_pending else "PASS","Approve or reject invoice matches and landed costs before closing."),
        ("EXPIRED_RELEASED_BATCH","Expired batches still released",expired_released,"BLOCKER" if expired_released else "PASS","Block or close expired fuel batches."),
        ("EXPIRED_CALIBRATION","Expired tank calibration",expired_calibration,"BLOCKER" if expired_calibration else "PASS","Renew expired measurement certificates."),
        ("OPEN_INCIDENTS","Open inventory incidents",open_incidents,"WARNING" if open_incidents else "PASS","Review unresolved loss, gain, spill, return or contamination incidents."),
        ("NEGATIVE_STOCK","Negative inventory positions",negative,"BLOCKER" if negative else "PASS","Investigate every negative tank or truck position."),
        ("FALLBACK_COST","Movements using system default cost",fallback,"WARNING" if fallback else "PASS","Review missing receipt prices or approve a product cost policy."),
        ("OPEN_CLAIMS","Open supplier claims",open_claims,"WARNING" if open_claims else "PASS","Open claims do not block close but remain a financial exposure."),
    ]
    if save:
        cursor=conn.cursor(); cursor.execute("DELETE FROM period_close_checks WHERE period_id=%s",(period_id,))
        for key,name,value,status,detail in rows:
            cursor.execute("""INSERT INTO period_close_checks(period_id,check_key,check_name,result_status,actual_value,detail,checked_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s)""",(period_id,key,name,status,value,detail,user))
        conn.commit(); record_event(conn,"RUN_CLOSE_CHECKS","Period Closing","Inventory Period",period_id,"Executed month-end control checks")
    return pd.DataFrame(rows,columns=["Check Key","Control","Actual","Status","Required Action"]),position,ledger,claims,commitments


def submit_period_close(conn, period_id, notes, user):
    checks,position,_,_,_=run_close_checks(conn,period_id,user,True)
    blockers=checks[checks["Status"]=="BLOCKER"]
    if not blockers.empty: raise ValueError("The period cannot be submitted while blocker checks remain. Resolve the red controls and run checks again.")
    cursor=conn.cursor(); cursor.execute("SELECT period_name,end_date,status FROM inventory_periods WHERE id=%s FOR UPDATE",(period_id,)); row=cursor.fetchone()
    if not row or row[2]!="OPEN": raise ValueError("Only an open period can be submitted for closing.")
    total_qty=float(position["Quantity (L)"].sum()) if not position.empty else 0.0; total_value=float(position["Inventory Value"].sum()) if not position.empty else 0.0
    request_id=submit_approval_request(conn,"PERIOD_CLOSE",f"Close inventory period · {row[0]}",{"period_id":int(period_id),"notes":notes.strip()},user,total_qty,total_value)
    cursor=conn.cursor(); cursor.execute("""UPDATE inventory_periods SET status='UNDER_REVIEW',close_request_id=%s,close_notes=%s,
        submitted_by=%s,submitted_at=CURRENT_TIMESTAMP,closing_quantity=%s,closing_value=%s WHERE id=%s""",(request_id,notes.strip(),user,total_qty,total_value,period_id)); conn.commit()
    return request_id


def execute_period_close(conn, period_id, reviewer, request_id):
    checks,position,_,_,_=run_close_checks(conn,period_id,reviewer,True)
    if (checks["Status"]=="BLOCKER").any(): raise ValueError("Closing checks now contain blockers. The close was not posted.")
    cursor=conn.cursor(); cursor.execute("SELECT period_name,status FROM inventory_periods WHERE id=%s FOR UPDATE",(period_id,)); row=cursor.fetchone()
    if not row or row[1]!="UNDER_REVIEW": raise ValueError("This period is no longer waiting for close approval.")
    cursor.execute("DELETE FROM inventory_valuation_snapshots WHERE period_id=%s",(period_id,))
    for item in position.to_dict("records"):
        cursor.execute("""INSERT INTO inventory_valuation_snapshots(period_id,asset_type,depot,asset,product,quantity_liters,unit_cost,inventory_value)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",(period_id,item["Asset Type"],item.get("Depot"),item["Asset"],item.get("Product"),float(item["Quantity (L)"]),float(item["Unit Cost"]),float(item["Inventory Value"])))
    cursor.execute("""UPDATE inventory_periods SET status='CLOSED',closed_by=%s,closed_at=CURRENT_TIMESTAMP,
        close_request_id=%s WHERE id=%s""",(reviewer,request_id,period_id)); conn.commit()
    record_event(conn,"CLOSE_PERIOD","Period Closing","Inventory Period",period_id,f"Closed {row[0]} and locked its valuation snapshot")
    return f"PC-{period_id}"


def submit_period_reopen(conn, period_id, reason, reference, user):
    cursor=conn.cursor(); cursor.execute("SELECT period_name,status FROM inventory_periods WHERE id=%s FOR UPDATE",(period_id,)); row=cursor.fetchone()
    if not row or row[1]!="CLOSED": raise ValueError("Only a closed period can be submitted for reopening.")
    request_id=submit_approval_request(conn,"PERIOD_REOPEN",f"Reopen inventory period · {row[0]}",{"period_id":int(period_id),"reason":reason.strip(),"reference":reference.strip()},user)
    cursor=conn.cursor(); cursor.execute("UPDATE inventory_periods SET status='REOPEN_REQUESTED',reopen_request_id=%s WHERE id=%s",(request_id,period_id)); conn.commit(); return request_id


def execute_period_reopen(conn, period_id, reviewer, request_id):
    cursor=conn.cursor(); cursor.execute("SELECT period_name,status FROM inventory_periods WHERE id=%s FOR UPDATE",(period_id,)); row=cursor.fetchone()
    if not row or row[1]!="REOPEN_REQUESTED": raise ValueError("This period is no longer waiting for reopening approval.")
    cursor.execute("""UPDATE inventory_periods SET status='OPEN',reopened_by=%s,reopened_at=CURRENT_TIMESTAMP,
        reopen_request_id=%s,closed_by=NULL,closed_at=NULL WHERE id=%s""",(reviewer,request_id,period_id)); conn.commit()
    record_event(conn,"REOPEN_PERIOD","Period Closing","Inventory Period",period_id,f"Reopened {row[0]} following approval")
    return f"PC-{period_id}"


def _write_table(ws,frame,start_row,name,profile):
    primary=str(profile.get("primary_color","#9E1B1B")).replace("#","")
    headers=list(frame.columns)
    for col,header in enumerate(headers,1):
        cell=ws.cell(start_row,col,str(header)); cell.fill=PatternFill("solid",fgColor=primary); cell.font=Font(bold=True,color="FFFFFF"); cell.alignment=Alignment(wrap_text=True)
    if frame.empty: ws.cell(start_row+1,1,"No records available"); return
    for row in frame.itertuples(index=False,name=None): ws.append([value.to_pydatetime().replace(tzinfo=None) if isinstance(value,pd.Timestamp) else value for value in row])
    end=start_row+len(frame); table=Table(displayName=name,ref=f"A{start_row}:{get_column_letter(max(len(headers),1))}{end}"); table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True); ws.add_table(table); ws.freeze_panes=f"A{start_row+1}"
    for i,header in enumerate(headers,1): ws.column_dimensions[get_column_letter(i)].width=28 if header in ("Asset","Control","Required Action") else 18


def build_period_close_report(period,checks,snapshot,profile):
    wb=Workbook(); cover=wb.active; cover.title="Closing Certificate"; primary=str(profile.get("primary_color","#9E1B1B")).replace("#",""); secondary=str(profile.get("secondary_color","#172033")).replace("#","")
    cover.sheet_view.showGridLines=False; cover.merge_cells("A1:H2"); cover["A1"]=f"{profile.get('company_name','Company')} | Inventory Closing Certificate"; cover["A1"].fill=PatternFill("solid",fgColor=secondary); cover["A1"].font=Font(size=20,bold=True,color="FFFFFF"); cover["A1"].alignment=Alignment(vertical="center")
    cover.merge_cells("A3:H3"); cover["A3"]=f"{period['period_name']} · {period['start_date']:%d %b %Y} to {period['end_date']:%d %b %Y}"; cover["A3"].fill=PatternFill("solid",fgColor=primary); cover["A3"].font=Font(color="FFFFFF",bold=True)
    closed_at=period.get("closed_at")
    if closed_at is not None and not pd.isna(closed_at): closed_at=pd.to_datetime(closed_at).strftime("%d %b %Y %H:%M")
    else: closed_at="—"
    fields=[("Status",period["status"]),("Closing quantity",period.get("closing_quantity") or 0),("Closing value",period.get("closing_value") or 0),("Currency",period.get("currency") or profile.get("currency","AED")),("Submitted by",period.get("submitted_by") or "—"),("Approved by",period.get("closed_by") or "—"),("Closed at",closed_at),("Notes",period.get("close_notes") or "—")]
    for idx,(label,value) in enumerate(fields,5): cover.cell(idx,1,label).font=Font(bold=True,color="475467"); cover.merge_cells(start_row=idx,start_column=2,end_row=idx,end_column=8); cover.cell(idx,2,value)
    cover.column_dimensions["A"].width=24
    check_ws=wb.create_sheet("Close Checks"); _write_table(check_ws,checks,1,"CloseChecks",profile)
    stock_ws=wb.create_sheet("Locked Valuation"); _write_table(stock_ws,snapshot,1,"LockedValuation",profile)
    output=io.BytesIO(); wb.save(output); return output.getvalue()


def render_period_close(conn):
    ensure_period_close_schema(conn); profile=st.session_state.get("company_profile",{}); user=st.session_state.get("user","System")
    editable=st.session_state.get("role") in ("ADMIN","INVENTORY_MANAGER")
    page_header("Month-End Inventory Closing","Check, approve, lock and certify each inventory accounting period.")
    periods=pd.read_sql_query("SELECT * FROM inventory_periods ORDER BY end_date DESC,id DESC",conn)
    open_count=int(periods["status"].isin(["OPEN","UNDER_REVIEW"]).sum()) if not periods.empty else 0; closed_count=int((periods["status"]=="CLOSED").sum()) if not periods.empty else 0
    a,b,c=st.columns(3); a.metric("Open / review",open_count); b.metric("Closed periods",closed_count); c.metric("Latest closed",periods.loc[periods["status"]=="CLOSED","period_name"].iloc[0] if closed_count else "—")
    setup,control,history=st.tabs(["Create period","Close control","History & certificate"])
    with setup:
        with st.form("create_period"):
            name=st.text_input("Period name",value=date.today().strftime("%B %Y")); x,y=st.columns(2); start=x.date_input("Start date",date.today().replace(day=1)); end=y.date_input("End date",date.today()); notes=st.text_area("Preparation notes"); create=st.form_submit_button("Create inventory period",type="primary",disabled=not editable)
        if not editable: st.info("Inventory Manager permission is required to create, submit or reopen a period. You can still review history and download certificates.")
        if create:
            if not name.strip() or end<start: st.error("Enter a period name and valid start/end dates.")
            else:
                try:
                    cursor=conn.cursor(); cursor.execute("""INSERT INTO inventory_periods(period_name,start_date,end_date,close_notes,created_by,currency)
                        VALUES (%s,%s,%s,%s,%s,%s) RETURNING id""",(name.strip(),start,end,notes.strip() or None,user,profile.get("currency","AED"))); period_id=cursor.fetchone()[0]; conn.commit(); record_event(conn,"CREATE_PERIOD","Period Closing","Inventory Period",period_id,f"Created {name.strip()}"); st.success(f"PC-{period_id} created."); st.rerun()
                except Exception as error: conn.rollback(); st.error(str(error))
    with control:
        active=periods[periods["status"].isin(["OPEN","UNDER_REVIEW","CLOSED","REOPEN_REQUESTED"])] if not periods.empty else periods
        if active.empty: st.info("Create the first inventory period to begin closing controls.")
        else:
            labels={f"PC-{r.id} · {r.period_name} · {r.status}":int(r.id) for r in active.itertuples()}; selected=st.selectbox("Period",list(labels)); period_id=labels[selected]; period=active[active["id"]==period_id].iloc[0]
            if period["status"]=="OPEN":
                if st.button("Run closing checks",type="primary",disabled=not editable):
                    run_close_checks(conn,period_id,user,True); st.success("Closing checks completed."); st.rerun()
                checks=pd.read_sql_query("SELECT check_name AS \"Control\",actual_value AS \"Actual\",result_status AS \"Status\",detail AS \"Required Action\",checked_at AS \"Checked At\" FROM period_close_checks WHERE period_id=%s ORDER BY id",conn,params=[period_id]); st.dataframe(checks,use_container_width=True,hide_index=True)
                with st.form("submit_close"):
                    close_notes=st.text_area("Closing notes and management explanation",value=period.get("close_notes") or ""); confirm=st.checkbox("I confirm the period is complete and ready for independent approval."); submit=st.form_submit_button("Submit period close for approval",type="primary",disabled=not editable)
                if submit:
                    if not confirm or len(close_notes.strip())<5: st.error("Confirm the close and enter clear closing notes.")
                    else:
                        try: request_id=submit_period_close(conn,period_id,close_notes,user); st.success(f"AP-{request_id} submitted. The period remains unlocked until another user approves it."); st.rerun()
                        except Exception as error: st.error(str(error))
            elif period["status"]=="UNDER_REVIEW": st.info(f"This period is awaiting approval as AP-{int(period['close_request_id'])}.")
            elif period["status"]=="REOPEN_REQUESTED": st.warning(f"Reopening is awaiting approval as AP-{int(period['reopen_request_id'])}.")
            else:
                st.success("This period is closed. Its dated transactions are locked and its valuation snapshot is preserved.")
                with st.form("reopen_period"):
                    reason=st.text_area("Reason for reopening"); reference=st.text_input("Authorization / supporting reference"); confirm=st.checkbox("I understand reopening removes the transaction lock for this period."); reopen=st.form_submit_button("Submit reopening request",disabled=not editable)
                if reopen:
                    if not confirm or len(reason.strip())<5 or not reference.strip(): st.error("Confirm the action and enter a clear reason and supporting reference.")
                    else:
                        try: request_id=submit_period_reopen(conn,period_id,reason,reference,user); st.success(f"AP-{request_id} submitted. The period remains closed until approval."); st.rerun()
                        except Exception as error: st.error(str(error))
    with history:
        st.dataframe(periods,use_container_width=True,hide_index=True,height=330)
        closed=periods[periods["status"]=="CLOSED"] if not periods.empty else periods
        if not closed.empty:
            choices={f"PC-{r.id} · {r.period_name}":int(r.id) for r in closed.itertuples()}; chosen=st.selectbox("Closing certificate",list(choices)); period_id=choices[chosen]; period=closed[closed["id"]==period_id].iloc[0].to_dict()
            checks=pd.read_sql_query("SELECT check_name AS \"Control\",actual_value AS \"Actual\",result_status AS \"Status\",detail AS \"Required Action\",checked_at AS \"Checked At\" FROM period_close_checks WHERE period_id=%s ORDER BY id",conn,params=[period_id]); snapshot=pd.read_sql_query("SELECT asset_type AS \"Asset Type\",depot AS \"Depot\",asset AS \"Asset\",product AS \"Product\",quantity_liters AS \"Quantity (L)\",unit_cost AS \"Unit Cost\",inventory_value AS \"Inventory Value\" FROM inventory_valuation_snapshots WHERE period_id=%s ORDER BY asset_type,depot,asset",conn,params=[period_id])
            report=build_period_close_report(period,checks,snapshot,profile); st.download_button("Download closing certificate",report,f"inventory_close_{period['end_date']:%Y%m%d}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
