from io import BytesIO
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from ui import page_header

CHECKS={
 "Negative truck balances":("CRITICAL","""SELECT tr.id,CONCAT(tr.emirate,' ',tr.plate_code,' ',tr.plate_number) asset,SUM(CASE WHEN x.type='IN' THEN x.liters ELSE -x.liters END) value FROM trucks tr JOIN transactions x ON x.truck_id=tr.id WHERE COALESCE(x.record_status,'POSTED')='POSTED' GROUP BY tr.id HAVING SUM(CASE WHEN x.type='IN' THEN x.liters ELSE -x.liters END)<-0.005"""),
 "Negative tank balances":("CRITICAL","""SELECT t.id,CONCAT(d.code,' / ',t.code) asset,SUM(CASE WHEN x.type='IN' THEN x.liters ELSE -x.liters END) value FROM storage_tanks t JOIN depots d ON d.id=t.depot_id JOIN tank_transactions x ON x.tank_id=t.id WHERE COALESCE(x.record_status,'POSTED')='POSTED' GROUP BY t.id,d.code HAVING SUM(CASE WHEN x.type='IN' THEN x.liters ELSE -x.liters END)<-0.005"""),
 "Tank above safe capacity":("CRITICAL","""SELECT t.id,CONCAT(d.code,' / ',t.code) asset,SUM(CASE WHEN x.type='IN' THEN x.liters ELSE -x.liters END)-t.safe_capacity_liters value FROM storage_tanks t JOIN depots d ON d.id=t.depot_id JOIN tank_transactions x ON x.tank_id=t.id WHERE COALESCE(x.record_status,'POSTED')='POSTED' GROUP BY t.id,d.code HAVING SUM(CASE WHEN x.type='IN' THEN x.liters ELSE -x.liters END)>t.safe_capacity_liters+0.005"""),
 "Broken truck transfer links":("CRITICAL","""SELECT a.id,CONCAT('TX-',a.id) asset,a.transfer_partner_id value FROM transactions a LEFT JOIN transactions b ON b.id=a.transfer_partner_id WHERE a.transfer_partner_id IS NOT NULL AND b.id IS NULL"""),
 "Expired batches still released":("CRITICAL","""SELECT id,batch_number asset,(CURRENT_DATE-expiry_date) value FROM fuel_batches WHERE status='RELEASED' AND expiry_date<CURRENT_DATE"""),
 "Supplier receipts without batch":("WARNING","""SELECT id,CONCAT('STX-',id) asset,COALESCE(accepted_liters,liters) value FROM tank_transactions WHERE movement_category='SUPPLIER_RECEIPT' AND batch_id IS NULL"""),
 "Storage OUT without batch allocation":("WARNING","""SELECT x.id,CONCAT('STX-',x.id) asset,x.liters value FROM tank_transactions x LEFT JOIN batch_issue_allocations a ON a.tank_transaction_id=x.id WHERE x.type='OUT' AND a.id IS NULL"""),
 "Pending approvals overdue":("WARNING","""SELECT id,CONCAT('AP-',id) asset,EXTRACT(EPOCH FROM (CURRENT_TIMESTAMP-due_at))/3600 value FROM approval_requests WHERE status='PENDING' AND due_at<CURRENT_TIMESTAMP"""),
}

def _run(conn):
 rows=[]
 for name,(severity,sql) in CHECKS.items():
  try:
   data=pd.read_sql_query(sql,conn)
   for item in data.itertuples(index=False): rows.append({"check":name,"severity":severity,"record_id":item.id,"asset":item.asset,"value":item.value})
  except Exception:
   conn.rollback(); rows.append({"check":name,"severity":"ERROR","record_id":None,"asset":"Check could not run","value":None})
 return pd.DataFrame(rows,columns=["check","severity","record_id","asset","value"])

def _xlsx(data,company):
 wb=Workbook(); ws=wb.active; ws.title="Integrity Exceptions"; ws.append([f"{company} | Inventory Integrity Health Check"]); ws.merge_cells("A1:E1"); ws["A1"].fill=PatternFill("solid",fgColor="9E1B1B"); ws["A1"].font=Font(color="FFFFFF",bold=True,size=16); ws.append(list(data.columns))
 for row in data.where(pd.notna(data),None).itertuples(index=False,name=None): ws.append(list(row))
 for c in ws[2]: c.fill=PatternFill("solid",fgColor="111827"); c.font=Font(color="FFFFFF",bold=True)
 ws.freeze_panes="A3"; ws.auto_filter.ref=ws.dimensions
 for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(max(14,max(len(str(x.value or '')) for x in col)+2),38)
 out=BytesIO(); wb.save(out); return out.getvalue()

def render_inventory_health(conn):
 page_header("Inventory Health Centre","Run read-only integrity checks across stock, transfers, batches, approvals and traceability.")
 data=_run(conn); critical=int(data.severity.eq("CRITICAL").sum()); warning=int(data.severity.eq("WARNING").sum()); errors=int(data.severity.eq("ERROR").sum())
 a,b,c,d=st.columns(4); a.metric("Checks",len(CHECKS)); b.metric("Critical exceptions",critical); c.metric("Warnings",warning); d.metric("Check errors",errors)
 if data.empty: st.success("All configured inventory integrity checks passed.")
 else:
  severity=st.multiselect("Severity",["CRITICAL","WARNING","ERROR"],default=["CRITICAL","WARNING","ERROR"]); st.dataframe(data[data.severity.isin(severity)],use_container_width=True,hide_index=True,height=500)
  st.warning("This page is diagnostic only. Correct exceptions through the controlled transaction, reconciliation, quality or approval pages.")
 company=st.session_state.get("company_profile",{}).get("company_name","Company"); st.download_button("Download integrity exception report",_xlsx(data,company),"inventory_integrity_exceptions.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
