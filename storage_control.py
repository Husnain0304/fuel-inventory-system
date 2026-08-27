from datetime import date
from io import BytesIO
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill
from audit import record_event
from period_close import assert_period_open
from ui import page_header

EDIT={"ADMIN","INVENTORY_MANAGER","STOREKEEPER","OPERATOR"}; REVIEW={"ADMIN","INVENTORY_MANAGER","APPROVER"}

def ensure_storage_control_schema(conn):
 c=conn.cursor()
 try:
  c.execute("""CREATE TABLE IF NOT EXISTS tank_calibrations(id BIGSERIAL PRIMARY KEY,tank_id INTEGER NOT NULL REFERENCES storage_tanks(id),certificate_number TEXT NOT NULL,calibration_date DATE NOT NULL,next_due_date DATE NOT NULL,calibration_company TEXT,maximum_error_percent REAL,document_reference TEXT,status TEXT DEFAULT 'VALID',notes TEXT,created_by TEXT,created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,UNIQUE(tank_id,certificate_number))""")
  c.execute("""CREATE TABLE IF NOT EXISTS tank_cycle_counts(id BIGSERIAL PRIMARY KEY,tank_id INTEGER NOT NULL REFERENCES storage_tanks(id),counted_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,system_liters REAL NOT NULL,physical_liters REAL NOT NULL CHECK(physical_liters>=0),variance_liters REAL NOT NULL,variance_percent REAL NOT NULL,measurement_method TEXT,reference TEXT,reason TEXT,status TEXT DEFAULT 'PENDING',counted_by TEXT,reviewed_by TEXT,reviewed_at TIMESTAMPTZ,review_comment TEXT,created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP)""")
  c.execute("""CREATE TABLE IF NOT EXISTS inventory_incidents(id BIGSERIAL PRIMARY KEY,incident_number TEXT UNIQUE,occurred_at TIMESTAMPTZ NOT NULL,asset_type TEXT NOT NULL,asset_reference TEXT NOT NULL,incident_type TEXT NOT NULL,estimated_liters REAL NOT NULL DEFAULT 0,product_id INTEGER REFERENCES products(id),description TEXT NOT NULL,immediate_action TEXT,status TEXT DEFAULT 'OPEN',root_cause TEXT,corrective_action TEXT,owner TEXT,created_by TEXT,created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,closed_by TEXT,closed_at TIMESTAMPTZ)""")
  c.execute("ALTER TABLE tank_cycle_counts ADD COLUMN IF NOT EXISTS adjustment_transaction_id BIGINT REFERENCES tank_transactions(id)")
  c.execute("ALTER TABLE tank_cycle_counts ADD COLUMN IF NOT EXISTS posted_by TEXT")
  c.execute("ALTER TABLE tank_cycle_counts ADD COLUMN IF NOT EXISTS posted_at TIMESTAMPTZ")
  c.execute("ALTER TABLE inventory_incidents ADD COLUMN IF NOT EXISTS closure_reference TEXT")
  conn.commit()
 except Exception: conn.rollback(); raise

def _tanks(conn): return pd.read_sql_query("""SELECT t.id,CONCAT(d.code,' / ',t.code,' · ',p.name) label,t.product_id,COALESCE(SUM(CASE WHEN x.type='IN' THEN x.liters ELSE -x.liters END),0) balance FROM storage_tanks t JOIN depots d ON d.id=t.depot_id JOIN products p ON p.id=t.product_id LEFT JOIN tank_transactions x ON x.tank_id=t.id AND COALESCE(x.record_status,'POSTED')='POSTED' GROUP BY t.id,d.code,p.name ORDER BY d.code,t.code""",conn)

def _export(parts,company):
 wb=Workbook(); wb.remove(wb.active)
 for name,data in parts.items():
  ws=wb.create_sheet(name); ws.append(list(data.columns))
  for row in data.where(pd.notna(data),None).itertuples(index=False,name=None): ws.append([str(v) if isinstance(v,pd.Timestamp) else v for v in row])
  for c in ws[1]: c.fill=PatternFill("solid",fgColor="111827"); c.font=Font(color="FFFFFF",bold=True)
  ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
  for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(max(14,max(len(str(x.value or '')) for x in col)+2),36)
 out=BytesIO(); wb.save(out); return out.getvalue()

def render_storage_control(conn):
 role=st.session_state.get("role","VIEWER"); user=st.session_state.get("user","System"); tanks=_tanks(conn); tm={r.label:r for r in tanks.itertuples()}
 counts=pd.read_sql_query("""SELECT c.id,c.tank_id,t.product_id,c.counted_at,t.code tank,d.code depot,c.system_liters,c.physical_liters,c.variance_liters,c.variance_percent,c.measurement_method,c.reference,c.reason,c.status,c.counted_by,c.reviewed_by,c.adjustment_transaction_id,c.posted_by,c.posted_at FROM tank_cycle_counts c JOIN storage_tanks t ON t.id=c.tank_id JOIN depots d ON d.id=t.depot_id ORDER BY c.id DESC""",conn)
 calibrations=pd.read_sql_query("""SELECT c.id,d.code depot,t.code tank,c.certificate_number,c.calibration_date,c.next_due_date,c.calibration_company,c.maximum_error_percent,CASE WHEN c.next_due_date<CURRENT_DATE THEN 'EXPIRED' WHEN c.next_due_date<=CURRENT_DATE+30 THEN 'DUE SOON' ELSE c.status END status,c.document_reference,c.created_by FROM tank_calibrations c JOIN storage_tanks t ON t.id=c.tank_id JOIN depots d ON d.id=t.depot_id ORDER BY c.next_due_date""",conn)
 incidents=pd.read_sql_query("SELECT * FROM inventory_incidents ORDER BY id DESC",conn)
 page_header("Measurement & Loss Control","Control tank measurements, calibration validity, physical counts and inventory incidents.")
 a,b,c,d=st.columns(4); a.metric("Pending tank counts",int(counts.status.eq('PENDING').sum()) if not counts.empty else 0); b.metric("Calibration overdue",int(calibrations.status.eq('EXPIRED').sum()) if not calibrations.empty else 0); c.metric("Open incidents",int(incidents.status.eq('OPEN').sum()) if not incidents.empty else 0); d.metric("Open estimated variance",f"{incidents.loc[incidents.status.eq('OPEN'),'estimated_liters'].sum():,.0f} L" if not incidents.empty else "0 L")
 count_tab,cal_tab,incident_tab,review_tab,report_tab=st.tabs(["Tank count","Calibration","Loss / gain incident","Review & close","Control report"])
 with count_tab:
  st.dataframe(counts,use_container_width=True,hide_index=True,height=280)
  if role in EDIT and tm:
   with st.form("tank_count"):
    label=st.selectbox("Tank",list(tm)); row=tm[label]; st.info(f"Current system quantity: {float(row.balance):,.2f} L"); physical=st.number_input("Physical measured quantity",min_value=0.0); method=st.selectbox("Measurement method",["Dip reading","Automatic tank gauge","Meter totalizer","Other"]); reference=st.text_input("Count sheet / evidence reference"); reason=st.text_area("Initial variance explanation"); submit=st.form_submit_button("Submit physical count",type="primary")
   if submit:
    variance=physical-float(row.balance); percent=variance/float(row.balance)*100 if abs(float(row.balance))>.005 else 0; c=conn.cursor(); c.execute("INSERT INTO tank_cycle_counts(tank_id,system_liters,physical_liters,variance_liters,variance_percent,measurement_method,reference,reason,counted_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",(int(row.id),float(row.balance),physical,variance,percent,method,reference.strip() or None,reason.strip() or None,user)); cid=c.fetchone()[0]; conn.commit(); record_event(conn,"TANK_CYCLE_COUNT","Measurement Control","Tank Cycle Count",cid,f"Tank count variance {variance:+,.2f} L",severity="WARNING" if abs(variance)>.005 else "INFO"); st.success(f"TC-{cid} submitted for review. No adjustment was posted."); st.rerun()
 with cal_tab:
  st.dataframe(calibrations,use_container_width=True,hide_index=True)
  if role in EDIT and tm:
   with st.form("calibration"):
    label=st.selectbox("Tank",list(tm),key="cal_tank"); certificate=st.text_input("Certificate number"); x,y=st.columns(2); cal_date=x.date_input("Calibration date",date.today()); due=y.date_input("Next due date"); company=st.text_input("Calibration company"); error=st.number_input("Maximum error %",min_value=0.0); document=st.text_input("Evidence document reference"); submit=st.form_submit_button("Register calibration",type="primary")
   if submit:
    if due<=cal_date or len(certificate.strip())<2: st.error("Enter a certificate and a future due date.")
    else:
     c=conn.cursor(); c.execute("INSERT INTO tank_calibrations(tank_id,certificate_number,calibration_date,next_due_date,calibration_company,maximum_error_percent,document_reference,created_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",(int(tm[label].id),certificate.strip(),cal_date,due,company.strip() or None,error,document.strip() or None,user)); cid=c.fetchone()[0]; conn.commit(); record_event(conn,"REGISTER_CALIBRATION","Measurement Control","Tank Calibration",cid,certificate.strip()); st.success(f"CAL-{cid} registered."); st.rerun()
 with incident_tab:
  st.dataframe(incidents,use_container_width=True,hide_index=True,height=260)
  if role in EDIT:
   products=pd.read_sql_query("SELECT id,name FROM products WHERE active=TRUE",conn); pm=dict(zip(products.name,products.id))
   with st.form("incident"):
    x,y=st.columns(2); kind=x.selectbox("Incident type",["LOSS","GAIN","SPILL","CONTAMINATION","RETURN","METER_VARIANCE","OTHER"]); asset=y.text_input("Asset / location reference"); product=st.selectbox("Product",list(pm)); liters=st.number_input("Estimated quantity",min_value=0.0); description=st.text_area("What happened?"); action=st.text_area("Immediate containment action"); owner=st.text_input("Investigation owner"); submit=st.form_submit_button("Open incident",type="primary")
   if submit:
    if len(asset.strip())<2 or len(description.strip())<5: st.error("Enter the affected asset and a clear description.")
    else:
     c=conn.cursor(); c.execute("INSERT INTO inventory_incidents(occurred_at,asset_type,asset_reference,incident_type,estimated_liters,product_id,description,immediate_action,owner,created_by) VALUES (CURRENT_TIMESTAMP,'INVENTORY',%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",(asset.strip(),kind,liters,pm[product],description.strip(),action.strip() or None,owner.strip() or None,user)); iid=c.fetchone()[0]; c.execute("UPDATE inventory_incidents SET incident_number=%s WHERE id=%s",(f"INC-{iid}",iid)); conn.commit(); record_event(conn,"OPEN_INVENTORY_INCIDENT","Measurement Control","Inventory Incident",iid,f"{kind} estimated {liters:,.2f} L",severity="WARNING"); st.success(f"INC-{iid} opened. No inventory adjustment was posted."); st.rerun()
 with review_tab:
  pending=counts[counts.status.eq("PENDING")] if not counts.empty else counts
  if role in REVIEW and not pending.empty:
   cm={f"TC-{int(r.id)} · {r.depot}/{r.tank} · {float(r.variance_liters):+,.2f} L":int(r.id) for r in pending.itertuples()}; selected=st.selectbox("Pending count",list(cm)); decision=st.selectbox("Decision",["REVIEWED","REJECTED"]); comment=st.text_area("Review comment")
   if st.button("Save count review",type="primary"):
    if len(comment.strip())<5: st.error("Enter a review comment.")
    else:
     cid=cm[selected]; c=conn.cursor(); c.execute("UPDATE tank_cycle_counts SET status=%s,reviewed_by=%s,reviewed_at=CURRENT_TIMESTAMP,review_comment=%s WHERE id=%s",(decision,user,comment.strip(),cid)); conn.commit(); record_event(conn,"REVIEW_TANK_COUNT","Measurement Control","Tank Cycle Count",cid,decision); st.success("Review saved. Use controlled reconciliation for any adjustment."); st.rerun()
  elif pending.empty: st.success("No tank count is waiting for review.")
  st.divider(); st.subheader("Post reviewed tank adjustment")
  reviewed=counts[(counts.status.eq("REVIEWED")) & (counts.adjustment_transaction_id.isna())] if not counts.empty else counts
  if role not in REVIEW: st.info("An Inventory Manager or Administrator must post a reviewed adjustment.")
  elif reviewed.empty: st.success("No reviewed tank count is waiting for adjustment.")
  else:
   am={f"TC-{int(r.id)} · {r.depot}/{r.tank} · {float(r.variance_liters):+,.2f} L":r for r in reviewed.itertuples()}; choice=st.selectbox("Reviewed count",list(am),key="post_tank_count"); item=am[choice]; authorization=st.text_input("Adjustment authorization / supporting reference"); confirm=st.checkbox("I confirm the physical count evidence has been reviewed and this adjustment is required.")
   if st.button("Post controlled tank adjustment",type="primary",disabled=not confirm):
    if item.counted_by==user and role!="ADMIN": st.error("Separation of duties: the person who counted the tank cannot post its adjustment.")
    elif len(authorization.strip())<3: st.error("Enter an authorization or supporting reference.")
    else:
     try:
      assert_period_open(conn,date.today()); c=conn.cursor(); c.execute("SELECT status,adjustment_transaction_id FROM tank_cycle_counts WHERE id=%s FOR UPDATE",(int(item.id),)); current=c.fetchone()
      if current[0]!="REVIEWED" or current[1] is not None: raise ValueError("This count is no longer awaiting adjustment.")
      variance=float(item.variance_liters); tx=None
      if abs(variance)>.005:
       c.execute("""INSERT INTO tank_transactions(tank_id,movement_at,liters,type,movement_category,product_id,reference,notes,created_by)
        VALUES (%s,CURRENT_TIMESTAMP,%s,%s,'COUNT_ADJUSTMENT',%s,%s,%s,%s) RETURNING id""",(int(item.tank_id),abs(variance),'IN' if variance>0 else 'OUT',int(item.product_id),authorization.strip(),f"Physical count TC-{int(item.id)}",user)); tx=c.fetchone()[0]
      c.execute("UPDATE tank_cycle_counts SET status='POSTED',adjustment_transaction_id=%s,posted_by=%s,posted_at=CURRENT_TIMESTAMP WHERE id=%s",(tx,user,int(item.id))); conn.commit(); record_event(conn,"POST_TANK_COUNT_ADJUSTMENT","Measurement Control","Tank Cycle Count",int(item.id),f"Adjustment {variance:+,.2f} L; transaction {tx}",severity="WARNING" if abs(variance)>.005 else "INFO"); st.success("The reviewed count is closed and the controlled adjustment has been posted."); st.rerun()
     except Exception as error: conn.rollback(); st.error(str(error))
  st.divider(); st.subheader("Close inventory incident")
  open_incidents=incidents[incidents.status.eq("OPEN")] if not incidents.empty else incidents
  if role not in REVIEW: st.info("An authorized reviewer must close incidents.")
  elif open_incidents.empty: st.success("No inventory incident is awaiting closure.")
  else:
   im={f"{r.incident_number} · {r.incident_type} · {r.asset_reference}":r for r in open_incidents.itertuples()}; selected_incident=st.selectbox("Open incident",list(im)); incident=im[selected_incident]; root=st.text_area("Confirmed root cause"); corrective=st.text_area("Corrective / preventive action"); closure_ref=st.text_input("Closure evidence / authorization reference")
   if st.button("Close inventory incident",type="primary"):
    if incident.created_by==user and role!="ADMIN": st.error("Separation of duties: the person who opened the incident cannot close it.")
    elif min(len(root.strip()),len(corrective.strip()),len(closure_ref.strip()))<3: st.error("Enter the root cause, corrective action and closure reference.")
    else:
     c=conn.cursor(); c.execute("UPDATE inventory_incidents SET status='CLOSED',root_cause=%s,corrective_action=%s,closure_reference=%s,closed_by=%s,closed_at=CURRENT_TIMESTAMP WHERE id=%s AND status='OPEN'",(root.strip(),corrective.strip(),closure_ref.strip(),user,int(incident.id))); conn.commit(); record_event(conn,"CLOSE_INVENTORY_INCIDENT","Measurement Control","Inventory Incident",int(incident.id),closure_ref.strip()); st.success("Incident closed with a complete audit trail."); st.rerun()
 with report_tab:
  company=st.session_state.get("company_profile",{}).get("company_name","Company"); st.download_button("Download measurement & loss control report",_export({"Tank Counts":counts,"Calibrations":calibrations,"Incidents":incidents},company),"measurement_loss_control.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
