from io import BytesIO
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from audit import record_event
from ui import page_header

EDIT={"ADMIN","INVENTORY_MANAGER","STOREKEEPER","OPERATOR"}

def ensure_reservation_schema(conn):
 c=conn.cursor()
 try:
  c.execute("""CREATE TABLE IF NOT EXISTS stock_reservations(id BIGSERIAL PRIMARY KEY,reservation_number TEXT UNIQUE,
   customer_name TEXT NOT NULL,external_reference TEXT,source_system TEXT,product_id INTEGER NOT NULL REFERENCES products(id),
   requested_liters REAL NOT NULL CHECK(requested_liters>0),reserved_liters REAL NOT NULL CHECK(reserved_liters>=0),
   fulfilled_liters REAL NOT NULL DEFAULT 0,required_date DATE NOT NULL,expires_at TIMESTAMPTZ,priority TEXT NOT NULL DEFAULT 'NORMAL',
   status TEXT NOT NULL DEFAULT 'ACTIVE',notes TEXT,created_by TEXT,created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
   closed_by TEXT,closed_at TIMESTAMPTZ)""")
  c.execute("""CREATE TABLE IF NOT EXISTS reservation_fulfillments(id BIGSERIAL PRIMARY KEY,reservation_id BIGINT NOT NULL REFERENCES stock_reservations(id),
   tank_transaction_id BIGINT UNIQUE NOT NULL REFERENCES tank_transactions(id),fulfilled_liters REAL NOT NULL CHECK(fulfilled_liters>0),created_by TEXT,created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP)""")
  c.execute("CREATE INDEX IF NOT EXISTS idx_res_status ON stock_reservations(status,required_date)"); conn.commit()
 except Exception: conn.rollback(); raise

def _atp(conn):
 return pd.read_sql_query("""WITH usable AS (SELECT b.product_id,SUM(GREATEST(COALESCE(r.received,0)-COALESCE(a.used,0),0)) usable
  FROM fuel_batches b LEFT JOIN (SELECT batch_id,SUM(COALESCE(accepted_liters,liters)) received FROM tank_transactions WHERE batch_id IS NOT NULL AND type='IN' GROUP BY batch_id) r ON r.batch_id=b.id
  LEFT JOIN (SELECT batch_id,SUM(allocated_liters) used FROM batch_issue_allocations GROUP BY batch_id) a ON a.batch_id=b.id
  WHERE b.status='RELEASED' AND (b.expiry_date IS NULL OR b.expiry_date>=CURRENT_DATE) GROUP BY b.product_id),
  reserved AS (SELECT product_id,SUM(GREATEST(reserved_liters-fulfilled_liters,0)) reserved FROM stock_reservations WHERE status IN ('ACTIVE','PARTIALLY_FULFILLED') AND (expires_at IS NULL OR expires_at>CURRENT_TIMESTAMP) GROUP BY product_id)
  SELECT p.id product_id,p.name product,COALESCE(u.usable,0) usable_released_liters,COALESCE(r.reserved,0) active_reserved_liters,
  GREATEST(COALESCE(u.usable,0)-COALESCE(r.reserved,0),0) available_to_promise FROM products p LEFT JOIN usable u ON u.product_id=p.id LEFT JOIN reserved r ON r.product_id=p.id WHERE p.active=TRUE ORDER BY p.name""",conn)

def _reservations(conn):
 return pd.read_sql_query("""SELECT r.id,r.reservation_number,r.customer_name,r.external_reference,r.source_system,p.name product,r.requested_liters,r.reserved_liters,r.fulfilled_liters,
  GREATEST(r.reserved_liters-r.fulfilled_liters,0) outstanding_liters,r.required_date,r.expires_at,r.priority,r.status,r.notes,r.created_by,r.created_at
  FROM stock_reservations r JOIN products p ON p.id=r.product_id ORDER BY CASE r.priority WHEN 'CRITICAL' THEN 1 WHEN 'URGENT' THEN 2 ELSE 3 END,r.required_date,r.id""",conn)

def _report(atp,reservations,fulfilments,company):
 wb=Workbook(); wb.remove(wb.active)
 for name,data in {"Availability Summary":atp,"Demand Register":reservations,"Fulfilment History":fulfilments}.items():
  ws=wb.create_sheet(name); ws.append(list(data.columns))
  for row in data.where(pd.notna(data),None).itertuples(index=False,name=None): ws.append([str(v) if isinstance(v,pd.Timestamp) else v for v in row])
  for x in ws[1]: x.fill=PatternFill("solid",fgColor="111827"); x.font=Font(color="FFFFFF",bold=True)
  ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
  for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(max(14,max(len(str(v.value or '')) for v in col)+2),32)
 ws=wb["Availability Summary"]; ws.insert_rows(1,2); ws["A1"]=f"{company} | Available-to-Promise & Demand Control"; ws["A1"].fill=PatternFill("solid",fgColor="9E1B1B"); ws["A1"].font=Font(color="FFFFFF",bold=True,size=16)
 out=BytesIO(); wb.save(out); return out.getvalue()

def render_stock_reservations(conn):
 user=st.session_state.get("user","System"); role=st.session_state.get("role","VIEWER"); atp=_atp(conn); reservations=_reservations(conn)
 page_header("Reservations & Available to Promise","Control customer demand without double-promising released usable inventory.")
 a,b,c,d=st.columns(4); a.metric("Usable released",f"{atp.usable_released_liters.sum():,.0f} L"); b.metric("Actively reserved",f"{atp.active_reserved_liters.sum():,.0f} L"); c.metric("Available to promise",f"{atp.available_to_promise.sum():,.0f} L"); d.metric("Open requests",int(reservations.status.isin(['ACTIVE','PARTIALLY_FULFILLED']).sum()) if not reservations.empty else 0)
 overview,create,manage,fulfil,report=st.tabs(["Availability","Create reservation","Manage requests","Fulfil reservation","Demand report"])
 with overview: st.caption("ATP excludes quarantined, rejected, expired, already consumed and already reserved stock."); st.dataframe(atp,use_container_width=True,hide_index=True)
 with create:
  if role not in EDIT: st.info("Your role has view-only access.")
  else:
   products={r.product:int(r.product_id) for r in atp.itertuples()}
   with st.form("new_reservation"):
    x,y=st.columns(2); customer=x.text_input("Customer / requesting department"); reference=y.text_input("Customer order / external reference")
    x,y=st.columns(2); product=x.selectbox("Product",list(products)); source=y.text_input("Source system",value="Manual")
    x,y,z=st.columns(3); liters=x.number_input("Requested litres",min_value=0.01); required=y.date_input("Required date"); priority=z.selectbox("Priority",["NORMAL","URGENT","CRITICAL"])
    allow_partial=st.checkbox("Allow partial reservation",value=True); notes=st.text_area("Notes"); submit=st.form_submit_button("Create reservation",type="primary")
   if submit:
    available=float(atp.loc[atp.product.eq(product),'available_to_promise'].iloc[0]); reserved=min(float(liters),available) if allow_partial else float(liters)
    if len(customer.strip())<2: st.error("Enter the customer or requesting department.")
    elif not allow_partial and liters>available+0.005: st.error(f"Only {available:,.2f} L is available to promise.")
    elif reserved<=0: st.error("No released usable stock is available for this product.")
    else:
     c=conn.cursor(); c.execute("""INSERT INTO stock_reservations(customer_name,external_reference,source_system,product_id,requested_liters,reserved_liters,required_date,priority,status,notes,created_by)
      VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'ACTIVE',%s,%s) RETURNING id""",(customer.strip(),reference.strip() or None,source.strip() or "Manual",products[product],liters,reserved,required,priority,notes.strip() or None,user)); rid=c.fetchone()[0]; number=f"RS-{rid}"; c.execute("UPDATE stock_reservations SET reservation_number=%s WHERE id=%s",(number,rid)); conn.commit(); record_event(conn,"CREATE_RESERVATION","Demand Control","Stock Reservation",rid,f"Reserved {reserved:,.2f} L of requested {liters:,.2f} L for {customer.strip()}"); st.success(f"{number} created. {reserved:,.2f} L reserved; physical inventory was not changed."); st.rerun()
 with manage:
  st.dataframe(reservations,use_container_width=True,hide_index=True,height=420)
  open_rows=reservations[reservations.status.isin(["ACTIVE","PARTIALLY_FULFILLED"])] if not reservations.empty else reservations
  if role in EDIT and not open_rows.empty:
   choices={f"{r.reservation_number} · {r.customer_name} · {float(r.outstanding_liters):,.2f} L":int(r.id) for r in open_rows.itertuples()}; selected=st.selectbox("Open request",list(choices)); reason=st.text_input("Cancellation / release reason")
   if st.button("Cancel reservation"):
    if len(reason.strip())<5: st.error("Enter a clear reason.")
    else:
     rid=choices[selected]; c=conn.cursor(); c.execute("UPDATE stock_reservations SET status='CANCELLED',closed_by=%s,closed_at=CURRENT_TIMESTAMP,notes=CONCAT(COALESCE(notes,''),%s) WHERE id=%s",(user," | Cancelled: "+reason.strip(),rid)); conn.commit(); record_event(conn,"CANCEL_RESERVATION","Demand Control","Stock Reservation",rid,reason.strip()); st.success("Reservation cancelled and its outstanding quantity returned to ATP."); st.rerun()
 with fulfil:
  open_rows=reservations[reservations.status.isin(["ACTIVE","PARTIALLY_FULFILLED"])] if not reservations.empty else reservations
  if open_rows.empty: st.info("No reservation is waiting for fulfilment.")
  elif role not in EDIT: st.info("Your role has view-only access.")
  else:
   choices={f"{r.reservation_number} · {r.customer_name} · {float(r.outstanding_liters):,.2f} L":r for r in open_rows.itertuples()}; label=st.selectbox("Reservation",list(choices),key="fulfil_res"); res=choices[label]
   issues=pd.read_sql_query("""SELECT x.id,x.movement_at,x.liters,x.reference FROM tank_transactions x LEFT JOIN reservation_fulfillments f ON f.tank_transaction_id=x.id WHERE x.type='OUT' AND x.product_id=(SELECT product_id FROM stock_reservations WHERE id=%s) AND f.id IS NULL ORDER BY x.id DESC""",conn,params=[int(res.id)])
   if issues.empty: st.warning("No unlinked OUT transaction exists for this product. Post the delivery in Storage Operations first.")
   else:
    issue_map={f"STX-{int(r.id)} · {float(r.liters):,.2f} L · {r.reference or 'No reference'}":r for r in issues.itertuples()}; issue=issue_map[st.selectbox("Posted OUT transaction",list(issue_map))]
    amount=st.number_input("Litres fulfilled",min_value=0.01,max_value=min(float(issue.liters),float(res.outstanding_liters)),value=min(float(issue.liters),float(res.outstanding_liters)))
    if st.button("Link fulfilment",type="primary"):
     c=conn.cursor(); c.execute("INSERT INTO reservation_fulfillments(reservation_id,tank_transaction_id,fulfilled_liters,created_by) VALUES (%s,%s,%s,%s) RETURNING id",(int(res.id),int(issue.id),amount,user)); fid=c.fetchone()[0]; c.execute("UPDATE stock_reservations SET fulfilled_liters=fulfilled_liters+%s,status=CASE WHEN fulfilled_liters+%s>=reserved_liters-0.005 THEN 'FULFILLED' ELSE 'PARTIALLY_FULFILLED' END,closed_by=CASE WHEN fulfilled_liters+%s>=reserved_liters-0.005 THEN %s ELSE closed_by END,closed_at=CASE WHEN fulfilled_liters+%s>=reserved_liters-0.005 THEN CURRENT_TIMESTAMP ELSE closed_at END WHERE id=%s",(amount,amount,amount,user,amount,int(res.id))); conn.commit(); record_event(conn,"FULFIL_RESERVATION","Demand Control","Reservation Fulfilment",fid,f"Linked STX-{int(issue.id)}; {amount:,.2f} L fulfilled"); st.success("Fulfilment linked. No additional inventory movement was posted."); st.rerun()
 fulfilments=pd.read_sql_query("""SELECT f.id,r.reservation_number,r.customer_name,x.id tank_transaction_id,f.fulfilled_liters,f.created_by,f.created_at FROM reservation_fulfillments f JOIN stock_reservations r ON r.id=f.reservation_id JOIN tank_transactions x ON x.id=f.tank_transaction_id ORDER BY f.id DESC""",conn)
 with report:
  st.dataframe(fulfilments,use_container_width=True,hide_index=True); company=st.session_state.get("company_profile",{}).get("company_name","Company"); st.download_button("Download ATP & demand report",_report(atp,reservations,fulfilments,company),"available_to_promise_demand_report.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
