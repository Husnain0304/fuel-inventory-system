from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from audit import record_event
from period_close import assert_period_open
from ui import page_header

OPERATORS={"ADMIN","INVENTORY_MANAGER","STOREKEEPER","OPERATOR"}


def ensure_transit_schema(conn):
    c=conn.cursor()
    try:
        c.execute("""CREATE TABLE IF NOT EXISTS inventory_transfers(
            id BIGSERIAL PRIMARY KEY,transfer_number TEXT UNIQUE,source_tank_id INTEGER NOT NULL REFERENCES storage_tanks(id),
            destination_tank_id INTEGER NOT NULL REFERENCES storage_tanks(id),product_id INTEGER NOT NULL REFERENCES products(id),
            planned_liters REAL NOT NULL CHECK(planned_liters>0),dispatched_liters REAL,received_liters REAL,variance_liters REAL,
            vehicle_number TEXT,driver_name TEXT,seal_number TEXT,dispatch_reference TEXT,receipt_reference TEXT,
            planned_dispatch_at TIMESTAMPTZ,status TEXT NOT NULL DEFAULT 'PLANNED',source_transaction_id BIGINT UNIQUE REFERENCES tank_transactions(id),
            destination_transaction_id BIGINT UNIQUE REFERENCES tank_transactions(id),created_by TEXT,created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
            dispatched_by TEXT,dispatched_at TIMESTAMPTZ,received_by TEXT,received_at TIMESTAMPTZ,cancelled_by TEXT,cancelled_at TIMESTAMPTZ,
            cancellation_reason TEXT,notes TEXT,CHECK(source_tank_id<>destination_tank_id))""")
        c.execute("CREATE INDEX IF NOT EXISTS idx_inventory_transfers_status ON inventory_transfers(status,planned_dispatch_at)")
        conn.commit()
    except Exception: conn.rollback(); raise


def _tank_positions(conn):
    return pd.read_sql_query("""SELECT t.id,t.product_id,CONCAT(d.code,' / ',t.code,' · ',p.name) label,t.safe_capacity_liters,
        COALESCE(SUM(CASE WHEN x.type='IN' THEN x.liters ELSE -x.liters END),0) balance
        FROM storage_tanks t JOIN depots d ON d.id=t.depot_id JOIN products p ON p.id=t.product_id
        LEFT JOIN tank_transactions x ON x.tank_id=t.id AND COALESCE(x.record_status,'POSTED')='POSTED'
        WHERE t.status='AVAILABLE' GROUP BY t.id,d.code,p.name ORDER BY d.code,t.code""",conn)


def _transfers(conn):
    return pd.read_sql_query("""SELECT m.id,m.transfer_number,sd.code source_depot,st.code source_tank,dd.code destination_depot,dt.code destination_tank,
        p.name product,m.planned_liters,m.dispatched_liters,m.received_liters,m.variance_liters,m.vehicle_number,m.driver_name,m.seal_number,
        m.planned_dispatch_at,m.dispatched_at,m.received_at,m.status,m.dispatch_reference,m.receipt_reference,m.created_by,m.dispatched_by,m.received_by
        FROM inventory_transfers m JOIN storage_tanks st ON st.id=m.source_tank_id JOIN depots sd ON sd.id=st.depot_id
        JOIN storage_tanks dt ON dt.id=m.destination_tank_id JOIN depots dd ON dd.id=dt.depot_id JOIN products p ON p.id=m.product_id
        ORDER BY m.id DESC""",conn)


def _report(data,company):
    wb=Workbook(); ws=wb.active; ws.title="Stock in Transit"; ws.append([f"{company} | Stock in Transit Control"]); ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=max(len(data.columns),1)); ws["A1"].fill=PatternFill("solid",fgColor="9E1B1B"); ws["A1"].font=Font(color="FFFFFF",bold=True,size=16); ws.append(list(data.columns))
    for row in data.where(pd.notna(data),None).itertuples(index=False,name=None): ws.append([str(v) if isinstance(v,pd.Timestamp) else v for v in row])
    for c in ws[2]: c.fill=PatternFill("solid",fgColor="111827"); c.font=Font(color="FFFFFF",bold=True)
    ws.freeze_panes="A3"; ws.auto_filter.ref=ws.dimensions
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(max(14,max(len(str(v.value or "")) for v in col)+2),32)
    out=BytesIO(); wb.save(out); return out.getvalue()


def render_stock_transit(conn):
    role=st.session_state.get("role","VIEWER"); user=st.session_state.get("user","System"); positions=_tank_positions(conn); transfers=_transfers(conn)
    page_header("Stock in Transit","Control depot-to-depot fuel movements from planning through dispatch, transit and destination receipt.")
    a,b,c,d=st.columns(4); a.metric("Planned",int(transfers.status.eq("PLANNED").sum()) if not transfers.empty else 0); b.metric("In transit",int(transfers.status.eq("IN_TRANSIT").sum()) if not transfers.empty else 0); c.metric("Transit quantity",f"{transfers.loc[transfers.status.eq('IN_TRANSIT'),'dispatched_liters'].sum():,.0f} L" if not transfers.empty else "0 L"); d.metric("Received variance",f"{transfers.loc[transfers.status.eq('RECEIVED'),'variance_liters'].sum():+,.0f} L" if not transfers.empty else "0 L")
    board,plan,dispatch,receive,cancel,report=st.tabs(["Transit board","Plan transfer","Dispatch source","Receive destination","Cancel plan","Transit report"])
    with board: st.dataframe(transfers,use_container_width=True,hide_index=True,height=480)
    with plan:
        if role not in OPERATORS: st.info("Your role has view-only access.")
        elif len(positions)<2: st.warning("At least two available tanks are required.")
        else:
            pm={r.label:r for r in positions.itertuples()}
            with st.form("plan_transit"):
                source=st.selectbox("Source tank",list(pm)); source_row=pm[source]; destinations=positions[(positions.product_id==source_row.product_id)&(positions.id!=source_row.id)]; dm={r.label:r for r in destinations.itertuples()}; destination=st.selectbox("Destination tank",list(dm)) if dm else None
                a,b=st.columns(2); liters=a.number_input("Planned quantity",min_value=0.01,max_value=max(float(source_row.balance),0.01)); planned=b.datetime_input("Planned dispatch",value=datetime.now()); a,b,c=st.columns(3); vehicle=a.text_input("Vehicle number"); driver=b.text_input("Driver"); seal=c.text_input("Seal number"); notes=st.text_area("Transfer notes"); submit=st.form_submit_button("Create transfer plan",type="primary",disabled=not dm)
            if submit:
                destination_row=dm[destination]
                if liters>float(source_row.balance)+.005: st.error("Planned quantity exceeds current source stock.")
                elif liters+float(destination_row.balance)>float(destination_row.safe_capacity_liters)+.005: st.error("Planned quantity would exceed destination safe capacity.")
                else:
                    c=conn.cursor(); c.execute("""INSERT INTO inventory_transfers(source_tank_id,destination_tank_id,product_id,planned_liters,vehicle_number,driver_name,seal_number,planned_dispatch_at,notes,created_by)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",(int(source_row.id),int(destination_row.id),int(source_row.product_id),liters,vehicle.strip() or None,driver.strip() or None,seal.strip() or None,planned,notes.strip() or None,user)); tid=c.fetchone()[0]; c.execute("UPDATE inventory_transfers SET transfer_number=%s WHERE id=%s",(f"IT-{tid}",tid)); conn.commit(); record_event(conn,"PLAN_STOCK_TRANSFER","Stock Transit","Inventory Transfer",tid,f"Planned {liters:,.2f} L from {source} to {destination}"); st.success(f"IT-{tid} planned. Inventory has not changed."); st.rerun()
    with dispatch:
        pending=transfers[transfers.status.eq("PLANNED")] if not transfers.empty else transfers
        if pending.empty: st.success("No planned transfer is waiting for dispatch.")
        elif role not in OPERATORS: st.dataframe(pending,use_container_width=True,hide_index=True)
        else:
            choices={f"{r.transfer_number} · {r.source_depot}/{r.source_tank} → {r.destination_depot}/{r.destination_tank} · {float(r.planned_liters):,.2f} L":r for r in pending.itertuples()}; selected=st.selectbox("Planned transfer",list(choices)); row=choices[selected]; actual=st.number_input("Metered quantity dispatched",min_value=0.01,value=float(row.planned_liters)); reference=st.text_input("Dispatch document reference"); dispatch_variance_reason=st.text_area("Dispatch variance explanation")
            if st.button("Confirm source dispatch",type="primary"):
                if abs(actual-float(row.planned_liters))>.005 and len(dispatch_variance_reason.strip())<5: st.error("Explain why the dispatched quantity differs from the plan.")
                elif len(reference.strip())<2: st.error("Enter the dispatch document reference.")
                else:
                 try:
                    assert_period_open(conn,datetime.now()); c=conn.cursor(); c.execute("SELECT source_tank_id,product_id,status FROM inventory_transfers WHERE id=%s FOR UPDATE",(int(row.id),)); current=c.fetchone()
                    c.execute("SELECT COALESCE(SUM(CASE WHEN type='IN' THEN liters ELSE -liters END),0) FROM tank_transactions WHERE tank_id=%s AND COALESCE(record_status,'POSTED')='POSTED'",(current[0],)); balance=float(c.fetchone()[0] or 0)
                    if current[2]!="PLANNED": raise ValueError("Transfer is no longer waiting for dispatch.")
                    if actual>balance+.005: raise ValueError(f"Only {balance:,.2f} L is available in the source tank.")
                    c.execute("""INSERT INTO tank_transactions(tank_id,movement_at,liters,type,movement_category,product_id,reference,notes,created_by)
                        VALUES (%s,CURRENT_TIMESTAMP,%s,'OUT','TRANSIT_DISPATCH',%s,%s,%s,%s) RETURNING id""",(current[0],actual,current[1],reference.strip(),f"Stock in transit {row.transfer_number}. {dispatch_variance_reason.strip()}",user)); tx=c.fetchone()[0]
                    c.execute("UPDATE inventory_transfers SET status='IN_TRANSIT',dispatched_liters=%s,dispatch_reference=%s,source_transaction_id=%s,dispatched_by=%s,dispatched_at=CURRENT_TIMESTAMP WHERE id=%s",(actual,reference.strip(),tx,user,int(row.id))); conn.commit(); record_event(conn,"DISPATCH_STOCK_TRANSFER","Stock Transit","Inventory Transfer",int(row.id),f"Dispatched {actual:,.2f} L; STX-{tx}"); st.success("Source OUT posted. Fuel is now shown as in transit."); st.rerun()
                 except Exception as error: conn.rollback(); st.error(str(error))
    with receive:
        moving=transfers[transfers.status.eq("IN_TRANSIT")] if not transfers.empty else transfers
        if moving.empty: st.success("No transfer is currently in transit.")
        elif role not in OPERATORS: st.dataframe(moving,use_container_width=True,hide_index=True)
        else:
            choices={f"{r.transfer_number} · {float(r.dispatched_liters):,.2f} L dispatched":r for r in moving.itertuples()}; selected=st.selectbox("In-transit movement",list(choices)); row=choices[selected]; accepted=st.number_input("Quantity accepted at destination",min_value=0.0,value=float(row.dispatched_liters)); reference=st.text_input("Destination receipt reference"); explanation=st.text_area("Variance explanation")
            if st.button("Confirm destination receipt",type="primary"):
                variance=accepted-float(row.dispatched_liters)
                if len(reference.strip())<2: st.error("Enter the destination receipt reference.")
                elif abs(variance)>.005 and len(explanation.strip())<5: st.error("Enter a clear explanation for the transit variance.")
                else:
                    try:
                        assert_period_open(conn,datetime.now()); c=conn.cursor(); c.execute("SELECT destination_tank_id,product_id,status FROM inventory_transfers WHERE id=%s FOR UPDATE",(int(row.id),)); current=c.fetchone()
                        if current[2]!="IN_TRANSIT": raise ValueError("Transfer is no longer in transit.")
                        c.execute("SELECT t.safe_capacity_liters,COALESCE(SUM(CASE WHEN x.type='IN' THEN x.liters ELSE -x.liters END),0) FROM storage_tanks t LEFT JOIN tank_transactions x ON x.tank_id=t.id AND COALESCE(x.record_status,'POSTED')='POSTED' WHERE t.id=%s GROUP BY t.id",(current[0],)); capacity,balance=c.fetchone()
                        if accepted+float(balance)>float(capacity)+.005: raise ValueError("Accepted quantity would exceed destination safe capacity.")
                        if accepted>0:
                            c.execute("""INSERT INTO tank_transactions(tank_id,movement_at,liters,type,movement_category,product_id,reference,notes,created_by)
                                VALUES (%s,CURRENT_TIMESTAMP,%s,'IN','TRANSIT_RECEIPT',%s,%s,%s,%s) RETURNING id""",(current[0],accepted,current[1],reference.strip(),f"Stock in transit {row.transfer_number}; variance {variance:+,.2f} L. {explanation.strip()}",user)); tx=c.fetchone()[0]
                        else: tx=None
                        c.execute("UPDATE inventory_transfers SET status='RECEIVED',received_liters=%s,variance_liters=%s,receipt_reference=%s,destination_transaction_id=%s,received_by=%s,received_at=CURRENT_TIMESTAMP WHERE id=%s",(accepted,variance,reference.strip(),tx,user,int(row.id))); conn.commit(); record_event(conn,"RECEIVE_STOCK_TRANSFER","Stock Transit","Inventory Transfer",int(row.id),f"Received {accepted:,.2f} L; variance {variance:+,.2f} L",severity="WARNING" if abs(variance)>.005 else "INFO"); st.success("Destination receipt posted and transit closed."); st.rerun()
                    except Exception as error: conn.rollback(); st.error(str(error))
    with cancel:
        planned_rows=transfers[transfers.status.eq("PLANNED")] if not transfers.empty else transfers
        if planned_rows.empty: st.success("No planned transfer can be cancelled.")
        elif role not in OPERATORS: st.dataframe(planned_rows,use_container_width=True,hide_index=True)
        else:
            cm={f"{r.transfer_number} · {r.source_depot}/{r.source_tank} → {r.destination_depot}/{r.destination_tank}":r for r in planned_rows.itertuples()}; choice=st.selectbox("Planned transfer to cancel",list(cm)); row=cm[choice]; reason=st.text_area("Cancellation reason"); authorization=st.text_input("Authorization / supporting reference",key="transit_cancel_reference"); confirmed=st.checkbox("I confirm this plan has not been physically dispatched.")
            if st.button("Cancel transfer plan",disabled=not confirmed):
                if len(reason.strip())<5 or len(authorization.strip())<2: st.error("Enter a clear reason and supporting reference.")
                else:
                    c=conn.cursor(); c.execute("UPDATE inventory_transfers SET status='CANCELLED',cancelled_by=%s,cancelled_at=CURRENT_TIMESTAMP,cancellation_reason=%s WHERE id=%s AND status='PLANNED'",(user,f"{reason.strip()} | Reference: {authorization.strip()}",int(row.id)))
                    if c.rowcount!=1: conn.rollback(); st.error("This plan is no longer available for cancellation.")
                    else: conn.commit(); record_event(conn,"CANCEL_STOCK_TRANSFER","Stock Transit","Inventory Transfer",int(row.id),authorization.strip(),severity="WARNING"); st.success("Transfer plan cancelled. Inventory was not changed."); st.rerun()
    with report:
        company=st.session_state.get("company_profile",{}).get("company_name","Company"); st.download_button("Download stock-in-transit report",_report(transfers,company),"stock_in_transit_report.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
