from datetime import date, timedelta, datetime
import io

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from audit import record_event
from ui import page_header


def ensure_scorecard_schema(conn):
    cursor=conn.cursor()
    try:
        cursor.execute("""CREATE TABLE IF NOT EXISTS supplier_scorecard_settings(
            metric_key TEXT PRIMARY KEY,metric_name TEXT NOT NULL,weight_percent REAL NOT NULL,
            green_threshold REAL NOT NULL,amber_threshold REAL NOT NULL,updated_by TEXT,
            updated_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP)""")
        defaults=[("ON_TIME","On-time delivery",30,90,75),("FILL_RATE","Quantity fulfilment",30,95,85),("ACCURACY","Receipt accuracy",20,98,95),("CLAIM_FREE","Claim-free supply",20,99,95)]
        for row in defaults:
            cursor.execute("""INSERT INTO supplier_scorecard_settings(metric_key,metric_name,weight_percent,green_threshold,amber_threshold)
                VALUES (%s,%s,%s,%s,%s) ON CONFLICT(metric_key) DO NOTHING""",row)
        conn.commit()
    except Exception: conn.rollback(); raise


def _settings(conn):
    return pd.read_sql_query("SELECT * FROM supplier_scorecard_settings ORDER BY metric_key",conn)


def calculate_scorecards(conn,start,end):
    raw=pd.read_sql_query("""WITH release_rows AS (
            SELECT b.supplier_id,r.id,r.released_liters,r.planned_delivery_date,
                COALESCE(SUM(tx.accepted_liters),0) AS accepted_liters,MIN(tx.movement_at)::date AS first_receipt_date
            FROM procurement_releases r JOIN procurement_bookings b ON b.id=r.booking_id
            LEFT JOIN tank_transactions tx ON tx.booking_release_id=r.id
            WHERE r.release_date BETWEEN %s AND %s AND r.status<>'CANCELLED'
            GROUP BY b.supplier_id,r.id),
        releases AS (SELECT supplier_id,COUNT(*) AS releases,SUM(released_liters) AS released_liters,
            SUM(accepted_liters) AS release_received_liters,
            COUNT(*) FILTER(WHERE first_receipt_date IS NOT NULL AND planned_delivery_date IS NOT NULL) AS received_releases,
            COUNT(*) FILTER(WHERE first_receipt_date IS NOT NULL AND planned_delivery_date IS NOT NULL AND first_receipt_date<=planned_delivery_date) AS on_time_releases
            FROM release_rows GROUP BY supplier_id),
        receipts AS (SELECT supplier_id,COUNT(*) AS receipts,COALESCE(SUM(accepted_liters),0) AS received_liters,
            COALESCE(SUM(dispatched_liters),0) AS dispatched_liters,
            COALESCE(SUM(accepted_liters*unit_price),0) AS received_value
            FROM tank_transactions WHERE movement_category='SUPPLIER_RECEIPT'
            AND movement_at::date BETWEEN %s AND %s GROUP BY supplier_id),
        claims AS (SELECT supplier_id,COUNT(*) AS claims,COALESCE(SUM(claim_liters),0) AS claim_liters,
            COALESCE(SUM(claim_amount),0) AS claim_value FROM supplier_claims
            WHERE created_at::date BETWEEN %s AND %s GROUP BY supplier_id)
        SELECT s.id,s.supplier_code,s.name,s.status,COALESCE(r.releases,0) AS releases,
            COALESCE(r.released_liters,0) AS released_liters,COALESCE(r.release_received_liters,0) AS release_received_liters,
            COALESCE(r.received_releases,0) AS received_releases,COALESCE(r.on_time_releases,0) AS on_time_releases,
            COALESCE(x.receipts,0) AS receipts,COALESCE(x.received_liters,0) AS received_liters,
            COALESCE(x.dispatched_liters,0) AS dispatched_liters,COALESCE(x.received_value,0) AS received_value,
            COALESCE(c.claims,0) AS claims,COALESCE(c.claim_liters,0) AS claim_liters,COALESCE(c.claim_value,0) AS claim_value
        FROM suppliers s LEFT JOIN releases r ON r.supplier_id=s.id LEFT JOIN receipts x ON x.supplier_id=s.id
        LEFT JOIN claims c ON c.supplier_id=s.id ORDER BY s.name""",conn,params=[start,end,start,end,start,end])
    config=_settings(conn); weights=dict(zip(config.metric_key,config.weight_percent)); green=dict(zip(config.metric_key,config.green_threshold)); amber=dict(zip(config.metric_key,config.amber_threshold))
    total_weight=sum(weights.values()) or 100; preferred_cut=sum(weights[k]*green[k] for k in weights)/total_weight; approved_cut=sum(weights[k]*amber[k] for k in weights)/total_weight; watch_cut=max(50,approved_cut-15)
    def metrics(row):
        on_time=100*row.on_time_releases/row.received_releases if row.received_releases else None
        fill=min(100*row.release_received_liters/row.released_liters,100) if row.released_liters else None
        accuracy=max(0,100-100*abs(row.received_liters-row.dispatched_liters)/row.dispatched_liters) if row.dispatched_liters else None
        claim_free=max(0,100-100*row.claim_liters/row.received_liters) if row.received_liters else None
        values={"ON_TIME":on_time,"FILL_RATE":fill,"ACCURACY":accuracy,"CLAIM_FREE":claim_free}; available=[k for k,v in values.items() if v is not None]
        score=sum(values[k]*weights.get(k,0) for k in available)/sum(weights.get(k,0) for k in available) if available else None
        band="NO DATA" if score is None else ("PREFERRED" if score>=preferred_cut else "APPROVED" if score>=approved_cut else "WATCH" if score>=watch_cut else "HIGH RISK")
        return pd.Series([on_time,fill,accuracy,claim_free,score,band])
    raw[["On-Time %","Fill Rate %","Receipt Accuracy %","Claim-Free %","Overall Score","Risk Band"]]=raw.apply(metrics,axis=1)
    raw["Rank"]=raw["Overall Score"].rank(method="min",ascending=False,na_option="bottom").astype(int)
    return raw.sort_values(["Overall Score","name"],ascending=[False,True],na_position="last"),config


def build_scorecard_report(data,config,profile,start,end):
    wb=Workbook(); summary=wb.active; summary.title="Executive Scorecard"; primary=str(profile.get("primary_color","#9E1B1B")).replace("#",""); secondary=str(profile.get("secondary_color","#172033")).replace("#","")
    summary.sheet_view.showGridLines=False; summary.merge_cells("A1:L2"); summary["A1"]=f"{profile.get('company_name','Company')} | Supplier Performance Scorecard"; summary["A1"].fill=PatternFill("solid",fgColor=secondary); summary["A1"].font=Font(size=20,bold=True,color="FFFFFF"); summary["A1"].alignment=Alignment(vertical="center")
    summary.merge_cells("A3:L3"); summary["A3"]=f"Performance period: {start:%d %b %Y} to {end:%d %b %Y}"; summary["A3"].fill=PatternFill("solid",fgColor=primary); summary["A3"].font=Font(color="FFFFFF")
    report=data[["Rank","supplier_code","name","status","releases","receipts","received_liters","received_value","On-Time %","Fill Rate %","Receipt Accuracy %","Claim-Free %","Overall Score","Risk Band","claims","claim_value"]].copy(); report.columns=[str(c).replace("_"," ").title() for c in report.columns]
    row=5
    for col,value in enumerate(report.columns,1): cell=summary.cell(row,col,value); cell.fill=PatternFill("solid",fgColor=primary); cell.font=Font(bold=True,color="FFFFFF"); cell.alignment=Alignment(wrap_text=True)
    for values in report.itertuples(index=False,name=None): summary.append(list(values))
    if not report.empty:
        end_row=5+len(report); table=Table(displayName="SupplierScorecards",ref=f"A5:{get_column_letter(len(report.columns))}{end_row}"); table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True); summary.add_table(table)
        for col,header in enumerate(report.columns,1):
            summary.column_dimensions[get_column_letter(col)].width=24 if header in ("Name","Risk Band") else 15
            if "%" in header or header=="Overall Score":
                for r in range(6,end_row+1): summary.cell(r,col).number_format='0.0'
    summary.freeze_panes="A6"
    rules=wb.create_sheet("Scoring Rules"); rules.append(["Metric Key","Metric","Weight %","Preferred Threshold","Watch Threshold","Method"])
    methods={"ON_TIME":"Receipts on/before planned delivery date ÷ received releases","FILL_RATE":"Accepted quantity ÷ released quantity, capped at 100%","ACCURACY":"100% less absolute dispatched-to-accepted variance","CLAIM_FREE":"100% less claimed liters as a share of received liters"}
    for r in config.itertuples(): rules.append([r.metric_key,r.metric_name,r.weight_percent,r.green_threshold,r.amber_threshold,methods.get(r.metric_key,"")])
    rules["A8"]="Control"; rules["B8"]="Weights total"; rules["C8"]="=SUM(C2:C5)"; rules["D8"]="=IF(ABS(C8-100)<0.01,\"OK\",\"REVIEW\")"
    for cell in rules[1]: cell.fill=PatternFill("solid",fgColor=primary); cell.font=Font(bold=True,color="FFFFFF")
    for col,width in enumerate([16,26,14,20,18,60],1): rules.column_dimensions[get_column_letter(col)].width=width
    output=io.BytesIO(); wb.save(output); return output.getvalue()


def render_supplier_scorecards(conn):
    ensure_scorecard_schema(conn); from procurement import ensure_procurement_schema; ensure_procurement_schema(conn)
    profile=st.session_state.get("company_profile",{}); page_header("Supplier Performance Scorecards","Rank suppliers using consistent delivery, quantity, accuracy and claim controls.")
    a,b=st.columns(2); start=a.date_input("From",date.today()-timedelta(days=90),key="score_start"); end=b.date_input("To",date.today(),key="score_end")
    if end<start: st.error("The end date must be on or after the start date."); return
    data,config=calculate_scorecards(conn,start,end); active=data[data["receipts"]>0]
    avg=float(active["Overall Score"].mean()) if not active.empty else 0; high=int((data["Risk Band"]=="HIGH RISK").sum()); preferred=int((data["Risk Band"]=="PREFERRED").sum())
    a,b,c,d=st.columns(4); a.metric("Suppliers with receipts",len(active)); b.metric("Average score",f"{avg:.1f}"); c.metric("Preferred",preferred); d.metric("High risk",high)
    scorecard,detail,settings,report=st.tabs(["Ranking","Supplier detail","Scoring policy","Management report"])
    with scorecard:
        st.dataframe(data[["Rank","supplier_code","name","status","receipts","received_liters","On-Time %","Fill Rate %","Receipt Accuracy %","Claim-Free %","Overall Score","Risk Band"]],use_container_width=True,hide_index=True,height=500)
    with detail:
        choices={f"{r.supplier_code} · {r.name}":int(r.id) for r in data.itertuples()}; selected=st.selectbox("Supplier",list(choices)); row=data[data.id==choices[selected]].iloc[0]
        a,b,c,d=st.columns(4); a.metric("Overall",f"{row['Overall Score']:.1f}" if pd.notna(row["Overall Score"]) else "No data"); b.metric("On-time",f"{row['On-Time %']:.1f}%" if pd.notna(row["On-Time %"]) else "—"); c.metric("Fill rate",f"{row['Fill Rate %']:.1f}%" if pd.notna(row["Fill Rate %"]) else "—"); d.metric("Claim value",f"{profile.get('currency','AED')} {row['claim_value']:,.2f}")
        st.dataframe(pd.DataFrame({"Measure":["Releases","Receipts","Released liters","Received liters","Receipt accuracy","Claims","Risk band"],"Result":[row.releases,row.receipts,row.released_liters,row.received_liters,row["Receipt Accuracy %"],row.claims,row["Risk Band"]]}),use_container_width=True,hide_index=True)
    with settings:
        st.caption("Weights must total 100%. Changes affect future and historical scorecard views immediately.")
        editable=st.session_state.get("role") in ("ADMIN","INVENTORY_MANAGER"); edited=st.data_editor(config,column_config={"metric_key":None,"metric_name":"Metric","weight_percent":st.column_config.NumberColumn("Weight %",min_value=0,max_value=100),"green_threshold":st.column_config.NumberColumn("Preferred threshold",min_value=0,max_value=100),"amber_threshold":st.column_config.NumberColumn("Watch threshold",min_value=0,max_value=100),"updated_by":None,"updated_at":None},disabled=["metric_name"] if editable else list(config.columns),hide_index=True,use_container_width=True)
        if st.button("Save scoring policy",type="primary",disabled=not editable):
            if abs(float(edited.weight_percent.sum())-100)>0.01: st.error("Metric weights must total exactly 100%.")
            elif (edited.green_threshold<edited.amber_threshold).any(): st.error("Preferred thresholds cannot be below watch thresholds.")
            else:
                cursor=conn.cursor()
                for r in edited.itertuples(): cursor.execute("UPDATE supplier_scorecard_settings SET weight_percent=%s,green_threshold=%s,amber_threshold=%s,updated_by=%s,updated_at=CURRENT_TIMESTAMP WHERE metric_key=%s",(r.weight_percent,r.green_threshold,r.amber_threshold,st.session_state["user"],r.metric_key))
                conn.commit(); record_event(conn,"UPDATE_SCORECARD_POLICY","Supplier Scorecards","Scorecard Policy",1,"Updated supplier performance weights and thresholds"); st.success("Scoring policy saved."); st.rerun()
    with report:
        output=build_scorecard_report(data,config,profile,start,end); st.download_button("Download supplier scorecard workbook",output,f"supplier_scorecard_{start:%Y%m%d}_{end:%Y%m%d}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
        st.caption("Includes management ranking, operational KPIs, risk bands and the complete scoring methodology.")
