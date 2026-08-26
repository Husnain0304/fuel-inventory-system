from datetime import datetime
import io

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from audit import record_event
from procurement import booking_options, create_variance_claim, ensure_procurement_schema, refresh_booking_status, release_options
from ui import page_header
from approval_workflow import needs_approval, submit_approval_request


REPORT_COLUMNS = [
    "Movement ID", "Date", "Time", "Depot", "Tank", "Product", "Direction",
    "Quantity (L)", "Movement", "Supplier", "Ordered (L)", "Dispatched (L)",
    "Accepted (L)", "Variance (L)", "Variance %", "Reference", "Transport",
    "Vehicle", "Driver", "Linked Tank Movement", "Linked Truck Movement",
    "Status", "Recorded By", "Notes",
]


def _colour(value, fallback):
    value = str(value or "").strip().lstrip("#")
    return value.upper() if len(value) in (6, 8) else fallback


def _movement_label(value):
    labels = {
        "SUPPLIER_RECEIPT": "Supplier Receipt",
        "TANK_TRANSFER": "Tank Transfer",
        "TANK_TO_TRUCK": "Tank to Truck",
        "TRUCK_TO_TANK": "Truck Return to Tank",
        "OPENING": "Opening Stock",
        "STANDARD": "Standard Movement",
    }
    text = str(value or "STANDARD").strip().upper()
    return labels.get(text, text.replace("_", " ").title())


def _variance_label(value):
    if pd.isna(value):
        return "Not Applicable"
    amount = float(value)
    if amount < -0.005:
        return "Short Received"
    if amount > 0.005:
        return "Excess Received"
    return "Matched"


def _prepare_export(history):
    data = history.copy()
    timestamp = pd.to_datetime(data["movement_at"], errors="coerce", utc=True)
    local = timestamp.dt.tz_convert("Asia/Dubai").dt.tz_localize(None)
    records = pd.DataFrame({
        "Movement ID": data["id"].map(lambda value: f"STX-{int(value)}"),
        "Date": local.dt.date,
        "Time": local.dt.time,
        "Depot": data["depot"],
        "Tank": data["tank"],
        "Product": data["product"],
        "Direction": data["type"].map({"IN": "IN", "OUT": "OUT"}).fillna(data["type"]),
        "Quantity (L)": pd.to_numeric(data["liters"], errors="coerce").fillna(0.0),
        "Movement": data["movement_category"].map(_movement_label),
        "Supplier": data["supplier"],
        "Ordered (L)": pd.to_numeric(data["ordered_liters"], errors="coerce"),
        "Dispatched (L)": pd.to_numeric(data["dispatched_liters"], errors="coerce"),
        "Accepted (L)": pd.to_numeric(data["accepted_liters"], errors="coerce"),
        "Variance (L)": pd.to_numeric(data["variance_liters"], errors="coerce"),
        "Variance %": 0.0,
        "Reference": data["reference"],
        "Transport": data["transport_method"],
        "Vehicle": data["vehicle_number"],
        "Driver": data["driver_name"],
        "Linked Tank Movement": data["partner_tank_transaction_id"].map(
            lambda value: f"STX-{int(value)}" if pd.notna(value) else ""
        ),
        "Linked Truck Movement": data["truck_transaction_id"].map(
            lambda value: f"TX-{int(value)}" if pd.notna(value) else ""
        ),
        "Status": data["record_status"].fillna("POSTED").str.title(),
        "Recorded By": data["created_by"].fillna("System"),
        "Notes": data["notes"],
    })
    dispatched = records["Dispatched (L)"].replace(0, pd.NA)
    records["Variance %"] = records["Variance (L)"] / dispatched
    return records.where(pd.notna(records), None)


def _apply_report_header(sheet, title, subtitle, profile, end_column):
    primary = _colour(profile.get("primary_color"), "8C1C1C")
    secondary = _colour(profile.get("secondary_color"), "172033")
    end = get_column_letter(end_column)
    sheet.merge_cells(f"A1:{end}2")
    sheet["A1"] = title
    sheet["A1"].fill = PatternFill("solid", fgColor=secondary)
    sheet["A1"].font = Font(name="Aptos Display", size=22, bold=True, color="FFFFFF")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.merge_cells(f"A3:{end}3")
    sheet["A3"] = subtitle
    sheet["A3"].fill = PatternFill("solid", fgColor=primary)
    sheet["A3"].font = Font(name="Aptos", size=10, color="FFFFFF")
    sheet["A3"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 22
    sheet.sheet_view.showGridLines = False


def _style_table_sheet(sheet, headers, first_row, last_row, widths, table_name, profile):
    primary = _colour(profile.get("primary_color"), "8C1C1C")
    header_fill = PatternFill("solid", fgColor=primary)
    header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    light_border = Border(bottom=Side(style="thin", color="D9DEE8"))
    for cell in sheet[first_row]:
        if cell.column <= len(headers):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[first_row].height = 32
    for row in sheet.iter_rows(min_row=first_row + 1, max_row=last_row, max_col=len(headers)):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color="253047")
            cell.border = light_border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (9, 16, 24))
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    if last_row > first_row:
        table = Table(displayName=table_name, ref=f"A{first_row}:{get_column_letter(len(headers))}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        sheet.add_table(table)
    sheet.freeze_panes = f"A{first_row + 1}"
    sheet.auto_filter.ref = f"A{first_row}:{get_column_letter(len(headers))}{last_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = f"1:{first_row}"


def build_operations_report(history, profile):
    detail = _prepare_export(history)
    receipt_data = detail[detail["Movement"] == "Supplier Receipt"].copy()
    detail_end = 6 + max(len(detail), 1)
    receipt_end = 6 + max(len(receipt_data), 1)
    company = profile.get("company_name", "FILLIT")
    application = profile.get("application_name", "Fuel Inventory Control")
    footer = profile.get("report_footer", "Confidential inventory report")
    generated = datetime.now()
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Executive Summary"
    movements = workbook.create_sheet("Movement Details")
    receipts = workbook.create_sheet("Receipt Variance")
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    _apply_report_header(summary, f"{company} | Storage Operations Report", application, profile, 10)
    summary.merge_cells("A4:J4")
    summary["A4"] = f"Generated {generated:%d %b %Y, %I:%M %p} (Dubai)  •  {len(detail):,} movement records"
    summary["A4"].font = Font(name="Aptos", size=10, italic=True, color="667085")
    summary["A4"].alignment = Alignment(vertical="center")
    summary.row_dimensions[4].height = 24

    primary = _colour(profile.get("primary_color"), "8C1C1C")
    secondary = _colour(profile.get("secondary_color"), "172033")
    card_fill = PatternFill("solid", fgColor="F2F4F7")
    cards = [
        ("A6:B6", "A7:B8", "TOTAL MOVEMENTS", f"=COUNTA('Movement Details'!$A$7:$A${detail_end})", "#,##0"),
        ("D6:E6", "D7:E8", "TOTAL RECEIVED", f'=SUMIF(\'Movement Details\'!$G$7:$G${detail_end},"IN",\'Movement Details\'!$H$7:$H${detail_end})', '#,##0.00 "L"'),
        ("G6:H6", "G7:H8", "TOTAL ISSUED", f'=SUMIF(\'Movement Details\'!$G$7:$G${detail_end},"OUT",\'Movement Details\'!$H$7:$H${detail_end})', '#,##0.00 "L"'),
        ("I6:J6", "I7:J8", "NET MOVEMENT", "=D7-G7", '#,##0.00 "L";[Red]-#,##0.00 "L"'),
    ]
    for label_range, value_range, label, formula, number_format in cards:
        summary.merge_cells(label_range)
        summary.merge_cells(value_range)
        label_cell = summary[label_range.split(":")[0]]
        value_cell = summary[value_range.split(":")[0]]
        label_cell.value = label
        value_cell.value = formula
        label_cell.fill = PatternFill("solid", fgColor=secondary)
        label_cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.fill = card_fill
        value_cell.font = Font(name="Aptos Display", size=20, bold=True, color=primary)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format

    summary.merge_cells("A10:J10")
    summary["A10"] = "Supplier Receipt Control"
    summary["A10"].fill = PatternFill("solid", fgColor=secondary)
    summary["A10"].font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    summary["A10"].alignment = Alignment(vertical="center")
    summary_data = [
        ("Supplier receipts", f'=COUNTIF(\'Movement Details\'!$I$7:$I${detail_end},"Supplier Receipt")', "#,##0"),
        ("Ordered quantity", f"=SUM('Receipt Variance'!$F$7:$F${receipt_end})", '#,##0.00 "L"'),
        ("Dispatched quantity", f"=SUM('Receipt Variance'!$G$7:$G${receipt_end})", '#,##0.00 "L"'),
        ("Accepted quantity", f"=SUM('Receipt Variance'!$H$7:$H${receipt_end})", '#,##0.00 "L"'),
        ("Net receipt variance", f"=SUM('Receipt Variance'!$I$7:$I${receipt_end})", '#,##0.00 "L";[Red]-#,##0.00 "L"'),
        ("Short receipts requiring review", f'=COUNTIF(\'Receipt Variance\'!$K$7:$K${receipt_end},"Short Received")', "#,##0"),
    ]
    for row_number, (label, formula, number_format) in enumerate(summary_data, 11):
        summary.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=5)
        summary.merge_cells(start_row=row_number, start_column=6, end_row=row_number, end_column=10)
        summary.cell(row_number, 1, label)
        summary.cell(row_number, 6, formula)
        summary.cell(row_number, 1).font = Font(name="Aptos", size=10, color="475467")
        summary.cell(row_number, 6).font = Font(name="Aptos", size=11, bold=True, color="172033")
        summary.cell(row_number, 6).number_format = number_format
        summary.cell(row_number, 6).alignment = Alignment(horizontal="right")
        for cell in summary[row_number]:
            cell.fill = PatternFill("solid", fgColor="FFFFFF" if row_number % 2 else "F8FAFC")
            cell.border = Border(bottom=Side(style="thin", color="E4E7EC"))

    summary.merge_cells("A19:J19")
    summary["A19"] = "Report Guide"
    summary["A19"].fill = PatternFill("solid", fgColor=secondary)
    summary["A19"].font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    guide = [
        "Movement Details: complete audit-ready register of all tank movements.",
        "Receipt Variance: ordered, dispatched and accepted comparison for supplier receipts.",
        "Negative variance means less fuel was accepted than the supplier dispatched and should be investigated.",
    ]
    for row_number, text in enumerate(guide, 20):
        summary.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=10)
        summary.cell(row_number, 1, f"• {text}")
        summary.cell(row_number, 1).font = Font(name="Aptos", size=10, color="475467")
        summary.cell(row_number, 1).alignment = Alignment(wrap_text=True, vertical="center")
        summary.row_dimensions[row_number].height = 24
    for column in range(1, 11):
        summary.column_dimensions[get_column_letter(column)].width = 14
    summary.sheet_view.showGridLines = False
    summary.freeze_panes = "A5"
    summary.oddFooter.center.text = footer
    summary.oddFooter.right.text = "Page &P of &N"
    summary.page_setup.orientation = "landscape"
    summary.page_setup.fitToWidth = 1
    summary.sheet_properties.pageSetUpPr.fitToPage = True

    _apply_report_header(movements, "Movement Details", "Complete storage movement and audit register", profile, len(REPORT_COLUMNS))
    movements.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(REPORT_COLUMNS))
    movements.cell(4, 1, f"Generated {generated:%d %b %Y, %I:%M %p} (Dubai)  •  Use the filters below to investigate movements")
    movements.cell(4, 1).font = Font(name="Aptos", size=10, italic=True, color="667085")
    movements.append([])
    movements.append(REPORT_COLUMNS)
    for values in detail.itertuples(index=False, name=None):
        movements.append(list(values))
    detail_last = max(movements.max_row, 7)
    if len(detail) == 0:
        movements.append(["No storage movements found"] + [None] * (len(REPORT_COLUMNS) - 1))
        detail_last = movements.max_row
    widths = [14, 13, 11, 12, 15, 14, 11, 15, 22, 25, 15, 17, 15, 15, 13, 22, 22, 16, 18, 20, 20, 13, 16, 34]
    _style_table_sheet(movements, REPORT_COLUMNS, 6, detail_last, widths, "StorageMovements", profile)
    for row in range(7, detail_last + 1):
        movements.cell(row, 2).number_format = "dd mmm yyyy"
        movements.cell(row, 3).number_format = "hh:mm"
        for column in (8, 11, 12, 13, 14):
            movements.cell(row, column).number_format = '#,##0.00 "L";[Red]-#,##0.00 "L"'
        movements.cell(row, 15).number_format = "0.00%;[Red]-0.00%"
    movements.conditional_formatting.add(
        f"N7:N{detail_last}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FEE4E2"))
    )
    movements.conditional_formatting.add(
        f"N7:N{detail_last}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FEF0C7"))
    )
    movements.oddFooter.center.text = footer
    movements.oddFooter.right.text = "Page &P of &N"

    receipt_headers = [
        "Movement ID", "Date", "Depot", "Tank", "Supplier", "Ordered (L)",
        "Dispatched (L)", "Accepted (L)", "Variance (L)", "Accepted %",
        "Variance Status", "Reference", "Transport", "Vehicle", "Driver", "Recorded By",
    ]
    _apply_report_header(receipts, "Supplier Receipt Variance", "Ordered, dispatched and accepted fuel control", profile, len(receipt_headers))
    receipts.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(receipt_headers))
    receipts.cell(4, 1, "Short receipts are highlighted in red. Excess receipts are highlighted in amber.")
    receipts.cell(4, 1).font = Font(name="Aptos", size=10, italic=True, color="667085")
    receipts.append([])
    receipts.append(receipt_headers)
    for item in receipt_data.to_dict("records"):
        receipts.append([
            item["Movement ID"], item["Date"], item["Depot"], item["Tank"], item["Supplier"],
            item["Ordered (L)"], item["Dispatched (L)"], item["Accepted (L)"], item["Variance (L)"],
            None, _variance_label(item["Variance (L)"]), item["Reference"], item["Transport"],
            item["Vehicle"], item["Driver"], item["Recorded By"],
        ])
        current = receipts.max_row
        receipts.cell(current, 10, f'=IF(G{current}=0,"",H{current}/G{current})')
    if receipt_data.empty:
        receipts.append(["No supplier receipts found"] + [None] * (len(receipt_headers) - 1))
    receipt_last = receipts.max_row
    receipt_widths = [14, 13, 12, 15, 26, 15, 17, 15, 15, 13, 18, 22, 22, 16, 18, 16]
    _style_table_sheet(receipts, receipt_headers, 6, receipt_last, receipt_widths, "SupplierReceipts", profile)
    for row in range(7, receipt_last + 1):
        receipts.cell(row, 2).number_format = "dd mmm yyyy"
        for column in (6, 7, 8, 9):
            receipts.cell(row, column).number_format = '#,##0.00 "L";[Red]-#,##0.00 "L"'
        receipts.cell(row, 10).number_format = "0.00%"
    receipts.conditional_formatting.add(
        f"A7:P{receipt_last}", FormulaRule(formula=["$K7=\"Short Received\""], fill=PatternFill("solid", fgColor="FEE4E2"))
    )
    receipts.conditional_formatting.add(
        f"A7:P{receipt_last}", FormulaRule(formula=["$K7=\"Excess Received\""], fill=PatternFill("solid", fgColor="FEF0C7"))
    )
    receipts.conditional_formatting.add(
        f"A7:P{receipt_last}", FormulaRule(formula=["$K7=\"Matched\""], fill=PatternFill("solid", fgColor="DCFCE7"))
    )
    receipts.oddFooter.center.text = footer
    receipts.oddFooter.right.text = "Page &P of &N"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def ensure_operations_schema(conn):
    cursor=conn.cursor()
    try:
        for statement in (
            "ALTER TABLE transactions ADD COLUMN IF NOT EXISTS tank_transaction_id BIGINT",
            "ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS supplier_id INTEGER REFERENCES suppliers(id)",
            "ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS transport_method TEXT",
            "ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS vehicle_number TEXT",
            "ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS driver_name TEXT",
            "ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS ordered_liters REAL",
            "ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS dispatched_liters REAL",
            "ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS accepted_liters REAL",
            "ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS variance_liters REAL",
        ): cursor.execute(statement)
        conn.commit()
    except Exception: conn.rollback(); raise


def _tanks(conn):
    return pd.read_sql_query("""SELECT t.id,t.depot_id,CONCAT(d.code,' · ',t.code,' · ',t.name) AS label,
        d.code AS depot,t.code,t.name,p.name AS product,t.product_id,t.safe_capacity_liters,t.status,
        COALESCE(SUM(CASE WHEN tx.type='IN' THEN tx.liters ELSE -tx.liters END),0) AS balance
        FROM storage_tanks t JOIN depots d ON d.id=t.depot_id JOIN products p ON p.id=t.product_id
        LEFT JOIN tank_transactions tx ON tx.tank_id=t.id GROUP BY t.id,d.code,p.name ORDER BY d.code,t.code""",conn)


def _trucks(conn):
    return pd.read_sql_query("""SELECT t.id,CONCAT(t.emirate,' ',t.plate_code,' ',t.plate_number) AS label,
        p.name AS product,t.product_id,t.capacity_liters,t.operational_status AS status,
        COALESCE(SUM(CASE WHEN tx.type='IN' THEN tx.liters ELSE -tx.liters END),0) AS balance
        FROM trucks t LEFT JOIN products p ON p.id=t.product_id LEFT JOIN transactions tx ON tx.truck_id=t.id
        GROUP BY t.id,p.name ORDER BY label""",conn)


def _tank_balance(cursor,tank_id):
    cursor.execute("SELECT COALESCE(SUM(CASE WHEN type='IN' THEN liters ELSE -liters END),0) FROM tank_transactions WHERE tank_id=%s",(tank_id,))
    return float(cursor.fetchone()[0] or 0)


def _truck_balance(cursor,truck_id):
    cursor.execute("SELECT COALESCE(SUM(CASE WHEN type='IN' THEN liters ELSE -liters END),0) FROM transactions WHERE truck_id=%s",(truck_id,))
    return float(cursor.fetchone()[0] or 0)


def post_supplier_receipt(conn,tank_id,movement_at,ordered,dispatched,accepted,supplier_id,method,vehicle,driver,reference,notes,user,
                          purchase_type="Credit purchase",booking_id=None,release_id=None,unit_price=0.0):
    from period_close import assert_period_open
    assert_period_open(conn,movement_at)
    cursor=conn.cursor()
    try:
        cursor.execute("SELECT product_id,safe_capacity_liters,status FROM storage_tanks WHERE id=%s FOR UPDATE",(tank_id,)); tank=cursor.fetchone()
        if not tank: raise ValueError("Tank no longer exists.")
        if tank[2] not in ("AVAILABLE","RECEIVING"): raise ValueError("Tank must be Available or Receiving.")
        balance=_tank_balance(cursor,tank_id)
        if accepted<=0: raise ValueError("Accepted quantity must be greater than zero.")
        if balance+accepted>float(tank[1])+0.001: raise ValueError(f"Safe capacity exceeded. Available: {max(float(tank[1])-balance,0):,.2f} L.")
        if purchase_type=="Advance booking":
            if not booking_id: raise ValueError("Select the supplier booking used for this receipt.")
            cursor.execute("""SELECT supplier_id,product_id,status,booked_liters-COALESCE((SELECT SUM(accepted_liters) FROM tank_transactions WHERE booking_id=%s),0)
                FROM procurement_bookings WHERE id=%s FOR UPDATE""",(booking_id,booking_id)); booking=cursor.fetchone()
            if not booking or booking[2] not in ("OPEN","PARTIALLY_USED"): raise ValueError("The selected booking is no longer open.")
            if int(booking[0])!=int(supplier_id) or int(booking[1])!=int(tank[0]): raise ValueError("Booking supplier or product does not match this receipt.")
            if accepted>float(booking[3] or 0)+0.005: raise ValueError(f"Accepted quantity exceeds booking balance of {float(booking[3] or 0):,.2f} L.")
            if release_id:
                cursor.execute("""SELECT booking_id,status,released_liters-COALESCE((SELECT SUM(accepted_liters) FROM tank_transactions WHERE booking_release_id=%s),0)
                    FROM procurement_releases WHERE id=%s FOR UPDATE""",(release_id,release_id)); release=cursor.fetchone()
                if not release or int(release[0])!=int(booking_id) or release[1] not in ("OPEN","PARTIALLY_RECEIVED"): raise ValueError("The selected booking release is no longer open.")
                if accepted>float(release[2] or 0)+0.005: raise ValueError(f"Accepted quantity exceeds release balance of {float(release[2] or 0):,.2f} L.")
        variance=accepted-dispatched
        cursor.execute("""INSERT INTO tank_transactions(tank_id,movement_at,liters,type,movement_category,product_id,
            supplier_id,transport_method,vehicle_number,driver_name,ordered_liters,dispatched_liters,accepted_liters,
            variance_liters,reference,notes,created_by,purchase_type,booking_id,booking_release_id,unit_price)
            VALUES (%s,%s,%s,'IN','SUPPLIER_RECEIPT',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (tank_id,movement_at,accepted,tank[0],supplier_id,method,vehicle or None,driver or None,ordered,dispatched,accepted,variance,reference,notes or None,user,purchase_type,booking_id,release_id,unit_price))
        tx_id=cursor.fetchone()[0]; create_variance_claim(conn,tx_id,booking_id,supplier_id,variance,unit_price,user)
        if booking_id: refresh_booking_status(conn,booking_id,release_id)
        cursor.execute('INSERT INTO audit_log ("user",action,timestamp) VALUES (%s,%s,CURRENT_TIMESTAMP)',(user,f"RECEIVED {accepted:,.2f} L into tank; STX-{tx_id}")); conn.commit()
        record_event(conn,"SUPPLIER_RECEIPT","Storage Operations","Tank Transaction",tx_id,f"Accepted {accepted:,.2f} L; dispatch variance {variance:+,.2f} L")
        return tx_id,variance
    except Exception: conn.rollback(); raise


def post_tank_transfer(conn,source_id,destination_id,movement_at,liters,reference,notes,user):
    from period_close import assert_period_open
    assert_period_open(conn,movement_at)
    cursor=conn.cursor()
    try:
        cursor.execute("SELECT id,product_id,safe_capacity_liters,status FROM storage_tanks WHERE id=ANY(%s) ORDER BY id FOR UPDATE",(sorted([source_id,destination_id]),)); rows={r[0]:r for r in cursor.fetchall()}
        if len(rows)!=2: raise ValueError("One selected tank no longer exists.")
        source,destination=rows[source_id],rows[destination_id]
        if source[1]!=destination[1]: raise ValueError("Source and destination products must match.")
        if source[3] not in ("AVAILABLE","ISSUING") or destination[3] not in ("AVAILABLE","RECEIVING"): raise ValueError("Tank statuses do not permit this transfer.")
        source_balance=_tank_balance(cursor,source_id); destination_balance=_tank_balance(cursor,destination_id)
        if liters<=0 or liters>source_balance: raise ValueError(f"Insufficient source stock. Available: {source_balance:,.2f} L.")
        if destination_balance+liters>float(destination[2])+0.001: raise ValueError(f"Destination safe capacity exceeded. Available: {max(float(destination[2])-destination_balance,0):,.2f} L.")
        cursor.execute("""INSERT INTO tank_transactions(tank_id,movement_at,liters,type,movement_category,product_id,reference,notes,created_by)
            VALUES (%s,%s,%s,'OUT','TANK_TRANSFER',%s,%s,%s,%s) RETURNING id""",(source_id,movement_at,liters,source[1],reference,notes or None,user)); out_id=cursor.fetchone()[0]
        cursor.execute("""INSERT INTO tank_transactions(tank_id,movement_at,liters,type,movement_category,product_id,reference,notes,created_by)
            VALUES (%s,%s,%s,'IN','TANK_TRANSFER',%s,%s,%s,%s) RETURNING id""",(destination_id,movement_at,liters,destination[1],reference,notes or None,user)); in_id=cursor.fetchone()[0]
        cursor.execute("UPDATE tank_transactions SET partner_tank_transaction_id=%s WHERE id=%s",(in_id,out_id)); cursor.execute("UPDATE tank_transactions SET partner_tank_transaction_id=%s WHERE id=%s",(out_id,in_id)); conn.commit()
        record_event(conn,"TANK_TRANSFER","Storage Operations","Tank Transfer",out_id,f"Transferred {liters:,.2f} L; STX-{out_id} linked to STX-{in_id}"); return out_id,in_id
    except Exception: conn.rollback(); raise


def post_tank_truck(conn,tank_id,truck_id,movement_at,liters,direction,reference,notes,user):
    from period_close import assert_period_open
    assert_period_open(conn,movement_at)
    cursor=conn.cursor()
    try:
        cursor.execute("SELECT product_id,safe_capacity_liters,status FROM storage_tanks WHERE id=%s FOR UPDATE",(tank_id,)); tank=cursor.fetchone()
        cursor.execute("SELECT product_id,capacity_liters,operational_status FROM trucks WHERE id=%s FOR UPDATE",(truck_id,)); truck=cursor.fetchone()
        if not tank or not truck: raise ValueError("Tank or truck no longer exists.")
        if tank[0]!=truck[0]: raise ValueError("Tank and truck fuel products must match.")
        if truck[2]!="ACTIVE": raise ValueError("Truck must be Active.")
        tank_balance=_tank_balance(cursor,tank_id); truck_balance=_truck_balance(cursor,truck_id)
        if liters<=0: raise ValueError("Quantity must be greater than zero.")
        if direction=="TANK_TO_TRUCK":
            if tank[2] not in ("AVAILABLE","ISSUING"): raise ValueError("Tank is not available for issuing.")
            if liters>tank_balance: raise ValueError(f"Insufficient tank stock. Available: {tank_balance:,.2f} L.")
            if truck[1] and truck_balance+liters>float(truck[1])+0.001: raise ValueError(f"Truck capacity exceeded. Available: {max(float(truck[1])-truck_balance,0):,.2f} L.")
            tank_type,truck_type,category="OUT","IN","TANK_TO_TRUCK"
        else:
            if tank[2] not in ("AVAILABLE","RECEIVING"): raise ValueError("Tank is not available for receiving.")
            if liters>truck_balance: raise ValueError(f"Insufficient truck stock. Available: {truck_balance:,.2f} L.")
            if tank_balance+liters>float(tank[1])+0.001: raise ValueError(f"Tank safe capacity exceeded. Available: {max(float(tank[1])-tank_balance,0):,.2f} L.")
            tank_type,truck_type,category="IN","OUT","TRUCK_TO_TANK"
        cursor.execute("""INSERT INTO tank_transactions(tank_id,movement_at,liters,type,movement_category,product_id,reference,notes,created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",(tank_id,movement_at,liters,tank_type,category,tank[0],reference,notes or None,user)); tank_tx=cursor.fetchone()[0]
        cursor.execute("""INSERT INTO transactions(truck_id,date,liters,type,created_by,product_id,movement_category,tank_transaction_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",(truck_id,str(movement_at.date()),liters,truck_type,user,truck[0],category,tank_tx)); truck_tx=cursor.fetchone()[0]
        cursor.execute("UPDATE tank_transactions SET truck_transaction_id=%s WHERE id=%s",(truck_tx,tank_tx)); conn.commit()
        record_event(conn,category,"Storage Operations","Tank Transaction",tank_tx,f"Posted {liters:,.2f} L; STX-{tank_tx} linked to TX-{truck_tx}"); return tank_tx,truck_tx
    except Exception: conn.rollback(); raise


def render_storage_operations(conn):
    ensure_operations_schema(conn); ensure_procurement_schema(conn); page_header("Storage Operations","Receive and move fuel safely between suppliers, tanks and trucks.")
    tanks=_tanks(conn); trucks=_trucks(conn); user=st.session_state.get("user","System")
    if tanks.empty: st.warning("Create a depot and storage tank first."); return
    tab_receipt,tab_transfer,tab_loading,tab_return,tab_history=st.tabs(["Supplier receipt","Tank transfer","Load truck","Truck return","Movement history"])
    tank_map=dict(zip(tanks["label"],tanks["id"])); truck_map=dict(zip(trucks["label"],trucks["id"])) if not trucks.empty else {}
    with tab_receipt:
        suppliers=pd.read_sql_query("SELECT id,name FROM suppliers WHERE COALESCE(status,'ACTIVE')='ACTIVE' ORDER BY name",conn); supplier_map=dict(zip(suppliers["name"],suppliers["id"]))
        selected=st.selectbox("Receiving tank",list(tank_map),key="receipt_tank"); tank=tanks[tanks["id"]==tank_map[selected]].iloc[0]; st.info(f"Current {tank['balance']:,.2f} L · Available {max(tank['safe_capacity_liters']-tank['balance'],0):,.2f} L")
        a,b=st.columns(2); supplier=a.selectbox("Supplier",list(supplier_map),key="receipt_supplier"); purchase_type=b.selectbox("Purchase source",["Advance booking","Credit purchase","Cash purchase"],key="receipt_purchase_type")
        booking_id=None; release_id=None; booking_price=0.0
        if purchase_type=="Advance booking":
            available=booking_options(conn,supplier_map[supplier]); available=available[available["product_id"]==int(tank["product_id"])] if not available.empty else available
            if available.empty: st.warning("No open booking is available for this supplier and product.")
            else:
                booking_map=dict(zip(available["label"],available["id"])); booking_label=st.selectbox("Supplier booking",list(booking_map)); booking_id=int(booking_map[booking_label]); booking_row=available[available["id"]==booking_id].iloc[0]; booking_price=float(booking_row["unit_price"] or 0)
                releases=release_options(conn,booking_id)
                if not releases.empty:
                    releases["label"]=releases.apply(lambda r:f"{r['release_number']} · {max(r['released_liters']-r['received_liters'],0):,.2f} L remaining",axis=1); release_map=dict(zip(releases["label"],releases["id"])); release_label=st.selectbox("Booking release",["No specific release"]+list(release_map)); release_id=None if release_label=="No specific release" else int(release_map[release_label])
        with st.form("supplier_receipt"):
            method=st.selectbox("Transport method",["Supplier delivery","Company collection","Third-party transporter","Other"])
            c,d,e=st.columns(3); ordered=c.number_input("Ordered/released quantity",min_value=0.0); dispatched=d.number_input("Supplier dispatched quantity",min_value=0.0); accepted=e.number_input("Accepted into tank",min_value=0.0)
            f,g=st.columns(2); vehicle=f.text_input("Vehicle number"); driver=g.text_input("Driver name"); unit_price=st.number_input("Unit price",min_value=0.0,value=booking_price,format="%.4f",disabled=purchase_type=="Advance booking"); reference=st.text_input("Delivery note / ticket reference"); notes=st.text_area("Receipt notes"); submitted=st.form_submit_button("Post supplier receipt",type="primary")
        if submitted:
            if not reference.strip(): st.error("Delivery note or ticket reference is required.")
            elif purchase_type=="Advance booking" and not booking_id: st.error("An open booking is required for an advance-booking receipt.")
            else:
                try:
                    receipt_price=booking_price if purchase_type=="Advance booking" else unit_price
                    receipt_value=float(accepted)*float(receipt_price or 0)
                    controlled=needs_approval(conn,"SUPPLIER_RECEIPT_QUANTITY",accepted) or needs_approval(conn,"SUPPLIER_RECEIPT_VALUE",receipt_value)
                    if controlled:
                        payload={"tank_id":int(tank["id"]),"movement_at":datetime.now().isoformat(),"ordered":ordered,"dispatched":dispatched,"accepted":accepted,
                                 "supplier_id":int(supplier_map[supplier]),"method":method,"vehicle":vehicle,"driver":driver,"reference":reference.strip(),"notes":notes,
                                 "purchase_type":purchase_type,"booking_id":booking_id,"release_id":release_id,"unit_price":receipt_price}
                        request_id=submit_approval_request(conn,"SUPPLIER_RECEIPT",f"Supplier receipt · {selected} · {accepted:,.2f} L",payload,user,accepted,receipt_value)
                        st.success(f"AP-{request_id} submitted for approval. Stock has not changed."); st.rerun()
                    else:
                        tx,variance=post_supplier_receipt(conn,int(tank["id"]),datetime.now(),ordered,dispatched,accepted,supplier_map[supplier],method,vehicle,driver,reference.strip(),notes,user,purchase_type,booking_id,release_id,receipt_price)
                        st.success(f"STX-{tx} posted. Dispatch-to-accepted variance: {variance:+,.2f} L."); st.rerun()
                except Exception as error: st.error(str(error))
    with tab_transfer:
        source=st.selectbox("Source tank",list(tank_map),key="tank_source"); destinations=[x for x in tank_map if x!=source]; destination=st.selectbox("Destination tank",destinations,key="tank_destination") if destinations else None
        source_row=tanks[tanks["id"]==tank_map[source]].iloc[0]; st.info(f"Source stock: {source_row['balance']:,.2f} L")
        with st.form("tank_transfer"):
            liters=st.number_input("Transfer quantity",min_value=0.0); reference=st.text_input("Transfer reference"); notes=st.text_area("Transfer notes"); submit=st.form_submit_button("Post tank transfer",type="primary")
        if submit:
            if not destination: st.error("A destination tank is required.")
            elif not reference.strip(): st.error("Transfer reference is required.")
            else:
                try: out_id,in_id=post_tank_transfer(conn,tank_map[source],tank_map[destination],datetime.now(),liters,reference.strip(),notes,user); st.success(f"STX-{out_id} OUT linked to STX-{in_id} IN."); st.rerun()
                except Exception as error: st.error(str(error))
    with tab_loading:
        if not truck_map: st.info("No trucks available.")
        else:
            with st.form("tank_to_truck"):
                tank_label=st.selectbox("Source tank",list(tank_map)); truck_label=st.selectbox("Destination truck",list(truck_map)); liters=st.number_input("Loading quantity",min_value=0.0); reference=st.text_input("Loading ticket reference"); notes=st.text_area("Loading notes"); submit=st.form_submit_button("Load truck",type="primary")
            if submit:
                if not reference.strip(): st.error("Loading ticket reference is required.")
                else:
                    try: stx,tx=post_tank_truck(conn,tank_map[tank_label],truck_map[truck_label],datetime.now(),liters,"TANK_TO_TRUCK",reference.strip(),notes,user); st.success(f"STX-{stx} tank OUT linked to TX-{tx} truck IN."); st.rerun()
                    except Exception as error: st.error(str(error))
    with tab_return:
        if not truck_map: st.info("No trucks available.")
        else:
            with st.form("truck_to_tank"):
                truck_label=st.selectbox("Source truck",list(truck_map),key="return_truck"); tank_label=st.selectbox("Receiving tank",list(tank_map),key="return_tank"); liters=st.number_input("Return quantity",min_value=0.0); reference=st.text_input("Return reference"); notes=st.text_area("Return notes"); submit=st.form_submit_button("Post truck return",type="primary")
            if submit:
                if not reference.strip(): st.error("Return reference is required.")
                else:
                    try: stx,tx=post_tank_truck(conn,tank_map[tank_label],truck_map[truck_label],datetime.now(),liters,"TRUCK_TO_TANK",reference.strip(),notes,user); st.success(f"TX-{tx} truck OUT linked to STX-{stx} tank IN."); st.rerun()
                    except Exception as error: st.error(str(error))
    with tab_history:
        history=pd.read_sql_query("""SELECT tx.id,tx.movement_at,d.code AS depot,t.code AS tank,p.name AS product,tx.type,tx.liters,tx.movement_category,
            s.name AS supplier,tx.ordered_liters,tx.dispatched_liters,tx.accepted_liters,tx.variance_liters,tx.reference,tx.vehicle_number,tx.driver_name,
            tx.transport_method,tx.partner_tank_transaction_id,tx.truck_transaction_id,tx.record_status,tx.created_by,tx.notes
            FROM tank_transactions tx JOIN storage_tanks t ON t.id=tx.tank_id
            JOIN depots d ON d.id=t.depot_id LEFT JOIN products p ON p.id=tx.product_id LEFT JOIN suppliers s ON s.id=tx.supplier_id ORDER BY tx.id DESC""",conn)
        display=_prepare_export(history)
        display_columns=["Movement ID","Date","Time","Depot","Tank","Direction","Quantity (L)","Movement","Supplier","Variance (L)","Reference","Recorded By"]
        st.dataframe(
            display[display_columns],use_container_width=True,hide_index=True,height=480,
            column_config={
                "Quantity (L)": st.column_config.NumberColumn(format="%.2f L"),
                "Variance (L)": st.column_config.NumberColumn(format="%+.2f L"),
            },
        )
        if not history.empty:
            profile=st.session_state.get("company_profile",{})
            report=build_operations_report(history,profile)
            st.download_button(
                "Download professional operations report",report,
                f"storage_operations_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
