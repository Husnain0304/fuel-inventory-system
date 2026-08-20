from datetime import date, datetime, time
import io

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from approval_workflow import submit_approval_request
from audit import record_event
from procurement import ensure_procurement_schema
from ui import page_header


def ensure_valuation_schema(conn):
    ensure_procurement_schema(conn)
    cursor=conn.cursor()
    try:
        cursor.execute("""CREATE TABLE IF NOT EXISTS product_cost_policies(
            id BIGSERIAL PRIMARY KEY,product_id INTEGER NOT NULL REFERENCES products(id),
            effective_from DATE NOT NULL,default_unit_cost REAL NOT NULL CHECK(default_unit_cost>=0),
            reason TEXT NOT NULL,reference TEXT NOT NULL,status TEXT NOT NULL DEFAULT 'ACTIVE',
            approved_request_id BIGINT REFERENCES approval_requests(id),created_by TEXT NOT NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP)""")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_cost_policy_product ON product_cost_policies(product_id,effective_from DESC)")
        conn.commit()
    except Exception:
        conn.rollback(); raise


def _default_costs(conn,as_of):
    data=pd.read_sql_query("""SELECT DISTINCT ON (product_id) product_id,default_unit_cost,effective_from
        FROM product_cost_policies WHERE status='ACTIVE' AND effective_from<=%s
        ORDER BY product_id,effective_from DESC,id DESC""",conn,params=[as_of])
    result={int(r.product_id):float(r.default_unit_cost) for r in data.itertuples()}
    fallback=pd.read_sql_query("SELECT cost_per_liter FROM settings ORDER BY id LIMIT 1",conn)
    return result,float(fallback.iloc[0,0] or 0) if not fallback.empty else 0.0


def _asset_master(conn):
    tanks=pd.read_sql_query("""SELECT t.id,CONCAT(d.code,' · ',t.code,' · ',t.name) AS asset,
        d.code AS depot,p.name AS product,t.product_id,'TANK' AS asset_type
        FROM storage_tanks t JOIN depots d ON d.id=t.depot_id JOIN products p ON p.id=t.product_id""",conn)
    trucks=pd.read_sql_query("""SELECT t.id,CONCAT(t.emirate,' ',t.plate_code,' ',t.plate_number) AS asset,
        'Fleet' AS depot,COALESCE(p.name,'Unassigned') AS product,t.product_id,'TRUCK' AS asset_type
        FROM trucks t LEFT JOIN products p ON p.id=t.product_id""",conn)
    return tanks,trucks


def calculate_valuation(conn,as_of_date):
    policies,fallback=_default_costs(conn,as_of_date)
    tanks,trucks=_asset_master(conn)
    tank_meta={int(r.id):r for r in tanks.itertuples()}; truck_meta={int(r.id):r for r in trucks.itertuples()}
    end_at=datetime.combine(as_of_date,time.max)
    tank_tx=pd.read_sql_query("""SELECT id,tank_id,movement_at,liters,type,movement_category,product_id,
        partner_tank_transaction_id,truck_transaction_id,unit_price,supplier_id,reference
        FROM tank_transactions WHERE movement_at<=%s AND COALESCE(record_status,'POSTED')='POSTED'""",conn,params=[end_at])
    truck_tx=pd.read_sql_query("""SELECT id,truck_id,date,liters,type,movement_category,product_id,
        transfer_partner_id,tank_transaction_id,supplier_id,ticket_number
        FROM transactions WHERE date::date<=%s AND COALESCE(record_status,'POSTED')='POSTED'""",conn,params=[as_of_date])
    linked_time={int(r.truck_transaction_id):pd.to_datetime(r.movement_at,utc=True) for r in tank_tx.itertuples() if pd.notna(r.truck_transaction_id)}
    events=[]
    for r in tank_tx.itertuples():
        category=str(r.movement_category or "STANDARD")
        priority=10
        if category=="TANK_TRANSFER": priority=10 if r.type=="OUT" else 30
        elif category=="TANK_TO_TRUCK": priority=10
        elif category=="TRUCK_TO_TANK": priority=30
        events.append({"sort_time":pd.to_datetime(r.movement_at,utc=True),"priority":priority,"asset_type":"TANK","asset_id":int(r.tank_id),"tx_id":int(r.id),"type":r.type,"liters":float(r.liters),"category":category,"product_id":int(r.product_id) if pd.notna(r.product_id) else None,"partner_id":int(r.partner_tank_transaction_id) if pd.notna(r.partner_tank_transaction_id) else None,"linked_id":int(r.truck_transaction_id) if pd.notna(r.truck_transaction_id) else None,"explicit_cost":float(r.unit_price) if pd.notna(r.unit_price) and float(r.unit_price)>0 else None,"reference":r.reference})
    for r in truck_tx.itertuples():
        category=str(r.movement_category or "STANDARD")
        linked=int(r.tank_transaction_id) if pd.notna(r.tank_transaction_id) else None
        event_time=linked_time.get(int(r.id),pd.to_datetime(str(r.date),utc=True)+pd.Timedelta(hours=12))
        priority=20
        if category=="TRUCK_TO_TANK": priority=10
        elif category=="TANK_TO_TRUCK": priority=20
        events.append({"sort_time":event_time,"priority":priority,"asset_type":"TRUCK","asset_id":int(r.truck_id),"tx_id":int(r.id),"type":r.type,"liters":float(r.liters),"category":category,"product_id":int(r.product_id) if pd.notna(r.product_id) else None,"partner_id":int(r.transfer_partner_id) if pd.notna(r.transfer_partner_id) else None,"linked_id":linked,"explicit_cost":None,"reference":r.ticket_number})
    events.sort(key=lambda x:(x["sort_time"],x["priority"],x["asset_type"],x["tx_id"]))
    states={}; transfer_cost={}; ledger=[]
    for event in events:
        key=(event["asset_type"],event["asset_id"]); state=states.setdefault(key,{"qty":0.0,"value":0.0})
        meta=tank_meta.get(event["asset_id"]) if event["asset_type"]=="TANK" else truck_meta.get(event["asset_id"])
        product_id=event["product_id"] or (int(meta.product_id) if meta and pd.notna(meta.product_id) else None)
        fallback_cost=float(policies.get(product_id,fallback))
        opening_qty,opening_value=state["qty"],state["value"]
        current_wac=(opening_value/opening_qty) if opening_qty>0.000001 else fallback_cost
        source="Moving weighted average"
        if event["type"]=="OUT":
            unit_cost=current_wac; movement_value=event["liters"]*unit_cost
            state["qty"]-=event["liters"]; state["value"]-=movement_value
            if abs(state["qty"])<0.000001: state["qty"],state["value"]=0.0,0.0
            transfer_cost[(event["asset_type"],event["tx_id"])]=unit_cost
            if event["partner_id"]: transfer_cost[(event["asset_type"],event["partner_id"])]=unit_cost
            if event["linked_id"]:
                other="TRUCK" if event["asset_type"]=="TANK" else "TANK"; transfer_cost[(other,event["linked_id"])]=unit_cost
        else:
            linked_cost=transfer_cost.get((event["asset_type"],event["tx_id"]))
            if event["explicit_cost"] is not None: unit_cost=event["explicit_cost"]; source="Supplier receipt price"
            elif linked_cost is not None: unit_cost=linked_cost; source="Linked transfer cost"
            else: unit_cost=fallback_cost; source="Approved fallback cost" if product_id in policies else "System default cost"
            movement_value=event["liters"]*unit_cost; state["qty"]+=event["liters"]; state["value"]+=movement_value
        if state["qty"] < -0.005:
            state["value"]=state["qty"]*unit_cost
        closing_wac=(state["value"]/state["qty"]) if state["qty"]>0.000001 else 0.0
        ledger.append({"Date & Time":event["sort_time"].tz_convert("Asia/Dubai").tz_localize(None),"Asset Type":event["asset_type"],"Asset":meta.asset if meta else str(event["asset_id"]),"Depot":meta.depot if meta else "","Product":meta.product if meta else "","Transaction":f"{'STX' if event['asset_type']=='TANK' else 'TX'}-{event['tx_id']}","Direction":event["type"],"Movement":event["category"],"Quantity (L)":event["liters"],"Unit Cost":unit_cost,"Movement Value":movement_value,"Closing Quantity (L)":state["qty"],"Closing Value":state["value"],"Closing WAC":closing_wac,"Cost Source":source,"Reference":event["reference"]})
    position=[]
    for key,state in states.items():
        asset_type,asset_id=key; meta=tank_meta.get(asset_id) if asset_type=="TANK" else truck_meta.get(asset_id)
        if not meta: continue
        position.append({"Asset Type":asset_type,"Depot":meta.depot,"Asset":meta.asset,"Product":meta.product,"Quantity (L)":state["qty"],"Unit Cost":(state["value"]/state["qty"] if state["qty"]>0.000001 else 0.0),"Inventory Value":state["value"]})
    ledger_df=pd.DataFrame(ledger); position_df=pd.DataFrame(position)
    claims=pd.read_sql_query("""SELECT c.id,s.name AS supplier,c.claim_liters,c.unit_price,c.claim_amount,c.status,
        c.credit_note_number,c.created_at FROM supplier_claims c JOIN suppliers s ON s.id=c.supplier_id
        WHERE DATE(c.created_at AT TIME ZONE 'Asia/Dubai')<=%s ORDER BY c.id DESC""",conn,params=[as_of_date])
    commitments=pd.read_sql_query("""SELECT b.id,b.booking_number,s.name AS supplier,p.name AS product,b.status,
        b.booked_liters,b.unit_price,b.booked_liters*b.unit_price AS booked_value,
        COALESCE(SUM(tx.accepted_liters),0) AS received_liters,
        COALESCE(SUM(tx.accepted_liters),0)*b.unit_price AS received_value,
        GREATEST(b.booked_liters-COALESCE(SUM(tx.accepted_liters),0),0)*b.unit_price AS open_commitment
        FROM procurement_bookings b JOIN suppliers s ON s.id=b.supplier_id JOIN products p ON p.id=b.product_id
        LEFT JOIN tank_transactions tx ON tx.booking_id=b.id WHERE b.booking_date<=%s
        GROUP BY b.id,s.name,p.name ORDER BY b.id DESC""",conn,params=[as_of_date])
    return position_df,ledger_df,claims,commitments


def _style_sheet(ws,title,subtitle,profile,column_count):
    primary=str(profile.get("primary_color","#8C1C1C")).replace("#",""); secondary=str(profile.get("secondary_color","#172033")).replace("#","")
    ws.sheet_view.showGridLines=False; ws.merge_cells(start_row=1,start_column=1,end_row=2,end_column=column_count)
    ws.cell(1,1,title); ws.cell(1,1).fill=PatternFill("solid",fgColor=secondary); ws.cell(1,1).font=Font(name="Aptos Display",size=20,bold=True,color="FFFFFF"); ws.cell(1,1).alignment=Alignment(vertical="center")
    ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=column_count); ws.cell(3,1,subtitle); ws.cell(3,1).fill=PatternFill("solid",fgColor=primary); ws.cell(3,1).font=Font(name="Aptos",size=10,color="FFFFFF")


def _write_table(ws,frame,start_row,name,profile):
    primary=str(profile.get("primary_color","#8C1C1C")).replace("#",""); headers=list(frame.columns)
    for col,value in enumerate(headers,1):
        cell=ws.cell(start_row,col,str(value)); cell.fill=PatternFill("solid",fgColor=primary); cell.font=Font(bold=True,color="FFFFFF"); cell.alignment=Alignment(wrap_text=True)
    if frame.empty: ws.cell(start_row+1,1,"No records available"); return start_row+1
    for row_idx,row in enumerate(frame.itertuples(index=False,name=None),start_row+1):
        for col_idx,value in enumerate(row,1):
            if isinstance(value,pd.Timestamp): value=value.to_pydatetime().replace(tzinfo=None)
            ws.cell(row_idx,col_idx,value)
    end=start_row+len(frame); table=Table(displayName=name,ref=f"A{start_row}:{get_column_letter(len(headers))}{end}"); table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True,showColumnStripes=False); ws.add_table(table); ws.freeze_panes=f"A{start_row+1}"
    for idx,header in enumerate(headers,1):
        width=15
        if header in ("Asset","Supplier","Product","Cost Source","Reference","Movement"): width=24
        ws.column_dimensions[get_column_letter(idx)].width=width
        if "Date" in header: ws.column_dimensions[get_column_letter(idx)].width=19
        if any(x in header for x in ("Value","Cost","Amount","Commitment")):
            for row in range(start_row+1,end+1): ws.cell(row,idx).number_format='#,##0.00;[Red](#,##0.00);-'
        if "Quantity" in header or "Liters" in header or "liters" in header:
            for row in range(start_row+1,end+1): ws.cell(row,idx).number_format='#,##0.00 "L";[Red](#,##0.00 "L");-'
    return end


def build_valuation_report(position,ledger,claims,commitments,profile,as_of):
    wb=Workbook(); summary=wb.active; summary.title="Executive Summary"; currency=profile.get("currency","AED")
    _style_sheet(summary,f"{profile.get('company_name','Company')} | Inventory Valuation",f"Moving weighted-average valuation as of {as_of:%d %b %Y} · Currency: {currency}",profile,8)
    total_qty=float(position["Quantity (L)"].sum()) if not position.empty else 0; total_value=float(position["Inventory Value"].sum()) if not position.empty else 0
    open_claims=float(claims.loc[~claims["status"].isin(["CLOSED","REJECTED"]),"claim_amount"].sum()) if not claims.empty else 0; open_commit=float(commitments.loc[~commitments["status"].isin(["COMPLETED","CANCELLED","EXPIRED"]),"open_commitment"].sum()) if not commitments.empty else 0
    cards=[("Inventory quantity",total_qty,'#,##0.00 "L"'),("Inventory value",total_value,'#,##0.00'),("Open supplier claims",open_claims,'#,##0.00'),("Open booking commitments",open_commit,'#,##0.00')]
    for i,(label,value,fmt) in enumerate(cards):
        col=1+i*2; summary.merge_cells(start_row=5,start_column=col,end_row=5,end_column=col+1); summary.cell(5,col,label); summary.cell(5,col).fill=PatternFill("solid",fgColor="E8EDF5"); summary.cell(5,col).font=Font(bold=True,color="475467")
        summary.merge_cells(start_row=6,start_column=col,end_row=7,end_column=col+1); summary.cell(6,col,value); summary.cell(6,col).number_format=fmt; summary.cell(6,col).font=Font(size=18,bold=True,color="172033"); summary.cell(6,col).alignment=Alignment(vertical="center")
    by_product=position.groupby("Product",as_index=False).agg(**{"Quantity (L)":("Quantity (L)","sum"),"Inventory Value":("Inventory Value","sum")}) if not position.empty else pd.DataFrame(columns=["Product","Quantity (L)","Inventory Value"])
    _write_table(summary,by_product,10,"SummaryByProduct",profile); summary.column_dimensions["A"].width=26
    assumptions=wb.create_sheet("Methodology & Checks"); _style_sheet(assumptions,"Valuation Methodology & Control Checks","Audit notes, assumptions and reconciliation controls",profile,6)
    assumptions.append([]); assumptions.append(["Control","Actual","Expected","Difference","Status","Notes"])
    checks=[("Position quantity vs ledger",total_qty,float(ledger.groupby(["Asset Type","Asset"])["Closing Quantity (L)"].last().sum()) if not ledger.empty else 0,"Latest movement balance must equal position"),("Position value vs ledger",total_value,float(ledger.groupby(["Asset Type","Asset"])["Closing Value"].last().sum()) if not ledger.empty else 0,"Latest movement value must equal position"),("Negative asset balances",float((position["Quantity (L)"]<-0.005).sum()) if not position.empty else 0,0,"Negative stock requires operational investigation")]
    for idx,(label,actual,expected,note) in enumerate(checks,6):
        assumptions.cell(idx,1,label); assumptions.cell(idx,2,actual); assumptions.cell(idx,3,expected); assumptions.cell(idx,4,f"=B{idx}-C{idx}"); assumptions.cell(idx,5,f'=IF(ABS(D{idx})<0.01,"OK","REVIEW")'); assumptions.cell(idx,6,note)
    assumptions["A11"]="Method"; assumptions["B11"]="Perpetual moving weighted-average cost"
    assumptions["A12"]="Supplier receipts"; assumptions["B12"]="Accepted liters × recorded unit price"
    assumptions["A13"]="Transfers"; assumptions["B13"]="Carry the issuing asset's weighted-average cost"
    assumptions["A14"]="Missing historical cost"; assumptions["B14"]="Most recent approved product fallback cost; otherwise system default"
    assumptions["A15"]="Important"; assumptions["B15"]="Management inventory valuation; accounting posting remains in the finance system"
    assumptions.column_dimensions["A"].width=30; assumptions.column_dimensions["B"].width=34; assumptions.column_dimensions["F"].width=48
    for name,frame,title in (("Inventory Position",position,"Inventory Position"),("Cost Movement Ledger",ledger,"Cost Movement Ledger"),("Supplier Exposure",commitments,"Supplier Commitments"),("Claims",claims,"Supplier Claims")):
        ws=wb.create_sheet(name); _style_sheet(ws,title,f"As of {as_of:%d %b %Y} · {currency}",profile,max(len(frame.columns),6)); _write_table(ws,frame,5,name.replace(" ","")[:20]+"Table",profile)
    output=io.BytesIO(); wb.save(output); return output.getvalue()


def render_valuation(conn):
    profile=st.session_state.get("company_profile",{}); currency=profile.get("currency","AED")
    page_header("Inventory Costing & Valuation","Understand the financial value of fuel across tanks, trucks, suppliers and commitments.")
    as_of=st.date_input("Valuation date",date.today()); position,ledger,claims,commitments=calculate_valuation(conn,as_of)
    total_qty=float(position["Quantity (L)"].sum()) if not position.empty else 0; total_value=float(position["Inventory Value"].sum()) if not position.empty else 0
    issued=ledger[ledger["Direction"]=="OUT"] if not ledger.empty else ledger; issue_value=float(issued["Movement Value"].sum()) if not issued.empty else 0
    open_claims=float(claims.loc[~claims["status"].isin(["CLOSED","REJECTED"]),"claim_amount"].sum()) if not claims.empty else 0
    a,b,c,d=st.columns(4); a.metric("Inventory quantity",f"{total_qty:,.2f} L"); b.metric("Inventory value",f"{currency} {total_value:,.2f}"); c.metric("Historical issued value",f"{currency} {issue_value:,.2f}"); d.metric("Open claim value",f"{currency} {open_claims:,.2f}")
    overview,movements,exposure,cost_policy,report=st.tabs(["Valuation position","Cost movement ledger","Supplier exposure","Cost policies","Financial report"])
    with overview:
        if position.empty: st.info("No inventory movements are available for valuation.")
        else:
            x,y=st.columns(2); selected_type=x.multiselect("Asset type",sorted(position["Asset Type"].unique())); selected_product=y.multiselect("Product",sorted(position["Product"].unique())); view=position.copy()
            if selected_type: view=view[view["Asset Type"].isin(selected_type)]
            if selected_product: view=view[view["Product"].isin(selected_product)]
            st.dataframe(view,use_container_width=True,hide_index=True,height=470,column_config={"Quantity (L)":st.column_config.NumberColumn(format="%.2f L"),"Unit Cost":st.column_config.NumberColumn(format=f"{currency} %.4f"),"Inventory Value":st.column_config.NumberColumn(format=f"{currency} %.2f")})
    with movements:
        if ledger.empty: st.info("No costed movements are available.")
        else:
            kinds=st.multiselect("Movement type",sorted(ledger["Movement"].unique())); view=ledger[ledger["Movement"].isin(kinds)] if kinds else ledger
            st.dataframe(view,use_container_width=True,hide_index=True,height=520,column_config={"Quantity (L)":st.column_config.NumberColumn(format="%.2f L"),"Unit Cost":st.column_config.NumberColumn(format=f"{currency} %.4f"),"Movement Value":st.column_config.NumberColumn(format=f"{currency} %.2f"),"Closing Value":st.column_config.NumberColumn(format=f"{currency} %.2f"),"Closing WAC":st.column_config.NumberColumn(format=f"{currency} %.4f")})
    with exposure:
        st.subheader("Supplier booking commitments"); st.dataframe(commitments,use_container_width=True,hide_index=True,height=300)
        st.subheader("Claims and expected credits"); st.dataframe(claims,use_container_width=True,hide_index=True,height=300)
    with cost_policy:
        policies=pd.read_sql_query("""SELECT c.id,p.name AS product,c.effective_from,c.default_unit_cost,c.reason,c.reference,
            c.status,c.created_by,c.created_at FROM product_cost_policies c JOIN products p ON p.id=c.product_id ORDER BY c.effective_from DESC,c.id DESC""",conn)
        st.caption("Fallback costs apply only when an older inventory receipt has no recorded unit price. They never overwrite supplier receipt prices.")
        st.dataframe(policies,use_container_width=True,hide_index=True,height=300)
        products=pd.read_sql_query("SELECT id,name FROM products WHERE active=TRUE ORDER BY name",conn); product_map=dict(zip(products["name"],products["id"]))
        with st.form("cost_policy_request"):
            product=st.selectbox("Product",list(product_map)); effective=st.date_input("Effective from",date.today()); cost=st.number_input("Fallback unit cost",min_value=0.0,format="%.4f"); reason=st.text_area("Reason for cost change"); reference=st.text_input("Supporting reference"); submit=st.form_submit_button("Submit cost policy for approval",type="primary")
        if submit:
            if cost<=0 or len(reason.strip())<5 or not reference.strip(): st.error("Enter a positive unit cost, clear reason and supporting reference.")
            else:
                payload={"product_id":int(product_map[product]),"effective_from":str(effective),"unit_cost":cost,"reason":reason.strip(),"reference":reference.strip()}
                request_id=submit_approval_request(conn,"COST_POLICY_CHANGE",f"Cost policy · {product} · {currency} {cost:.4f}/L",payload,st.session_state["user"],monetary_value=cost)
                st.success(f"AP-{request_id} submitted. The cost policy remains unchanged until approval."); st.rerun()
    with report:
        report_bytes=build_valuation_report(position,ledger,claims,commitments,profile,as_of)
        st.download_button("Download financial valuation workbook",report_bytes,f"inventory_valuation_{as_of:%Y%m%d}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
        st.caption("The workbook includes executive summary, position, movement ledger, supplier commitments, claims, methodology and control checks.")
