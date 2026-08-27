from io import BytesIO
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill
from audit import record_event
from ui import page_header

EDIT={"ADMIN","INVENTORY_MANAGER","PROCUREMENT_USER"}; APPROVE={"ADMIN","INVENTORY_MANAGER","APPROVER"}

def ensure_receipt_cost_schema(conn):
 c=conn.cursor()
 try:
  c.execute("ALTER TABLE tank_transactions ADD COLUMN IF NOT EXISTS landed_unit_cost REAL")
  c.execute("""CREATE TABLE IF NOT EXISTS supplier_receipt_invoices(id BIGSERIAL PRIMARY KEY,tank_transaction_id BIGINT UNIQUE NOT NULL REFERENCES tank_transactions(id),supplier_invoice_number TEXT NOT NULL,invoice_date DATE NOT NULL,invoiced_liters REAL NOT NULL,invoiced_unit_price REAL NOT NULL,freight_cost REAL DEFAULT 0,testing_cost REAL DEFAULT 0,handling_cost REAL DEFAULT 0,other_cost REAL DEFAULT 0,tax_amount REAL DEFAULT 0,invoice_total REAL NOT NULL,quantity_variance REAL NOT NULL,price_variance REAL NOT NULL,landed_cost_total REAL NOT NULL,landed_unit_cost REAL NOT NULL,status TEXT DEFAULT 'DRAFT',exception_reason TEXT,document_reference TEXT,created_by TEXT,created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,approved_by TEXT,approved_at TIMESTAMPTZ)""")
  c.execute("CREATE INDEX IF NOT EXISTS idx_receipt_invoice_status ON supplier_receipt_invoices(status,invoice_date)"); conn.commit()
 except Exception: conn.rollback(); raise

def _receipts(conn): return pd.read_sql_query("""SELECT x.id,x.movement_at,s.name supplier,p.name product,COALESCE(x.accepted_liters,x.liters) accepted_liters,COALESCE(x.unit_price,b.unit_price,0) expected_unit_price,x.reference,b.booking_number FROM tank_transactions x JOIN suppliers s ON s.id=x.supplier_id JOIN products p ON p.id=x.product_id LEFT JOIN procurement_bookings b ON b.id=x.booking_id LEFT JOIN supplier_receipt_invoices i ON i.tank_transaction_id=x.id WHERE x.movement_category='SUPPLIER_RECEIPT' AND i.id IS NULL ORDER BY x.id DESC""",conn)

def _invoices(conn): return pd.read_sql_query("""SELECT i.id,x.id receipt_id,i.supplier_invoice_number,i.invoice_date,s.name supplier,p.name product,COALESCE(x.accepted_liters,x.liters) accepted_liters,i.invoiced_liters,i.invoiced_unit_price,i.quantity_variance,i.price_variance,i.freight_cost,i.testing_cost,i.handling_cost,i.other_cost,i.tax_amount,i.invoice_total,i.landed_cost_total,i.landed_unit_cost,i.status,i.exception_reason,i.document_reference,i.created_by,i.approved_by FROM supplier_receipt_invoices i JOIN tank_transactions x ON x.id=i.tank_transaction_id JOIN suppliers s ON s.id=x.supplier_id JOIN products p ON p.id=x.product_id ORDER BY i.id DESC""",conn)

def _xlsx(data,company):
 wb=Workbook(); ws=wb.active; ws.title="Receipt Costing"; ws.append([f"{company} | Receipt Invoice Match & Landed Cost"]); ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=max(len(data.columns),1)); ws["A1"].fill=PatternFill("solid",fgColor="9E1B1B"); ws["A1"].font=Font(color="FFFFFF",bold=True,size=16); ws.append(list(data.columns))
 for row in data.where(pd.notna(data),None).itertuples(index=False,name=None): ws.append([str(v) if isinstance(v,pd.Timestamp) else v for v in row])
 for c in ws[2]: c.fill=PatternFill("solid",fgColor="111827"); c.font=Font(color="FFFFFF",bold=True)
 ws.freeze_panes="A3"; ws.auto_filter.ref=ws.dimensions
 for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(max(14,max(len(str(x.value or '')) for x in col)+2),30)
 out=BytesIO(); wb.save(out); return out.getvalue()

def render_receipt_costing(conn):
 role=st.session_state.get("role","VIEWER"); user=st.session_state.get("user","System"); receipts=_receipts(conn); invoices=_invoices(conn); currency=st.session_state.get("company_profile",{}).get("currency","AED")
 page_header("Receipt Costing & Invoice Match","Match supplier invoices to accepted receipts and calculate auditable landed inventory cost per litre.")
 a,b,c,d=st.columns(4); a.metric("Unmatched receipts",len(receipts)); b.metric("Invoice exceptions",int(invoices.status.eq('EXCEPTION').sum()) if not invoices.empty else 0); c.metric("Awaiting approval",int(invoices.status.eq('MATCHED').sum()) if not invoices.empty else 0); d.metric("Approved landed value",f"{currency} {invoices.loc[invoices.status.eq('APPROVED'),'landed_cost_total'].sum():,.2f}" if not invoices.empty else f"{currency} 0.00")
 match,register,approve,report=st.tabs(["Match invoice","Cost register","Approve cost","Cost report"])
 with match:
  if receipts.empty: st.success("Every supplier receipt has an invoice match record.")
  elif role not in EDIT: st.dataframe(receipts,use_container_width=True,hide_index=True)
  else:
   rm={f"STX-{int(r.id)} · {r.supplier} · {float(r.accepted_liters):,.2f} L":r for r in receipts.itertuples()}; selected=st.selectbox("Supplier receipt",list(rm)); row=rm[selected]; st.info(f"Accepted {float(row.accepted_liters):,.2f} L · expected price {currency} {float(row.expected_unit_price):,.4f}/L")
   with st.form("invoice_match"):
    x,y=st.columns(2); number=x.text_input("Supplier invoice number"); invoice_date=y.date_input("Invoice date"); x,y=st.columns(2); qty=x.number_input("Invoiced litres",min_value=0.01,value=float(row.accepted_liters)); price=y.number_input("Invoice unit price",min_value=0.0,value=float(row.expected_unit_price),format="%.4f"); x,y,z,w=st.columns(4); freight=x.number_input("Freight",min_value=0.0); testing=y.number_input("Testing",min_value=0.0); handling=z.number_input("Handling",min_value=0.0); other=w.number_input("Other landed cost",min_value=0.0); tax=st.number_input("Tax amount (reported separately)",min_value=0.0); document=st.text_input("Invoice / evidence document reference"); explanation=st.text_area("Exception explanation"); submit=st.form_submit_button("Register invoice match",type="primary")
   if submit:
    qv=qty-float(row.accepted_liters); pv=price-float(row.expected_unit_price); base=qty*price; landed=base+freight+testing+handling+other; unit=landed/float(row.accepted_liters) if float(row.accepted_liters)>0 else 0; exception=abs(qv)>.005 or abs(pv)>.0001; status="EXCEPTION" if exception else "MATCHED"
    if len(number.strip())<2 or (exception and len(explanation.strip())<5): st.error("Enter the invoice number and explain any quantity or price exception.")
    else:
     c=conn.cursor(); c.execute("""INSERT INTO supplier_receipt_invoices(tank_transaction_id,supplier_invoice_number,invoice_date,invoiced_liters,invoiced_unit_price,freight_cost,testing_cost,handling_cost,other_cost,tax_amount,invoice_total,quantity_variance,price_variance,landed_cost_total,landed_unit_cost,status,exception_reason,document_reference,created_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",(int(row.id),number.strip(),invoice_date,qty,price,freight,testing,handling,other,tax,base+tax,qv,pv,landed,unit,status,explanation.strip() or None,document.strip() or None,user)); iid=c.fetchone()[0]; conn.commit(); record_event(conn,"MATCH_SUPPLIER_INVOICE","Receipt Costing","Supplier Receipt Invoice",iid,f"STX-{int(row.id)}; {status}; landed cost {currency} {unit:.4f}/L",severity="WARNING" if exception else "INFO"); st.success(f"INV-{iid} registered as {status}."); st.rerun()
 with register: st.dataframe(invoices,use_container_width=True,hide_index=True,height=470)
 with approve:
  pending=invoices[invoices.status.isin(["MATCHED","EXCEPTION"])] if not invoices.empty else invoices
  if pending.empty: st.success("No landed cost record is waiting for approval.")
  elif role not in APPROVE: st.info("Only an authorized approver can apply landed cost to valuation.")
  else:
   im={f"INV-{int(r.id)} · STX-{int(r.receipt_id)} · {r.supplier} · {currency} {float(r.landed_unit_cost):.4f}/L":r for r in pending.itertuples()}; selected=st.selectbox("Invoice match",list(im)); row=im[selected]; decision=st.selectbox("Decision",["APPROVED","REJECTED"]); comment=st.text_area("Approval comment / exception authorization")
   if st.button("Save costing decision",type="primary"):
    if len(comment.strip())<5: st.error("Enter an approval comment.")
    else:
     c=conn.cursor(); c.execute("UPDATE supplier_receipt_invoices SET status=%s,approved_by=%s,approved_at=CURRENT_TIMESTAMP,exception_reason=CONCAT(COALESCE(exception_reason,''),%s) WHERE id=%s",(decision,user," | Decision: "+comment.strip(),int(row.id)))
     if decision=="APPROVED": c.execute("UPDATE tank_transactions SET landed_unit_cost=%s WHERE id=%s",(float(row.landed_unit_cost),int(row.receipt_id)))
     conn.commit(); record_event(conn,"APPROVE_LANDED_COST" if decision=="APPROVED" else "REJECT_LANDED_COST","Receipt Costing","Supplier Receipt Invoice",int(row.id),comment.strip()); st.success("Decision saved. Inventory quantity was not changed."); st.rerun()
 with report:
  company=st.session_state.get("company_profile",{}).get("company_name","Company"); st.download_button("Download receipt costing report",_xlsx(invoices,company),"receipt_invoice_landed_cost.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
